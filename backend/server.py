from fastapi import FastAPI, APIRouter, HTTPException, Depends, BackgroundTasks, UploadFile, File, Query, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.background import BackgroundTask
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, logging, uuid, hmac, hashlib, smtplib, io, asyncio, base64
import httpx
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from jose import jwt, JWTError
from passlib.context import CryptContext
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
import razorpay
import qrcode
import aiofiles

try:
    import razorpay
    RAZORPAY_AVAILABLE = True
except ImportError:
    RAZORPAY_AVAILABLE = False

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
db_name = os.environ.get('DB_NAME', 'idsea_db')
db_client = AsyncIOMotorClient(mongo_url)
db = db_client[db_name]

app = FastAPI(title="IDSEA API")
api_router = APIRouter(prefix="/api")

SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'idsea-secret-key-2024')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

RAZORPAY_KEY_ID = os.environ.get('RAZORPAY_KEY_ID', '')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET', '')

async def get_effective_razorpay_keys():
    """Get Razorpay keys from DB first, fallback to env vars"""
    settings = await db.payment_settings.find_one({}, {"_id": 0})
    if settings:
        db_key = settings.get("razorpay_key_id", "")
        db_secret = settings.get("razorpay_key_secret", "")
        if db_key and db_secret:
            return db_key, db_secret
    return RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET
SMTP_HOST = os.environ.get('SMTP_HOST', '')
SMTP_PORT_VAL = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASS = os.environ.get('SMTP_PASS', '')
FROM_EMAIL = os.environ.get('FROM_EMAIL', 'noreply@idsea.org')

UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)


def now_iso():
    return datetime.now(timezone.utc).isoformat()


# =================== MODELS ===================

class AdminUser(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    password_hash: str
    role: str = "admin"
    created_at: str = Field(default_factory=now_iso)


class AdminLogin(BaseModel):
    email: str
    password: str


class Member(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    prefix: Optional[str] = ""
    name: str
    qualification: Optional[str] = ""
    specialization: Optional[str] = ""
    organization: Optional[str] = ""
    permanent_address: Optional[dict] = Field(default_factory=lambda: {"line1": "", "line2": "", "line3": "", "state": "", "district": "", "pincode": ""})
    contact_address: Optional[dict] = Field(default_factory=lambda: {"line1": "", "line2": "", "line3": "", "state": "", "district": "", "pincode": ""})
    contact_same_as_permanent: Optional[bool] = False
    address: Optional[str] = ""
    state: Optional[str] = ""
    country: Optional[str] = "India"
    phone: Optional[str] = ""
    email: str
    photo_url: Optional[str] = ""
    identity_proof_url: Optional[str] = ""
    membership_type: str
    membership_id: Optional[str] = ""
    join_date: Optional[str] = ""
    payment_status: str = "pending"
    status: str = "pending"
    amount_paid: Optional[float] = 0
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


class MemberCreate(BaseModel):
    prefix: Optional[str] = ""
    name: str
    qualification: Optional[str] = ""
    specialization: Optional[str] = ""
    organization: Optional[str] = ""
    permanent_address: Optional[dict] = Field(default_factory=lambda: {"line1": "", "line2": "", "line3": "", "state": "", "district": "", "pincode": ""})
    contact_address: Optional[dict] = Field(default_factory=lambda: {"line1": "", "line2": "", "line3": "", "state": "", "district": "", "pincode": ""})
    contact_same_as_permanent: Optional[bool] = False
    address: Optional[str] = ""
    state: Optional[str] = ""
    country: Optional[str] = "India"
    phone: Optional[str] = ""
    email: str
    photo_url: Optional[str] = ""
    identity_proof_url: Optional[str] = ""
    membership_type: str
    payment_status: str = "pending"


class MemberUpdate(BaseModel):
    prefix: Optional[str] = None
    name: Optional[str] = None
    qualification: Optional[str] = None
    specialization: Optional[str] = None
    organization: Optional[str] = None
    permanent_address: Optional[dict] = None
    contact_address: Optional[dict] = None
    contact_same_as_permanent: Optional[bool] = None
    address: Optional[str] = None
    state: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    photo_url: Optional[str] = None
    membership_type: Optional[str] = None
    payment_status: Optional[str] = None
    status: Optional[str] = None


class FeeTier(BaseModel):
    name: str = ""  # "early_bird" or "regular"
    deadline: str = ""
    fees: dict = {}  # {member: 0, non_member: 0, student: 0, international: 0}
    accommodation_fees: dict = {}  # {member: 0, non_member: 0, student: 0, international: 0}


class HotelOption(BaseModel):
    name: str = ""
    fee: float = 0
    description: Optional[str] = ""


class AccommodationConfig(BaseModel):
    enabled: bool = False
    self_option: bool = True
    free_categories: List[str] = []
    hotels: List[dict] = []


class Event(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    date: str
    end_date: Optional[str] = ""
    venue: str
    venue_map_link: Optional[str] = ""
    description: Optional[str] = ""
    registration_fee: Optional[float] = 0
    brochure_url: Optional[str] = ""
    status: str = "upcoming"
    image_url: Optional[str] = ""
    is_visible: bool = True
    registration_enabled: bool = False
    allow_membership_registration: bool = False
    fee_tiers: List[dict] = []  # [{name, deadline, fees: {member, non_member, student, international}}]
    accommodation: dict = {}
    additional_person_fee: float = 0
    additional_person_fee_usd: float = 0
    registration_addons: List[dict] = []
    premium_hotels: List[dict] = []
    created_at: str = Field(default_factory=now_iso)


class EventCreate(BaseModel):
    title: str
    date: str
    end_date: Optional[str] = ""
    venue: str
    venue_map_link: Optional[str] = ""
    description: Optional[str] = ""
    registration_fee: Optional[float] = 0
    brochure_url: Optional[str] = ""
    status: str = "upcoming"
    image_url: Optional[str] = ""
    is_visible: bool = True
    registration_enabled: bool = False
    allow_membership_registration: bool = False
    fee_tiers: List[dict] = []
    accommodation: dict = {}
    additional_person_fee: float = 0
    additional_person_fee_usd: float = 0
    registration_addons: List[dict] = []
    premium_hotels: List[dict] = []


class EventRegistration(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    is_member: bool = False
    member_id: Optional[str] = ""
    member_category: Optional[str] = ""
    registration_category: str = ""  # member, non_member, student, international
    name: str
    email: str
    phone: str
    qualification: Optional[str] = ""
    organization: Optional[str] = ""
    state: Optional[str] = ""
    # Address fields (non-member, international)
    address_line1: Optional[str] = ""
    address_line2: Optional[str] = ""
    district: Optional[str] = ""
    address_state: Optional[str] = ""
    country: Optional[str] = "India"
    pincode: Optional[str] = ""
    # Student fields
    college: Optional[str] = ""
    university: Optional[str] = ""
    # File uploads (PDFs)
    identity_proof_url: Optional[str] = ""
    bonafide_cert_url: Optional[str] = ""
    student_proof_url: Optional[str] = ""
    # Accommodation
    accommodation_choice: Optional[str] = ""  # none, default, self_hotel
    hotel_name: Optional[str] = ""
    hotel_room_type: Optional[str] = ""
    hotel_tax_percent: float = 0
    hotel_base_amount: float = 0
    # Additional persons for accommodation
    additional_persons: List[dict] = []  # [{name, age, mobile}]
    additional_persons_fee: float = 0
    # Membership
    wants_membership: bool = False
    membership_type: Optional[str] = ""
    # Add-ons
    selected_addons: List[str] = []
    addon_fee: float = 0
    # Fees
    registration_fee: float = 0
    accommodation_fee: float = 0
    hotel_tax_amount: float = 0
    membership_fee: float = 0
    total_amount: float = 0
    payment_status: str = "pending"
    payment_mode: str = "offline"
    # Admin fields
    assigned_room_no: Optional[str] = ""
    assigned_location: Optional[str] = ""
    assigned_location_type: Optional[str] = ""
    assigned_map_link: Optional[str] = ""
    accommodation_notified: bool = False
    created_at: str = Field(default_factory=now_iso)


class News(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    content: str
    category: str = "general"
    image_url: Optional[str] = ""
    published_date: str
    status: str = "published"
    created_at: str = Field(default_factory=now_iso)


class NewsCreate(BaseModel):
    title: str
    content: str
    category: str = "general"
    image_url: Optional[str] = ""
    published_date: str
    status: str = "published"


class GalleryAlbum(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = ""
    cover_image: Optional[str] = ""
    category: str = "conference"
    created_at: str = Field(default_factory=now_iso)


class GalleryAlbumCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    cover_image: Optional[str] = ""
    category: str = "conference"


class GalleryPhoto(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    album_id: str
    title: Optional[str] = ""
    image_url: str
    created_at: str = Field(default_factory=now_iso)


class GalleryPhotoCreate(BaseModel):
    album_id: str
    title: Optional[str] = ""
    image_url: str


class Publication(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    author: str
    abstract: Optional[str] = ""
    pdf_url: Optional[str] = ""
    category: str = "research"
    published_date: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class PublicationCreate(BaseModel):
    title: str
    author: str
    abstract: Optional[str] = ""
    pdf_url: Optional[str] = ""
    category: str = "research"
    published_date: Optional[str] = ""


class ExecutiveCommittee(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    member_id: Optional[str] = ""
    name: str
    designation: str
    sub_division: Optional[str] = ""
    frontend_section: Optional[str] = "office_bearers"
    affiliation: Optional[str] = ""
    profile: Optional[str] = ""
    photo_url: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    category: str = "council"
    order: int = 0
    created_at: str = Field(default_factory=now_iso)


class ExecutiveCreate(BaseModel):
    member_id: Optional[str] = ""
    name: str
    designation: str
    sub_division: Optional[str] = ""
    frontend_section: Optional[str] = "office_bearers"
    affiliation: Optional[str] = ""
    profile: Optional[str] = ""
    photo_url: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    category: str = "council"
    order: int = 0


class Payment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    member_id: Optional[str] = ""
    member_name: Optional[str] = ""
    member_email: Optional[str] = ""
    membership_type: Optional[str] = ""
    event_registration_id: Optional[str] = ""
    purpose: str = "membership"  # membership, event_registration
    amount: float
    currency: str = "INR"
    payment_method: str = "razorpay"  # razorpay, upi, bank_transfer, manual
    razorpay_order_id: Optional[str] = ""
    razorpay_payment_id: Optional[str] = ""
    utr_number: Optional[str] = ""
    status: str = "pending"  # pending, verification_pending, paid, rejected, refunded
    refund_status: Optional[str] = ""  # "", initiated, completed
    refund_date: Optional[str] = ""
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class PaymentOrder(BaseModel):
    amount: float
    currency: str = "INR"
    member_id: Optional[str] = ""
    member_name: Optional[str] = ""
    member_email: Optional[str] = ""
    membership_type: Optional[str] = ""
    event_registration_id: Optional[str] = ""
    purpose: str = "membership"


class PaymentVerify(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    payment_id: str


class EmailRequest(BaseModel):
    recipients: List[str] = []
    subject: str
    body: str
    recipient_group: Optional[str] = "custom"


class CertificateRequest(BaseModel):
    member_id: str
    certificate_type: str
    event_name: Optional[str] = ""
    issue_date: Optional[str] = ""


class Slider(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: Optional[str] = ""
    subtitle: Optional[str] = ""
    image_url: str
    link_url: Optional[str] = ""
    order: int = 0
    is_active: bool = True
    created_at: str = Field(default_factory=now_iso)


class SliderCreate(BaseModel):
    title: Optional[str] = ""
    subtitle: Optional[str] = ""
    image_url: str
    link_url: Optional[str] = ""
    order: int = 0
    is_active: bool = True


class CMSSettings(BaseModel):
    about_content: Optional[str] = ""
    vision: Optional[str] = ""
    mission: Optional[str] = ""
    contact_email: Optional[str] = ""
    contact_phone: Optional[str] = ""
    contact_address: Optional[str] = ""
    facebook_url: Optional[str] = ""
    twitter_url: Optional[str] = ""
    linkedin_url: Optional[str] = ""
    website_name: Optional[str] = "IDSEA"
    logo_url: Optional[str] = ""
    hero_title: Optional[str] = ""
    hero_subtitle: Optional[str] = ""
    favicon_url: Optional[str] = ""
    custom_head_scripts: Optional[str] = ""
    custom_body_start_scripts: Optional[str] = ""
    custom_body_end_scripts: Optional[str] = ""
    site_url: Optional[str] = ""


# =================== AUTH UTILS ===================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


async def get_current_admin(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        admin_id = payload.get("sub")
        if not admin_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        admin = await db.admins.find_one({"id": admin_id}, {"_id": 0})
        if not admin:
            raise HTTPException(status_code=401, detail="Admin not found")
        return admin
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


def get_razorpay_client():
    """Sync version - tries env keys only. Use async version for DB keys."""
    if RAZORPAY_AVAILABLE and RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET:
        return razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
    return None

async def get_razorpay_client_async():
    """Async version - checks DB keys first, then env vars"""
    key_id, key_secret = await get_effective_razorpay_keys()
    if RAZORPAY_AVAILABLE and key_id and key_secret:
        return razorpay.Client(auth=(key_id, key_secret)), key_id
    return None, ""


def send_email_smtp(recipients: List[str], subject: str, body: str, smtp_settings: dict = None, attachments: list = None):
    host = smtp_settings.get("smtp_host", SMTP_HOST) if smtp_settings else SMTP_HOST
    port = int(smtp_settings.get("smtp_port", SMTP_PORT_VAL) if smtp_settings else SMTP_PORT_VAL)
    user = smtp_settings.get("smtp_user", SMTP_USER) if smtp_settings else SMTP_USER
    passwd = smtp_settings.get("smtp_pass", SMTP_PASS) if smtp_settings else SMTP_PASS
    from_email = smtp_settings.get("from_email", FROM_EMAIL) if smtp_settings else FROM_EMAIL
    if not host or not user:
        logging.warning("SMTP not configured, email skipped")
        return False
    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = f"IDSEA <{from_email}>"
        msg['To'] = ', '.join(recipients[:50])
        msg.attach(MIMEText(body, 'html'))
        if attachments:
            for att in attachments:
                part = MIMEApplication(att["data"], Name=att["filename"])
                part['Content-Disposition'] = f'attachment; filename="{att["filename"]}"'
                msg.attach(part)
        with smtplib.SMTP(host, port) as server:
            server.starttls()
            server.login(user, passwd)
            server.sendmail(from_email, recipients, msg.as_string())
        return True
    except Exception as e:
        logging.error(f"Email error: {e}")
        return False


# =================== EMAIL TEMPLATE ENGINE ===================

DEFAULT_EMAIL_TEMPLATES = {
    "registration_submitted": {
        "name": "Registration Submitted",
        "subject": "IDSEA - Membership Application Received",
        "description": "Sent when a new member submits their application",
        "variables": ["member_name", "email", "phone", "qualification", "organization", "membership_type", "state", "payment_status", "application_date"],
        "body": """<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff;">
  <div style="background: #0c3c60; padding: 32px 24px; text-align: center;">
    <h1 style="color: white; margin: 0; font-size: 22px; font-weight: 700; letter-spacing: 0.5px;">IDSEA</h1>
    <p style="color: rgba(255,255,255,0.8); margin: 6px 0 0; font-size: 12px;">Indian Dairy Scientists and Entrepreneurs Association</p>
  </div>
  <div style="padding: 36px 28px;">
    <h2 style="color: #0c3c60; font-size: 20px; margin: 0 0 12px;">Application Received!</h2>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Dear <strong>{{member_name}}</strong>,</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Thank you for submitting your membership application to IDSEA. We have received your details and our team will review your application shortly.</p>
    <div style="background: #f0f9ff; border-radius: 10px; padding: 20px; margin: 24px 0; border-left: 4px solid #0c3c60;">
      <h3 style="color: #0c3c60; margin: 0 0 14px; font-size: 15px;">Application Summary</h3>
      <table style="width: 100%; font-size: 13px; color: #374151;">
        <tr><td style="padding: 6px 0; color: #6b7280; width: 40%;">Full Name</td><td style="padding: 6px 0; font-weight: 600;">{{member_name}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Email</td><td style="padding: 6px 0;">{{email}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Phone</td><td style="padding: 6px 0;">{{phone}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Qualification</td><td style="padding: 6px 0;">{{qualification}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Organization</td><td style="padding: 6px 0;">{{organization}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Membership Type</td><td style="padding: 6px 0;"><span style="background: #dbeafe; color: #1e40af; padding: 2px 10px; border-radius: 10px; font-size: 12px; font-weight: 600;">{{membership_type}}</span></td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">State</td><td style="padding: 6px 0;">{{state}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Payment Status</td><td style="padding: 6px 0;"><span style="background: #fef3c7; color: #92400e; padding: 2px 10px; border-radius: 10px; font-size: 12px; font-weight: 600;">{{payment_status}}</span></td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Application Date</td><td style="padding: 6px 0;">{{application_date}}</td></tr>
      </table>
    </div>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Our membership committee will review your application and notify you once it is approved. If you have any questions, please contact us.</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Warm regards,<br><strong style="color: #0c3c60;">IDSEA Team</strong></p>
  </div>
  <div style="background: #f8fafc; padding: 20px 28px; text-align: center; border-top: 1px solid #e5e7eb;">
    <p style="color: #9ca3af; font-size: 11px; margin: 0;">Indian Dairy Scientists and Entrepreneurs Association (IDSEA)</p>
    <p style="color: #9ca3af; font-size: 11px; margin: 4px 0 0;">This is an automated email. Please do not reply directly.</p>
  </div>
</div>"""
    },
    "membership_approved": {
        "name": "Membership Approved",
        "subject": "Welcome to IDSEA! - Membership Approved",
        "description": "Sent when membership is approved (with certificate attached)",
        "variables": ["member_name", "email", "membership_id", "membership_type", "qualification", "organization", "state", "join_date"],
        "body": """<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff;">
  <div style="background: #0c3c60; padding: 32px 24px; text-align: center;">
    <h1 style="color: white; margin: 0; font-size: 22px; font-weight: 700; letter-spacing: 0.5px;">IDSEA</h1>
    <p style="color: rgba(255,255,255,0.8); margin: 6px 0 0; font-size: 12px;">Indian Dairy Scientists and Entrepreneurs Association</p>
  </div>
  <div style="padding: 36px 28px;">
    <div style="text-align: center; margin-bottom: 24px;">
      <div style="width: 64px; height: 64px; background: #d1fae5; border-radius: 50%; display: inline-flex; align-items: center; justify-content: center; font-size: 28px;">&#10003;</div>
    </div>
    <h2 style="color: #1e7a4d; font-size: 22px; margin: 0 0 12px; text-align: center;">Welcome to IDSEA!</h2>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7; text-align: center;">Dear <strong>{{member_name}}</strong>, your membership has been approved.</p>
    <div style="background: #f0fdf4; border-radius: 10px; padding: 24px; margin: 24px 0; border: 1px solid #bbf7d0; text-align: center;">
      <p style="color: #6b7280; font-size: 12px; margin: 0 0 6px;">Your Membership ID</p>
      <p style="color: #0c3c60; font-size: 28px; font-weight: 800; margin: 0; letter-spacing: 2px;">{{membership_id}}</p>
    </div>
    <div style="background: #f0f9ff; border-radius: 10px; padding: 20px; margin: 24px 0;">
      <h3 style="color: #0c3c60; margin: 0 0 14px; font-size: 15px;">Membership Details</h3>
      <table style="width: 100%; font-size: 13px; color: #374151;">
        <tr><td style="padding: 6px 0; color: #6b7280; width: 40%;">Full Name</td><td style="padding: 6px 0; font-weight: 600;">{{member_name}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Email</td><td style="padding: 6px 0;">{{email}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Membership Type</td><td style="padding: 6px 0;"><span style="background: #d1fae5; color: #065f46; padding: 2px 10px; border-radius: 10px; font-size: 12px; font-weight: 600;">{{membership_type}}</span></td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Qualification</td><td style="padding: 6px 0;">{{qualification}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Organization</td><td style="padding: 6px 0;">{{organization}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">State</td><td style="padding: 6px 0;">{{state}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Member Since</td><td style="padding: 6px 0;">{{join_date}}</td></tr>
      </table>
    </div>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Your membership certificate is attached to this email. Please keep it for your records.</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">We look forward to your active participation in IDSEA activities.</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Warm regards,<br><strong style="color: #0c3c60;">IDSEA Team</strong></p>
  </div>
  <div style="background: #f8fafc; padding: 20px 28px; text-align: center; border-top: 1px solid #e5e7eb;">
    <p style="color: #9ca3af; font-size: 11px; margin: 0;">Indian Dairy Scientists and Entrepreneurs Association (IDSEA)</p>
    <p style="color: #9ca3af; font-size: 11px; margin: 4px 0 0;">This is an automated email. Please do not reply directly.</p>
  </div>
</div>"""
    },
    "event_notification": {
        "name": "New Event Notification",
        "subject": "IDSEA - New Event: {{event_title}}",
        "description": "Sent to approved members when a new event is created (batch: 50 per 30 mins)",
        "variables": ["member_name", "event_title", "event_date", "event_end_date", "event_venue", "venue_map_link", "event_description", "registration_fee"],
        "body": """<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff;">
  <div style="background: #0c3c60; padding: 32px 24px; text-align: center;">
    <h1 style="color: white; margin: 0; font-size: 22px; font-weight: 700; letter-spacing: 0.5px;">IDSEA</h1>
    <p style="color: rgba(255,255,255,0.8); margin: 6px 0 0; font-size: 12px;">Indian Dairy Scientists and Entrepreneurs Association</p>
  </div>
  <div style="padding: 36px 28px;">
    <h2 style="color: #0c3c60; font-size: 20px; margin: 0 0 8px;">New Event Announcement</h2>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Dear <strong>{{member_name}}</strong>,</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">We are excited to announce a new event:</p>
    <div style="background: #f0f9ff; border-radius: 10px; padding: 24px; margin: 20px 0; border-left: 4px solid #0c3c60;">
      <h3 style="color: #0c3c60; margin: 0 0 16px; font-size: 18px;">{{event_title}}</h3>
      <table style="width: 100%; font-size: 13px; color: #374151;">
        <tr><td style="padding: 6px 0; color: #6b7280; width: 35%;">Date</td><td style="padding: 6px 0; font-weight: 600;">{{event_date}} - {{event_end_date}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Venue</td><td style="padding: 6px 0;">{{event_venue}} {{venue_map_link}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Registration Fee</td><td style="padding: 6px 0; font-weight: 600;">₹{{registration_fee}}</td></tr>
      </table>
    </div>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">{{event_description}}</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">We encourage all members to participate. For registration and queries, please contact us.</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Warm regards,<br><strong style="color: #0c3c60;">IDSEA Team</strong></p>
  </div>
  <div style="background: #f8fafc; padding: 20px 28px; text-align: center; border-top: 1px solid #e5e7eb;">
    <p style="color: #9ca3af; font-size: 11px; margin: 0;">Indian Dairy Scientists and Entrepreneurs Association (IDSEA)</p>
    <p style="color: #9ca3af; font-size: 11px; margin: 4px 0 0;">This is an automated email. Please do not reply directly.</p>
  </div>
</div>"""
    },
    "event_participation": {
        "name": "Event Participation Certificate",
        "subject": "IDSEA - Thank You for Participating in {{event_title}}",
        "description": "Sent to participants when event is closed (with certificate attached)",
        "variables": ["member_name", "membership_id", "event_title", "event_date", "event_venue"],
        "body": """<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff;">
  <div style="background: #0c3c60; padding: 32px 24px; text-align: center;">
    <h1 style="color: white; margin: 0; font-size: 22px; font-weight: 700; letter-spacing: 0.5px;">IDSEA</h1>
    <p style="color: rgba(255,255,255,0.8); margin: 6px 0 0; font-size: 12px;">Indian Dairy Scientists and Entrepreneurs Association</p>
  </div>
  <div style="padding: 36px 28px;">
    <h2 style="color: #0c3c60; font-size: 20px; margin: 0 0 12px;">Thank You for Your Participation!</h2>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Dear <strong>{{member_name}}</strong>,</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Thank you for participating in <strong>{{event_title}}</strong> held at {{event_venue}} on {{event_date}}.</p>
    <div style="background: #f0fdf4; border-radius: 10px; padding: 24px; margin: 24px 0; text-align: center; border: 1px solid #bbf7d0;">
      <p style="color: #1e7a4d; font-size: 16px; font-weight: 700; margin: 0;">Your participation certificate is attached!</p>
      <p style="color: #6b7280; font-size: 12px; margin: 8px 0 0;">Member ID: {{membership_id}}</p>
    </div>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">We hope you found the event valuable. Please keep the attached certificate for your records.</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">We look forward to seeing you at future IDSEA events!</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Warm regards,<br><strong style="color: #0c3c60;">IDSEA Team</strong></p>
  </div>
  <div style="background: #f8fafc; padding: 20px 28px; text-align: center; border-top: 1px solid #e5e7eb;">
    <p style="color: #9ca3af; font-size: 11px; margin: 0;">Indian Dairy Scientists and Entrepreneurs Association (IDSEA)</p>
    <p style="color: #9ca3af; font-size: 11px; margin: 4px 0 0;">This is an automated email. Please do not reply directly.</p>
  </div>
</div>"""
    },
    "membership_rejected": {
        "name": "Membership Rejected",
        "subject": "IDSEA - Membership Application Update",
        "description": "Sent when membership application is rejected",
        "variables": ["member_name", "email", "membership_type", "rejection_reason"],
        "body": """<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff;">
  <div style="background: #0c3c60; padding: 32px 24px; text-align: center;">
    <h1 style="color: white; margin: 0; font-size: 22px; font-weight: 700;">IDSEA</h1>
    <p style="color: rgba(255,255,255,0.8); margin: 6px 0 0; font-size: 12px;">Indian Dairy Scientists and Entrepreneurs Association</p>
  </div>
  <div style="padding: 36px 28px;">
    <h2 style="color: #dc2626; font-size: 20px; margin: 0 0 12px;">Application Update</h2>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Dear <strong>{{member_name}}</strong>,</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">We regret to inform you that your membership application for <strong>{{membership_type}}</strong> category could not be approved at this time.</p>
    <div style="background: #fef2f2; border-radius: 10px; padding: 20px; margin: 24px 0; border-left: 4px solid #dc2626;">
      <p style="color: #991b1b; font-size: 13px; margin: 0;"><strong>Reason:</strong> {{rejection_reason}}</p>
    </div>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">You may reapply after addressing the above concerns. For queries, please contact us.</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Regards,<br><strong style="color: #0c3c60;">IDSEA Team</strong></p>
  </div>
  <div style="background: #f8fafc; padding: 20px 28px; text-align: center; border-top: 1px solid #e5e7eb;">
    <p style="color: #9ca3af; font-size: 11px; margin: 0;">Indian Dairy Scientists and Entrepreneurs Association (IDSEA)</p>
  </div>
</div>"""
    },
    "event_registration_confirmed": {
        "name": "Event Registration Confirmed",
        "subject": "IDSEA - Registration Confirmed for {{event_title}}",
        "description": "Sent when someone successfully registers for an event",
        "variables": ["participant_name", "participant_email", "event_title", "event_date", "event_venue", "registration_id", "registration_type", "amount_paid"],
        "body": """<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff;">
  <div style="background: #0c3c60; padding: 32px 24px; text-align: center;">
    <h1 style="color: white; margin: 0; font-size: 22px; font-weight: 700;">IDSEA</h1>
    <p style="color: rgba(255,255,255,0.8); margin: 6px 0 0; font-size: 12px;">Indian Dairy Scientists and Entrepreneurs Association</p>
  </div>
  <div style="padding: 36px 28px;">
    <div style="text-align: center; margin-bottom: 24px;">
      <div style="width: 64px; height: 64px; background: #d1fae5; border-radius: 50%; display: inline-flex; align-items: center; justify-content: center; font-size: 28px;">&#10003;</div>
    </div>
    <h2 style="color: #1e7a4d; font-size: 20px; margin: 0 0 12px; text-align: center;">Registration Confirmed!</h2>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Dear <strong>{{participant_name}}</strong>,</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Your registration for the following event has been confirmed:</p>
    <div style="background: #f0f9ff; border-radius: 10px; padding: 20px; margin: 20px 0; border-left: 4px solid #0c3c60;">
      <h3 style="color: #0c3c60; margin: 0 0 14px; font-size: 17px;">{{event_title}}</h3>
      <table style="width: 100%; font-size: 13px; color: #374151;">
        <tr><td style="padding: 6px 0; color: #6b7280; width: 40%;">Registration ID</td><td style="padding: 6px 0; font-weight: 600;">{{registration_id}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Date</td><td style="padding: 6px 0;">{{event_date}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Venue</td><td style="padding: 6px 0;">{{event_venue}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Category</td><td style="padding: 6px 0;">{{registration_type}}</td></tr>
        <tr><td style="padding: 6px 0; color: #6b7280;">Amount Paid</td><td style="padding: 6px 0; font-weight: 600; color: #1e7a4d;">{{amount_paid}}</td></tr>
      </table>
    </div>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Please save this email for your records. We look forward to seeing you!</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Warm regards,<br><strong style="color: #0c3c60;">IDSEA Team</strong></p>
  </div>
  <div style="background: #f8fafc; padding: 20px 28px; text-align: center; border-top: 1px solid #e5e7eb;">
    <p style="color: #9ca3af; font-size: 11px; margin: 0;">Indian Dairy Scientists and Entrepreneurs Association (IDSEA)</p>
  </div>
</div>"""
    },
    "certificate_issued": {
        "name": "Certificate Issued",
        "subject": "IDSEA - Your Certificate is Ready",
        "description": "Sent when a certificate is generated and available for download",
        "variables": ["member_name", "certificate_type", "event_title", "certificate_id", "verify_url"],
        "body": """<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff;">
  <div style="background: #0c3c60; padding: 32px 24px; text-align: center;">
    <h1 style="color: white; margin: 0; font-size: 22px; font-weight: 700;">IDSEA</h1>
    <p style="color: rgba(255,255,255,0.8); margin: 6px 0 0; font-size: 12px;">Indian Dairy Scientists and Entrepreneurs Association</p>
  </div>
  <div style="padding: 36px 28px;">
    <h2 style="color: #0c3c60; font-size: 20px; margin: 0 0 12px;">Certificate Issued</h2>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Dear <strong>{{member_name}}</strong>,</p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Your <strong>{{certificate_type}}</strong> certificate has been issued.</p>
    <div style="background: #f0fdf4; border-radius: 10px; padding: 24px; margin: 24px 0; text-align: center; border: 1px solid #bbf7d0;">
      <p style="color: #6b7280; font-size: 12px; margin: 0 0 6px;">Certificate ID</p>
      <p style="color: #0c3c60; font-size: 22px; font-weight: 800; margin: 0; letter-spacing: 1px;">{{certificate_id}}</p>
      <p style="color: #6b7280; font-size: 12px; margin: 10px 0 0;">{{event_title}}</p>
    </div>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">You can verify your certificate online at: <a href="{{verify_url}}" style="color: #2563eb;">{{verify_url}}</a></p>
    <p style="color: #4b5563; font-size: 14px; line-height: 1.7;">Warm regards,<br><strong style="color: #0c3c60;">IDSEA Team</strong></p>
  </div>
  <div style="background: #f8fafc; padding: 20px 28px; text-align: center; border-top: 1px solid #e5e7eb;">
    <p style="color: #9ca3af; font-size: 11px; margin: 0;">Indian Dairy Scientists and Entrepreneurs Association (IDSEA)</p>
  </div>
</div>"""
    },
}


def render_template(template_body: str, variables: dict) -> str:
    """Replace {{variable_name}} placeholders with actual values"""
    rendered = template_body
    for key, value in variables.items():
        rendered = rendered.replace("{{" + key + "}}", str(value or ""))
    return rendered


def generate_registration_pdf(member: dict) -> bytes:
    """Generate a membership application form PDF with all filled data."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40, leftMargin=40, rightMargin=40)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#0c3c60'), spaceAfter=4, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('SubStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#6b7280'), spaceAfter=20, alignment=TA_CENTER)
    section_style = ParagraphStyle('SectionStyle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#0c3c60'), spaceBefore=16, spaceAfter=8, borderPadding=(0, 0, 4, 0))
    elements = []

    # Header
    elements.append(Paragraph("IDSEA Membership Application", title_style))
    elements.append(Paragraph("Indian Dairy Scientists and Entrepreneurs Association", subtitle_style))
    elements.append(Spacer(1, 4))

    # Personal Information
    elements.append(Paragraph("Personal Information", section_style))
    full_name = f"{member.get('prefix', '')} {member.get('name', '')}".strip()
    personal_data = [
        ["Full Name", full_name],
        ["Email", member.get("email", "")],
        ["Phone", member.get("phone", "")],
        ["Qualification", member.get("qualification", "")],
        ["Specialization", member.get("specialization", "")],
        ["Organization", member.get("organization", "")],
        ["Country", member.get("country", "India")],
    ]
    t = Table(personal_data, colWidths=[140, 370])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#e5e7eb')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(t)

    # Address
    def format_addr(addr):
        if not addr or not isinstance(addr, dict):
            return ""
        parts = [addr.get("line1", ""), addr.get("line2", ""), addr.get("line3", "")]
        parts = [p for p in parts if p]
        loc = ", ".join([addr.get("district", ""), addr.get("state", ""), addr.get("pincode", "")])
        loc = ", ".join([p for p in loc.split(", ") if p])
        if loc:
            parts.append(loc)
        return ", ".join(parts)

    perm_addr = format_addr(member.get("permanent_address"))
    cont_addr = format_addr(member.get("contact_address"))

    if perm_addr or cont_addr:
        elements.append(Paragraph("Address Details", section_style))
        addr_data = []
        if perm_addr:
            addr_data.append(["Permanent Address", perm_addr])
        if cont_addr and cont_addr != perm_addr:
            addr_data.append(["Contact Address", cont_addr])
        elif member.get("contact_same_as_permanent"):
            addr_data.append(["Contact Address", "Same as Permanent Address"])
        if addr_data:
            t2 = Table(addr_data, colWidths=[140, 370])
            t2.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#e5e7eb')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(t2)

    # Membership Details
    elements.append(Paragraph("Membership Details", section_style))
    mtype_labels = {"academic": "Academic", "entrepreneur": "Entrepreneur", "corporate": "Corporate", "international": "International"}
    mem_data = [
        ["Membership Type", mtype_labels.get(member.get("membership_type", ""), member.get("membership_type", ""))],
        ["Application Date", member.get("join_date", datetime.now().strftime("%Y-%m-%d"))],
        ["Payment Status", member.get("payment_status", "pending").capitalize()],
        ["Application Status", member.get("status", "pending").capitalize()],
    ]
    t3 = Table(mem_data, colWidths=[140, 370])
    t3.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#e5e7eb')),
    ]))
    elements.append(t3)

    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle('FooterStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER)
    elements.append(Paragraph(f"This is a computer-generated application form. Generated on {datetime.now().strftime('%d %B %Y at %I:%M %p')}", footer_style))
    elements.append(Paragraph("Indian Dairy Scientists and Entrepreneurs Association (IDSEA) | www.idsea.in", footer_style))

    doc.build(elements)
    return buf.getvalue()


def generate_certificate_pdf_bytes(member_name: str, membership_id: str, cert_type: str = "membership", event_name: str = "", issue_date: str = "") -> bytes:
    """Generate certificate PDF in memory and return bytes (NOT stored in DB)"""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    w, h = landscape(A4)
    if not issue_date:
        issue_date = datetime.now().strftime("%d %B %Y")

    c.setStrokeColor(colors.HexColor('#0c3c60'))
    c.setLineWidth(4)
    c.rect(30, 30, w - 60, h - 60)
    c.setStrokeColor(colors.HexColor('#1e7a4d'))
    c.setLineWidth(2)
    c.rect(40, 40, w - 80, h - 80)

    c.setFillColor(colors.HexColor('#0c3c60'))
    c.setFont("Helvetica-Bold", 28)
    c.drawCentredString(w / 2, h - 100, "Indian Dairy Scientists and")
    c.drawCentredString(w / 2, h - 135, "Entrepreneurs Association")
    c.setFont("Helvetica", 12)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(w / 2, h - 158, "IDSEA")

    cert_type_label = {"membership": "Certificate of Membership", "event": "Certificate of Participation", "appreciation": "Certificate of Appreciation"}.get(cert_type, "Certificate")
    c.setFillColor(colors.HexColor('#1e7a4d'))
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(w / 2, h - 210, cert_type_label)

    c.setFillColor(colors.HexColor('#374151'))
    c.setFont("Helvetica", 14)
    c.drawCentredString(w / 2, h - 260, "This is to certify that")

    c.setFillColor(colors.HexColor('#0c3c60'))
    c.setFont("Helvetica-Bold", 26)
    c.drawCentredString(w / 2, h - 300, member_name)

    c.setFillColor(colors.HexColor('#6b7280'))
    c.setFont("Helvetica", 12)
    if membership_id:
        c.drawCentredString(w / 2, h - 325, f"Member ID: {membership_id}")

    c.setFillColor(colors.HexColor('#374151'))
    c.setFont("Helvetica", 13)
    if cert_type == "event" and event_name:
        c.drawCentredString(w / 2, h - 358, "has successfully participated in")
        c.setFont("Helvetica-Bold", 15)
        c.setFillColor(colors.HexColor('#1e7a4d'))
        c.drawCentredString(w / 2, h - 380, event_name)
    elif cert_type == "membership":
        c.drawCentredString(w / 2, h - 358, "is a registered member of IDSEA")
    else:
        c.drawCentredString(w / 2, h - 358, "is recognized by IDSEA for their contributions")

    c.setFillColor(colors.HexColor('#6b7280'))
    c.setFont("Helvetica", 11)
    c.drawString(80, 90, f"Date: {issue_date}")
    c.line(w - 280, 100, w - 80, 100)
    c.setFont("Helvetica", 11)
    c.drawCentredString(w - 180, 82, "Authorized Signatory")
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(w - 180, 110, "IDSEA")

    c.save()
    buf.seek(0)
    return buf.getvalue()


async def send_templated_email(template_key: str, recipients: list, variables: dict, attachments: list = None):
    """Send email using template from DB (or default), with optional attachments"""
    template = await db.email_templates.find_one({"key": template_key}, {"_id": 0})
    if not template:
        template = DEFAULT_EMAIL_TEMPLATES.get(template_key)
    if not template:
        logging.error(f"Email template not found: {template_key}")
        return

    subject = render_template(template.get("subject", ""), variables)
    body = render_template(template.get("body", ""), variables)
    smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})

    log_id = str(uuid.uuid4())
    log = {
        "id": log_id,
        "subject": subject,
        "body": body[:300],
        "recipients_count": len(recipients),
        "recipient_group": template_key,
        "sent_by": "system",
        "created_at": now_iso(),
        "sent_at": "",
        "status": "sending"
    }
    await db.email_logs.insert_one(log)

    try:
        success = send_email_smtp(recipients, subject, body, smtp_settings, attachments)
        if success:
            await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "sent", "sent_at": now_iso()}})
            logging.info(f"Templated email [{template_key}] sent to {len(recipients)} recipients")
        else:
            await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "failed", "error": "SMTP send returned False"}})
            logging.warning(f"Templated email [{template_key}] FAILED for {len(recipients)} recipients")
    except Exception as e:
        await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "failed", "error": str(e)[:200]}})
        logging.error(f"Templated email [{template_key}] error: {e}")


async def batch_event_notification(event: dict, membership_type_filter: str = "all"):
    """Send event notification in batches of 50 every 30 mins"""
    query = {"status": "approved"}
    if membership_type_filter and membership_type_filter not in ("all", ""):
        query["membership_type"] = membership_type_filter

    members = await db.members.find(query, {"_id": 0}).to_list(10000)
    members_with_email = [m for m in members if m.get("email")]
    batch_size = 50
    smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})

    template = await db.email_templates.find_one({"key": "event_notification"}, {"_id": 0})
    if not template:
        template = DEFAULT_EMAIL_TEMPLATES["event_notification"]

    for i in range(0, len(members_with_email), batch_size):
        batch = members_with_email[i:i + batch_size]
        for member in batch:
            variables = {
                "member_name": member.get("name", "Member"),
                "event_title": event.get("title", ""),
                "event_date": event.get("date", ""),
                "event_end_date": event.get("end_date", ""),
                "event_venue": event.get("venue", ""),
                "venue_map_link": f'<a href="{event.get("venue_map_link", "")}">[View on Map]</a>' if event.get("venue_map_link") else "",
                "event_description": event.get("description", ""),
                "registration_fee": str(event.get("registration_fee", 0)),
            }
            subject = render_template(template.get("subject", ""), variables)
            body = render_template(template.get("body", ""), variables)
            send_email_smtp([member["email"]], subject, body, smtp_settings)

        sent_count = min(i + batch_size, len(members_with_email))
        await db.email_logs.insert_one({
            "id": str(uuid.uuid4()),
            "subject": f"Event: {event.get('title', '')} (batch {i // batch_size + 1})",
            "body": f"Batch notification for event",
            "recipients_count": len(batch),
            "recipient_group": f"event_batch_{membership_type_filter}",
            "sent_by": "system",
            "sent_at": now_iso(),
            "status": "sent"
        })
        # WhatsApp event invite for members with phone
        for member in batch:
            if member.get("phone"):
                await send_whatsapp_notification(member["phone"], "event_invite", {
                    "name": member.get("name", "Member"),
                    "event_title": event.get("title", ""),
                    "event_date": event.get("date", ""),
                    "event_venue": event.get("venue", ""),
                    "event_description": event.get("description", "")[:200],
                })
                await asyncio.sleep(0.5)
        logging.info(f"Event notification batch {i // batch_size + 1}: sent to {len(batch)} members ({sent_count}/{len(members_with_email)})")

        if i + batch_size < len(members_with_email):
            await asyncio.sleep(1800)  # 30 minutes between batches


async def send_event_participation_certificates(event: dict):
    """Send participation certificates to all approved members when event closes"""
    members = await db.members.find({"status": "approved"}, {"_id": 0}).to_list(10000)
    members_with_email = [m for m in members if m.get("email")]
    smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})

    template = await db.email_templates.find_one({"key": "event_participation"}, {"_id": 0})
    if not template:
        template = DEFAULT_EMAIL_TEMPLATES["event_participation"]

    batch_size = 50
    for i in range(0, len(members_with_email), batch_size):
        batch = members_with_email[i:i + batch_size]
        for member in batch:
            variables = {
                "member_name": member.get("name", ""),
                "membership_id": member.get("membership_id", ""),
                "event_title": event.get("title", ""),
                "event_date": event.get("date", ""),
                "event_venue": event.get("venue", ""),
            }
            subject = render_template(template.get("subject", ""), variables)
            body = render_template(template.get("body", ""), variables)

            # Generate certificate on-the-fly (NOT stored in DB)
            cert_pdf = generate_certificate_pdf_bytes(
                member_name=member.get("name", ""),
                membership_id=member.get("membership_id", ""),
                cert_type="event",
                event_name=event.get("title", ""),
                issue_date=datetime.now().strftime("%d %B %Y")
            )
            attachments = [{"filename": f"IDSEA_Certificate_{event.get('title', 'Event').replace(' ', '_')}.pdf", "data": cert_pdf}]
            send_email_smtp([member["email"]], subject, body, smtp_settings, attachments)

            # WhatsApp: Send participation certificate
            if member.get("phone"):
                wa_vars = {
                    "name": member.get("name", ""),
                    "event_title": event.get("title", ""),
                    "event_date": event.get("date", ""),
                    "event_venue": event.get("venue", ""),
                }
                wa_msg = await _render_wa_template("event_certificate", wa_vars)
                if wa_msg:
                    await send_whatsapp_document_bytes(
                        member["phone"], cert_pdf,
                        f"IDSEA_Certificate_{event.get('title', 'Event').replace(' ', '_')}.pdf",
                        wa_msg
                    )
                    await asyncio.sleep(0.5)

        await db.email_logs.insert_one({
            "id": str(uuid.uuid4()),
            "subject": f"Participation Cert: {event.get('title', '')} (batch {i // batch_size + 1})",
            "body": "Auto-generated participation certificates",
            "recipients_count": len(batch),
            "recipient_group": "event_participation",
            "sent_by": "system",
            "sent_at": now_iso(),
            "status": "sent"
        })
        if i + batch_size < len(members_with_email):
            await asyncio.sleep(1800)


# =================== AUTH ROUTES ===================

@api_router.post("/auth/login")
async def admin_login(data: AdminLogin):
    admin = await db.admins.find_one({"email": data.email}, {"_id": 0})
    if not admin or not verify_password(data.password, admin.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_access_token({"sub": admin["id"], "role": admin["role"]})
    return {
        "access_token": token,
        "token_type": "bearer",
        "admin": {"id": admin["id"], "email": admin["email"], "role": admin["role"], "username": admin["username"]}
    }


@api_router.get("/auth/me")
async def get_me(admin=Depends(get_current_admin)):
    return {"id": admin["id"], "email": admin["email"], "role": admin["role"], "username": admin["username"]}


# =================== PUBLIC ROUTES ===================

@api_router.get("/public/stats")
async def get_public_stats():
    return {
        "total_members": await db.members.count_documents({"status": "approved"}),
        "total_events": await db.events.count_documents({}),
        "total_publications": await db.publications.count_documents({}),
        "upcoming_events": await db.events.count_documents({"status": "upcoming"})
    }


@api_router.get("/public/member-stats")
async def get_public_member_stats():
    """Category-wise member registration counts"""
    pipeline = [
        {"$group": {"_id": "$membership_type", "count": {"$sum": 1}}}
    ]
    results = await db.members.aggregate(pipeline).to_list(20)
    category_counts = {r["_id"]: r["count"] for r in results if r["_id"]}
    total = await db.members.count_documents({})
    return {
        "total": total,
        "categories": category_counts
    }


@api_router.get("/public/members")
async def get_public_members(
    state: Optional[str] = None,
    specialization: Optional[str] = None,
    membership_type: Optional[str] = None,
    search: Optional[str] = None
):
    query: dict = {"status": "approved"}
    if state and state != "all":
        query["state"] = state
    if specialization:
        query["specialization"] = {"$regex": specialization, "$options": "i"}
    if membership_type and membership_type != "all":
        query["membership_type"] = membership_type
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"organization": {"$regex": search, "$options": "i"}},
            {"specialization": {"$regex": search, "$options": "i"}}
        ]
    members = await db.members.find(query, {"_id": 0, "phone": 0, "address": 0}).to_list(500)
    return members


@api_router.post("/public/members/apply")
async def apply_membership(data: MemberCreate, background_tasks: BackgroundTasks):
    existing = await db.members.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    member = Member(**data.model_dump())
    member.join_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Derive state from permanent_address for backward compat
    if data.permanent_address and data.permanent_address.get("state"):
        member.state = data.permanent_address["state"]
    await db.members.insert_one(member.model_dump())

    # Send registration email with application form PDF
    if member.email:
        variables = {
            "member_name": f"{member.prefix} {member.name}".strip() if member.prefix else member.name,
            "email": member.email,
            "phone": member.phone or "",
            "qualification": member.qualification or "",
            "organization": member.organization or "",
            "membership_type": member.membership_type,
            "state": member.state or "",
            "payment_status": member.payment_status,
            "application_date": member.join_date,
        }
        # Generate application form PDF with all filled data
        app_pdf = generate_registration_pdf(member.model_dump())
        pdf_filename = f"IDSEA_Application_{member.name.replace(' ', '_')}.pdf"
        attachments = [{"filename": pdf_filename, "data": app_pdf}]
        background_tasks.add_task(send_templated_email, "registration_submitted", [member.email], variables, attachments)

    # Send WhatsApp notification with application PDF
    if member.phone:
        app_pdf_wa = generate_registration_pdf(member.model_dump()) if not member.email else app_pdf
        background_tasks.add_task(
            send_whatsapp_document_bytes, member.phone, app_pdf_wa,
            f"IDSEA_Application_{member.name.replace(' ', '_')}.pdf",
            f"Hello {member.prefix} {member.name}!\n\nThank you for applying for IDSEA {member.membership_type} membership. Please find your application form attached.\n\nRegards,\nIDSEA Team"
        )

    return {"message": "Application submitted successfully", "id": member.id}


@api_router.post("/public/upload-photo")
async def public_upload_photo(file: UploadFile = File(...)):
    try:
        ext = Path(file.filename).suffix.lower()
        if ext not in ['.jpg', '.jpeg', '.png', '.webp']:
            raise HTTPException(status_code=400, detail="Only image files (jpg, png, webp) allowed")
        content = await file.read()
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large (max 5MB)")
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        unique_name = f"member_{uuid.uuid4().hex[:12]}{ext}"
        file_path = UPLOAD_DIR / unique_name
        with open(file_path, 'wb') as f:
            f.write(content)
        if not file_path.exists():
            raise HTTPException(status_code=500, detail="File verification failed")
        logging.info(f"Photo upload OK: {file_path} size={len(content)}")
        return {"file_url": f"/api/uploads/{unique_name}"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Photo upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@api_router.get("/public/events")
async def get_public_events():
    return await db.events.find({"is_visible": {"$ne": False}}, {"_id": 0}).sort("date", -1).to_list(100)


@api_router.get("/public/events/{event_id}")
async def get_public_event(event_id: str):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    return event


@api_router.get("/public/events/{event_id}/details")
async def get_public_event_details(event_id: str):
    doc = await db.event_details.find_one({"event_id": event_id}, {"_id": 0})
    return doc or {"event_id": event_id}


@api_router.get("/admin/events/{event_id}/details")
async def get_admin_event_details(event_id: str, admin=Depends(get_current_admin)):
    doc = await db.event_details.find_one({"event_id": event_id}, {"_id": 0})
    return doc or {"event_id": event_id}


@api_router.put("/admin/events/{event_id}/details")
async def update_admin_event_details(event_id: str, data: dict, admin=Depends(get_current_admin)):
    data["event_id"] = event_id
    data["updated_at"] = now_iso()
    await db.event_details.update_one(
        {"event_id": event_id},
        {"$set": data},
        upsert=True
    )
    return {"message": "Event details updated"}


@api_router.get("/public/news")
async def get_public_news(category: Optional[str] = None):
    query: dict = {"status": "published"}
    if category and category != "all":
        query["category"] = category
    return await db.news.find(query, {"_id": 0}).sort("published_date", -1).to_list(50)


@api_router.get("/public/news/{news_id}")
async def get_public_news_single(news_id: str):
    item = await db.news.find_one({"id": news_id, "status": "published"}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    return item


@api_router.get("/public/gallery")
async def get_public_gallery():
    albums = await db.gallery_albums.find({}, {"_id": 0}).to_list(50)
    for album in albums:
        photos = await db.gallery_photos.find({"album_id": album["id"]}, {"_id": 0}).to_list(50)
        album["photos"] = photos
        album["photo_count"] = len(photos)
    return albums


@api_router.get("/public/publications")
async def get_public_publications(category: Optional[str] = None):
    query: dict = {}
    if category and category != "all":
        query["category"] = category
    return await db.publications.find(query, {"_id": 0}).to_list(100)


@api_router.get("/public/executive")
async def get_public_executive():
    return await db.executive_committee.find({"category": "council"}, {"_id": 0}).sort("order", 1).to_list(100)


@api_router.get("/public/founders")
async def get_public_founders():
    return await db.executive_committee.find({"category": "founder"}, {"_id": 0}).sort("order", 1).to_list(50)


@api_router.get("/public/cms")
async def get_public_cms():
    settings = await db.cms_settings.find_one({}, {"_id": 0})
    return settings or {}


@api_router.get("/public/page-content/{page}")
async def get_public_page_content(page: str):
    doc = await db.page_contents.find_one({"page": page}, {"_id": 0})
    return doc.get("content", {}) if doc else {}


@api_router.post("/public/contact")
async def submit_contact_form(data: dict, background_tasks: BackgroundTasks):
    name = data.get("name", "").strip()
    email = data.get("email", "").strip()
    subject = data.get("subject", "").strip()
    message = data.get("message", "").strip()
    if not name or not email or not subject or not message:
        raise HTTPException(status_code=400, detail="All fields are required")

    # Store in DB
    contact_entry = {
        "id": str(uuid.uuid4()),
        "name": name, "email": email, "subject": subject, "message": message,
        "status": "new", "created_at": now_iso()
    }
    await db.contact_messages.insert_one(contact_entry)

    # Send email to IDSEA admin with CC to sender
    to_email = "idsea2026@gmail.com"
    smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})
    email_body = f"""<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto;">
  <div style="background: #0c3c60; padding: 24px; text-align: center;">
    <h1 style="color: white; margin: 0; font-size: 20px;">New Contact Message</h1>
  </div>
  <div style="padding: 28px; background: white;">
    <table style="width: 100%; font-size: 14px; color: #374151; border-collapse: collapse;">
      <tr><td style="padding: 10px 0; color: #6b7280; width: 100px; vertical-align: top; font-weight: 600;">Name</td><td style="padding: 10px 0;">{name}</td></tr>
      <tr><td style="padding: 10px 0; color: #6b7280; vertical-align: top; font-weight: 600;">Email</td><td style="padding: 10px 0;"><a href="mailto:{email}">{email}</a></td></tr>
      <tr><td style="padding: 10px 0; color: #6b7280; vertical-align: top; font-weight: 600;">Subject</td><td style="padding: 10px 0; font-weight: 600;">{subject}</td></tr>
      <tr><td style="padding: 10px 0; color: #6b7280; vertical-align: top; font-weight: 600;">Message</td><td style="padding: 10px 0; line-height: 1.7;">{message}</td></tr>
    </table>
    <div style="margin-top: 20px; padding: 14px; background: #f0f9ff; border-radius: 8px; font-size: 12px; color: #6b7280;">
      Received on {datetime.now().strftime('%d %B %Y at %I:%M %p')}
    </div>
  </div>
  <div style="padding: 16px; background: #f8fafc; text-align: center; font-size: 11px; color: #9ca3af;">
    IDSEA Contact Form | www.idsea.in
  </div>
</div>"""

    async def _send_contact_email():
        try:
            host = smtp_settings.get("smtp_host", SMTP_HOST) if smtp_settings else SMTP_HOST
            port = int(smtp_settings.get("smtp_port", SMTP_PORT_VAL) if smtp_settings else SMTP_PORT_VAL)
            user = smtp_settings.get("smtp_user", SMTP_USER) if smtp_settings else SMTP_USER
            passwd = smtp_settings.get("smtp_pass", SMTP_PASS) if smtp_settings else SMTP_PASS
            from_email = smtp_settings.get("from_email", FROM_EMAIL) if smtp_settings else FROM_EMAIL
            if not host or not user:
                logging.warning("SMTP not configured, contact email skipped")
                return
            msg = MIMEMultipart('mixed')
            msg['Subject'] = f"[IDSEA Contact] {subject}"
            msg['From'] = f"IDSEA Website <{from_email}>"
            msg['To'] = to_email
            msg['Cc'] = email
            msg['Reply-To'] = email
            msg.attach(MIMEText(email_body, 'html'))
            with smtplib.SMTP(host, port) as server:
                server.starttls()
                server.login(user, passwd)
                server.sendmail(from_email, [to_email, email], msg.as_string())
            await db.contact_messages.update_one({"id": contact_entry["id"]}, {"$set": {"status": "emailed"}})
            logging.info(f"Contact form email sent to {to_email}, CC: {email}")
        except Exception as e:
            logging.error(f"Contact form email error: {e}")
            await db.contact_messages.update_one({"id": contact_entry["id"]}, {"$set": {"status": "email_failed", "error": str(e)[:200]}})

    background_tasks.add_task(_send_contact_email)
    return {"message": "Your message has been sent. We'll get back to you soon."}




@api_router.get("/public/robots.txt")
async def get_robots_txt():
    from fastapi.responses import PlainTextResponse
    base_url = "https://idsea.org"
    robots = f"""User-agent: *
Allow: /
Sitemap: {base_url}/api/public/sitemap.xml
"""
    return PlainTextResponse(content=robots, media_type="text/plain")


@api_router.get("/public/sitemap.xml")
async def get_sitemap_xml():
    from fastapi.responses import Response
    base_url = "https://idsea.org"
    pages = [
        {"loc": "/", "priority": "1.0", "changefreq": "weekly"},
        {"loc": "/about", "priority": "0.9", "changefreq": "monthly"},
        {"loc": "/ec-members", "priority": "0.8", "changefreq": "monthly"},
        {"loc": "/members", "priority": "0.8", "changefreq": "weekly"},
        {"loc": "/events", "priority": "0.9", "changefreq": "weekly"},
        {"loc": "/publications", "priority": "0.7", "changefreq": "monthly"},
        {"loc": "/gallery", "priority": "0.6", "changefreq": "monthly"},
        {"loc": "/contact", "priority": "0.7", "changefreq": "yearly"},
        {"loc": "/apply", "priority": "0.8", "changefreq": "monthly"},
    ]
    urls = ""
    for p in pages:
        urls += f"""  <url>
    <loc>{base_url}{p['loc']}</loc>
    <changefreq>{p['changefreq']}</changefreq>
    <priority>{p['priority']}</priority>
  </url>\n"""
    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
{urls}</urlset>"""
    return Response(content=xml, media_type="application/xml")


# =================== ADMIN MEMBER ROUTES ===================

@api_router.get("/admin/members")
async def admin_get_members(
    status: Optional[str] = None,
    membership_type: Optional[str] = None,
    search: Optional[str] = None,
    admin=Depends(get_current_admin)
):
    query: dict = {}
    if status and status != "all":
        query["status"] = status
    if membership_type and membership_type != "all":
        query["membership_type"] = membership_type
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"organization": {"$regex": search, "$options": "i"}}
        ]
    return await db.members.find(query, {"_id": 0}).to_list(1000)


@api_router.get("/admin/members/{member_id}")
async def admin_get_member(member_id: str, admin=Depends(get_current_admin)):
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    return member


@api_router.post("/admin/members")
async def admin_create_member(data: MemberCreate, admin=Depends(get_current_admin)):
    existing = await db.members.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    member = Member(**data.model_dump(), status="approved")
    member.join_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    member.membership_id = await generate_membership_id(data.membership_type)
    if data.permanent_address and data.permanent_address.get("state"):
        member.state = data.permanent_address["state"]
    await db.members.insert_one(member.model_dump())
    return member.model_dump()


@api_router.put("/admin/members/{member_id}")
async def admin_update_member(member_id: str, data: MemberUpdate, admin=Depends(get_current_admin)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = now_iso()
    await db.members.update_one({"id": member_id}, {"$set": update_data})
    return {"message": "Member updated"}


@api_router.delete("/admin/members/{member_id}")
async def admin_delete_member(member_id: str, admin=Depends(get_current_admin)):
    await db.members.delete_one({"id": member_id})
    return {"message": "Member deleted"}


@api_router.put("/admin/members/{member_id}/approve")
async def approve_member(member_id: str, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Generate membership ID using configurable prefix
    mtype = member.get("membership_type", "academic")
    membership_id = await generate_membership_id(mtype)

    # Derive state from permanent address if not set
    state = member.get("state", "") or member.get("permanent_address", {}).get("state", "")

    await db.members.update_one(
        {"id": member_id},
        {"$set": {"status": "approved", "membership_id": membership_id, "state": state, "updated_at": now_iso()}}
    )

    # Generate membership certificate using Certificate Designer template (if linked)
    cert_pdf = None
    cert_id = ""
    cert_filename = f"IDSEA_Membership_Certificate_{membership_id.replace('/', '_')}.pdf"
    cert_template = await db.certificate_templates.find_one(
        {"$or": [{"linked_membership_types": mtype}, {"linked_membership_type": mtype}]},
        {"_id": 0}
    )
    if cert_template:
        cert_id = _gen_cert_id("MEM")
        site_url = await _get_site_url()
        cert_data = {
            "name": f"{member.get('prefix', '')} {member.get('name', '')}".strip(),
            "membership_id": membership_id,
            "date": datetime.now().strftime("%d.%m.%Y"), "year": str(datetime.now().year),
            "email": member.get("email", ""), "phone": member.get("phone", ""),
            "qualification": member.get("qualification", ""), "specialization": member.get("specialization", ""),
            "organization": member.get("organization", ""), "membership_type": _membership_label(mtype),
            "state": state, "country": member.get("country", "India"),
            "certificate_id": cert_id, "_site_url": site_url,
        }
        try:
            cert_pdf = generate_template_pdf(cert_template, cert_data, cert_id=cert_id)
            # Store certificate record
            await db.certificate_records.insert_one({
                "cert_id": cert_id, "type": "membership", "template_id": cert_template.get("id", ""),
                "member_id": member_id, "recipient_name": cert_data["name"],
                "membership_id": membership_id, "membership_type": _membership_label(mtype),
                "issued_date": now_iso(), "data": cert_data,
            })
            logging.info(f"Certificate {cert_id} generated for member {member_id} using designer template")
        except Exception as e:
            logging.error(f"Certificate Designer template generation failed: {e}")
            cert_pdf = None
    # Fallback to basic certificate if no template linked or generation failed
    if not cert_pdf:
        cert_pdf = generate_certificate_pdf_bytes(
            member_name=member.get("name", ""),
            membership_id=membership_id,
            cert_type="membership",
        )

    # Send welcome email with membership certificate attached
    if member.get("email"):
        variables = {
            "member_name": member.get("name", ""),
            "email": member.get("email", ""),
            "membership_id": membership_id,
            "membership_type": _membership_label(member.get("membership_type", "")),
            "qualification": member.get("qualification", ""),
            "organization": member.get("organization", ""),
            "state": member.get("state", ""),
            "join_date": datetime.now().strftime("%d %B %Y"),
        }
        attachments = [{"filename": cert_filename, "data": cert_pdf}]
        background_tasks.add_task(send_templated_email, "membership_approved", [member["email"]], variables, attachments)

    # Send WhatsApp notification with certificate attached
    if member.get("phone"):
        wa_variables = {"name": member.get("name", ""), "membership_id": membership_id, "membership_type": _membership_label(mtype)}
        # Build WhatsApp message from template
        db_wa_template = await db.whatsapp_templates.find_one({"key": "membership_approved"}, {"_id": 0})
        wa_msg_text = ""
        if db_wa_template and db_wa_template.get("enabled", True):
            wa_msg_text = db_wa_template.get("message", "")
        else:
            wa_msg_text = "Congratulations {name}!\n\nYour IDSEA membership has been approved.\n\nMembership ID: {membership_id}\nType: {membership_type}\n\nWelcome to the IDSEA family!\n\nRegards,\nIDSEA Team"
        try:
            wa_msg_text = wa_msg_text.format(**wa_variables)
        except KeyError:
            for k, v in wa_variables.items():
                wa_msg_text = wa_msg_text.replace("{" + k + "}", str(v))
        # Send WhatsApp with certificate PDF attached
        background_tasks.add_task(
            send_whatsapp_document_bytes, member["phone"], cert_pdf, cert_filename, wa_msg_text
        )

    return {"message": "Member approved", "membership_id": membership_id}


@api_router.put("/admin/members/{member_id}/reject")
async def reject_member(member_id: str, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    await db.members.update_one(
        {"id": member_id},
        {"$set": {"status": "rejected", "updated_at": now_iso()}}
    )
    # Send WhatsApp notification
    if member and member.get("phone"):
        background_tasks.add_task(
            send_whatsapp_notification, member["phone"], "membership_denied",
            {"name": member.get("name", "")}
        )
    return {"message": "Member rejected"}


@api_router.put("/admin/members/{member_id}/hold")
async def hold_member(member_id: str, admin=Depends(get_current_admin)):
    await db.members.update_one(
        {"id": member_id},
        {"$set": {"status": "hold", "updated_at": now_iso()}}
    )
    return {"message": "Member put on hold"}


@api_router.put("/admin/members/{member_id}/change-type")
async def change_member_type(member_id: str, data: dict, admin=Depends(get_current_admin)):
    """Change membership type and regenerate membership ID if already approved"""
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    new_type = data.get("membership_type")
    if new_type not in ["academic", "entrepreneur", "corporate"]:
        raise HTTPException(status_code=400, detail="Invalid membership type")
    update = {"membership_type": new_type, "updated_at": now_iso()}
    # Regenerate ID if already approved
    if member.get("status") == "approved":
        update["membership_id"] = await generate_membership_id(new_type)
    await db.members.update_one({"id": member_id}, {"$set": update})
    return {"message": "Membership type changed", "membership_id": update.get("membership_id", member.get("membership_id", ""))}


@api_router.post("/admin/members/{member_id}/send-email")
async def send_member_email(member_id: str, data: dict, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    subject = data.get("subject", "")
    body = data.get("body", "")
    if not subject or not body:
        raise HTTPException(status_code=400, detail="Subject and body required")
    log_id = str(uuid.uuid4())
    await db.email_logs.insert_one({
        "id": log_id, "to": member["email"], "subject": subject,
        "status": "sending", "created_at": now_iso(), "sent_at": "", "template": "custom"
    })

    async def _send_and_update():
        smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})
        try:
            success = send_email_smtp([member["email"]], subject, body, smtp_settings)
            if success:
                await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "sent", "sent_at": now_iso()}})
            else:
                await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "failed", "error": "SMTP send failed"}})
        except Exception as e:
            await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "failed", "error": str(e)[:200]}})

    background_tasks.add_task(_send_and_update)
    return {"message": f"Email sending to {member['email']}"}


@api_router.post("/admin/members/{member_id}/send-whatsapp")
async def send_member_whatsapp(member_id: str, data: dict, admin=Depends(get_current_admin)):
    """Send a custom WhatsApp message to a specific member"""
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    if not member.get("phone"):
        raise HTTPException(status_code=400, detail="Member has no phone number")
    message = data.get("message", "")
    if not message:
        raise HTTPException(status_code=400, detail="Message is required")
    # Use custom_message template
    success = await send_whatsapp_notification(member["phone"], "custom_message", {
        "name": member.get("name", ""), "message": message
    })
    return {"status": "success" if success else "error", "message": "WhatsApp sent" if success else "Send failed"}


@api_router.post("/admin/members/send-payment-reminder")
async def send_payment_reminders(data: dict, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    """Send payment reminder to members with pending payments"""
    target = data.get("target", "all")  # all, specific member_ids list
    member_ids = data.get("member_ids", [])
    query = {"payment_status": {"$ne": "paid"}, "status": {"$in": ["pending", "approved"]}}
    if target == "specific" and member_ids:
        query["id"] = {"$in": member_ids}
    members = await db.members.find(query, {"_id": 0}).to_list(10000)
    sent_count = 0

    async def _send_reminders():
        nonlocal sent_count
        plans = {p["key"]: p for p in await db.membership_plans.find({}, {"_id": 0}).to_list(20)}
        for m in members:
            mtype = m.get("membership_type", "academic")
            plan = plans.get(mtype, {})
            fee = plan.get("fee_inr", 0) if mtype != "international" else plan.get("fee_usd", 100)
            if m.get("phone"):
                await send_whatsapp_notification(m["phone"], "payment_reminder", {
                    "name": m.get("name", ""), "membership_type": mtype,
                    "amount": str(fee), "membership_id": m.get("membership_id", ""),
                })
                sent_count += 1
                await asyncio.sleep(0.5)

    background_tasks.add_task(_send_reminders)
    return {"message": f"Payment reminders being sent to {len(members)} members"}


@api_router.post("/admin/members/send-renewal-reminder")
async def send_renewal_reminders(data: dict, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    """Send membership renewal reminders to approved members"""
    member_ids = data.get("member_ids", [])
    query = {"status": "approved"}
    if member_ids:
        query["id"] = {"$in": member_ids}
    members = await db.members.find(query, {"_id": 0}).to_list(10000)

    async def _send_renewals():
        for m in members:
            if m.get("phone"):
                await send_whatsapp_notification(m["phone"], "membership_renewal", {
                    "name": m.get("name", ""),
                    "membership_type": m.get("membership_type", ""),
                    "membership_id": m.get("membership_id", ""),
                })
                await asyncio.sleep(0.5)

    background_tasks.add_task(_send_renewals)
    return {"message": f"Renewal reminders being sent to {len(members)} members"}


# =================== ADMIN EVENT ROUTES ===================

@api_router.get("/admin/events")
async def admin_get_events(admin=Depends(get_current_admin)):
    return await db.events.find({}, {"_id": 0}).sort("date", -1).to_list(200)


@api_router.post("/admin/events")
async def admin_create_event(data: EventCreate, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    event = Event(**data.model_dump())
    await db.events.insert_one(event.model_dump())
    return event.model_dump()


@api_router.post("/admin/events/{event_id}/notify")
async def notify_event(event_id: str, background_tasks: BackgroundTasks, membership_type: str = "all", admin=Depends(get_current_admin)):
    """Send event notification to approved members in batches (50 per 30 mins)"""
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    background_tasks.add_task(batch_event_notification, event, membership_type)
    return {"message": f"Event notification queued for '{membership_type}' members (batch: 50 per 30 mins)"}


@api_router.put("/admin/events/{event_id}/close")
async def close_event(event_id: str, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    """Close event and auto-send participation certificates to all approved members"""
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    await db.events.update_one({"id": event_id}, {"$set": {"status": "completed", "updated_at": now_iso()}})
    background_tasks.add_task(send_event_participation_certificates, event)
    return {"message": "Event closed. Participation certificates are being sent to all approved members."}


@api_router.put("/admin/events/{event_id}")
async def admin_update_event(event_id: str, data: EventCreate, admin=Depends(get_current_admin)):
    await db.events.update_one({"id": event_id}, {"$set": data.model_dump()})
    return {"message": "Event updated"}


@api_router.delete("/admin/events/{event_id}")
async def admin_delete_event(event_id: str, admin=Depends(get_current_admin)):
    await db.events.delete_one({"id": event_id})
    return {"message": "Event deleted"}


# =================== EVENT REGISTRATION ===================

MEMBERSHIP_FEES = {"academic": 3100, "entrepreneur": 5100, "corporate": 25100}


@api_router.get("/public/events/{event_id}/registration-info")
async def get_event_registration_info(event_id: str):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    if not event.get("registration_enabled"):
        raise HTTPException(status_code=400, detail="Registration not enabled for this event")
    # Get membership plans from DB
    plans = await db.membership_plans.find({}, {"_id": 0}).to_list(100)
    if not plans:
        plans = [
            {"key": "academic", "label": "Academic", "fee_inr": 3100, "fee_usd": 50, "enabled": True},
            {"key": "entrepreneur", "label": "Entrepreneur", "fee_inr": 5100, "fee_usd": 75, "enabled": True},
            {"key": "corporate", "label": "Corporate", "fee_inr": 25100, "fee_usd": 300, "enabled": True},
            {"key": "international", "label": "International Delegates", "fee_inr": 0, "fee_usd": 100, "enabled": True},
        ]
    return {
        "id": event["id"],
        "title": event["title"],
        "date": event.get("date", ""),
        "end_date": event.get("end_date", ""),
        "venue": event.get("venue", ""),
        "venue_map_link": event.get("venue_map_link", ""),
        "description": event.get("description", ""),
        "image_url": event.get("image_url", ""),
        "fee_tiers": event.get("fee_tiers", []),
        "accommodation": event.get("accommodation", {}),
        "allow_membership_registration": event.get("allow_membership_registration", False),
        "membership_plans": plans,
        "additional_person_fee": event.get("additional_person_fee", 0),
        "additional_person_fee_usd": event.get("additional_person_fee_usd", 0),
        "registration_addons": event.get("registration_addons", []),
        "premium_hotels": event.get("premium_hotels", []),
    }


@api_router.get("/public/members/lookup")
async def lookup_member_by_phone(phone: str):
    if not phone or len(phone) < 5:
        raise HTTPException(status_code=400, detail="Invalid phone number")
    member = await db.members.find_one(
        {"phone": {"$regex": phone.strip().replace("+", "\\+"), "$options": "i"}, "status": "approved"},
        {"_id": 0, "name": 1, "email": 1, "phone": 1, "membership_type": 1, "membership_id": 1, "organization": 1, "qualification": 1, "state": 1, "id": 1}
    )
    if not member:
        return {"found": False}
    return {"found": True, "member": member}


@api_router.post("/public/events/{event_id}/register")
async def register_for_event(event_id: str, data: dict, background_tasks: BackgroundTasks):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    if not event.get("registration_enabled"):
        raise HTTPException(status_code=400, detail="Registration not enabled")

    # Duplicate check: if same email already registered for this event, return existing registration
    email = data.get("email", "").strip().lower()
    if email:
        import re as _re
        safe_email = _re.escape(email)
        existing_reg = await db.event_registrations.find_one(
            {"event_id": event_id, "email": {"$regex": f"^{safe_email}$", "$options": "i"}},
            {"_id": 0}
        )
        if existing_reg:
            return {"message": "Registration already exists", "registration": existing_reg, "duplicate": True}

    reg = EventRegistration(
        event_id=event_id,
        is_member=data.get("is_member", False),
        member_id=data.get("member_id", ""),
        member_category=data.get("member_category", ""),
        registration_category=data.get("registration_category", ""),
        name=data.get("name", ""),
        email=data.get("email", ""),
        phone=data.get("phone", ""),
        qualification=data.get("qualification", ""),
        organization=data.get("organization", ""),
        state=data.get("state", ""),
        address_line1=data.get("address_line1", ""),
        address_line2=data.get("address_line2", ""),
        district=data.get("district", ""),
        address_state=data.get("address_state", ""),
        country=data.get("country", "India"),
        pincode=data.get("pincode", ""),
        college=data.get("college", ""),
        university=data.get("university", ""),
        identity_proof_url=data.get("identity_proof_url", ""),
        bonafide_cert_url=data.get("bonafide_cert_url", ""),
        student_proof_url=data.get("student_proof_url", ""),
        accommodation_choice=data.get("accommodation_choice", "none"),
        hotel_name=data.get("hotel_name", ""),
        hotel_room_type=data.get("hotel_room_type", ""),
        hotel_tax_percent=float(data.get("hotel_tax_percent", 0)),
        hotel_base_amount=float(data.get("hotel_base_amount", 0)),
        additional_persons=data.get("additional_persons", []),
        additional_persons_fee=float(data.get("additional_persons_fee", 0)),
        wants_membership=data.get("wants_membership", False),
        membership_type=data.get("membership_type", ""),
        selected_addons=data.get("selected_addons", []),
        addon_fee=float(data.get("addon_fee", 0)),
        registration_fee=float(data.get("registration_fee", 0)),
        accommodation_fee=float(data.get("accommodation_fee", 0)),
        hotel_tax_amount=float(data.get("hotel_tax_amount", 0)),
        membership_fee=float(data.get("membership_fee", 0)),
        total_amount=float(data.get("total_amount", 0)),
        payment_status="pending",
        payment_mode=data.get("payment_mode", "offline"),
    )
    await db.event_registrations.insert_one(reg.model_dump())

    # If wants membership, create a pending member application
    if reg.wants_membership and reg.membership_type:
        existing = await db.members.find_one({"email": reg.email})
        if not existing:
            from pydantic import BaseModel as _BM
            new_member = Member(
                name=reg.name,
                email=reg.email,
                phone=reg.phone,
                qualification=reg.qualification or "",
                organization=reg.organization or "",
                state=reg.state or "",
                membership_type=reg.membership_type,
                status="pending",
                payment_status="pending",
            )
            new_member.join_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            await db.members.insert_one(new_member.model_dump())

    result = reg.model_dump()
    result.pop("_id", None)

    # Send WhatsApp notification for event registration
    if reg.phone:
        background_tasks.add_task(
            send_whatsapp_notification, reg.phone, "event_registered",
            {"name": reg.name, "event_title": event.get("title", ""), "event_date": event.get("date", ""),
             "event_venue": event.get("venue", ""), "venue_map_link": f'Map: {event.get("venue_map_link", "")}' if event.get("venue_map_link") else "", "total_amount": str(reg.total_amount), "payment_status": reg.payment_status}
        )

    return {"message": "Registration successful", "registration": result}


@api_router.get("/admin/events/{event_id}/registrations")
async def admin_get_event_registrations(event_id: str, admin=Depends(get_current_admin)):
    regs = await db.event_registrations.find({"event_id": event_id}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return regs


@api_router.get("/admin/events/{event_id}/registrations/export")
async def admin_export_event_registrations(event_id: str, admin=Depends(get_current_admin)):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    regs = await db.event_registrations.find({"event_id": event_id}, {"_id": 0}).sort("created_at", -1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"
    headers = [
        "S.No", "Name", "Email", "Phone", "Organization", "Qualification",
        "Reg Category", "Member", "Member ID", "Member Type",
        "Address Line 1", "Address Line 2", "District", "State", "Country", "Pincode",
        "College", "University",
        "Accommodation", "Hotel", "Room Type", "Hotel Base Amt", "Hotel Tax %", "Hotel Tax Amt",
        "Additional Persons", "Add. Persons Fee",
        "Add-ons", "Add-on Fee",
        "Wants Membership", "Membership Type", "Membership Fee",
        "Reg Fee", "Accom Fee", "Total Amount",
        "Payment Status", "Payment Mode",
        "Assigned Room", "Assigned Location", "Notified", "Registered On"
    ]
    header_fill = openpyxl.styles.PatternFill(start_color="0C3C60", end_color="0C3C60", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, r in enumerate(regs, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r.get("name", ""))
        ws.cell(row=row, column=3, value=r.get("email", ""))
        ws.cell(row=row, column=4, value=r.get("phone", ""))
        ws.cell(row=row, column=5, value=r.get("organization", ""))
        ws.cell(row=row, column=6, value=r.get("qualification", ""))
        ws.cell(row=row, column=7, value=r.get("registration_category", ""))
        ws.cell(row=row, column=8, value="Yes" if r.get("is_member") else "No")
        ws.cell(row=row, column=9, value=r.get("member_id", ""))
        ws.cell(row=row, column=10, value=r.get("member_category", ""))
        ws.cell(row=row, column=11, value=r.get("address_line1", ""))
        ws.cell(row=row, column=12, value=r.get("address_line2", ""))
        ws.cell(row=row, column=13, value=r.get("district", ""))
        ws.cell(row=row, column=14, value=r.get("address_state", r.get("state", "")))
        ws.cell(row=row, column=15, value=r.get("country", "India"))
        ws.cell(row=row, column=16, value=r.get("pincode", ""))
        ws.cell(row=row, column=17, value=r.get("college", ""))
        ws.cell(row=row, column=18, value=r.get("university", ""))
        ws.cell(row=row, column=19, value=r.get("accommodation_choice", ""))
        ws.cell(row=row, column=20, value=r.get("hotel_name", ""))
        ws.cell(row=row, column=21, value=r.get("hotel_room_type", ""))
        ws.cell(row=row, column=22, value=r.get("hotel_base_amount", 0))
        ws.cell(row=row, column=23, value=r.get("hotel_tax_percent", 0))
        ws.cell(row=row, column=24, value=r.get("hotel_tax_amount", 0))
        add_persons = r.get("additional_persons", [])
        ws.cell(row=row, column=25, value=", ".join([f"{p.get('name','')} (Age:{p.get('age','')}, Mob:{p.get('mobile','')})" for p in add_persons]) if add_persons else "")
        ws.cell(row=row, column=26, value=r.get("additional_persons_fee", 0))
        ws.cell(row=row, column=27, value=", ".join(r.get("selected_addons", [])))
        ws.cell(row=row, column=28, value=r.get("addon_fee", 0))
        ws.cell(row=row, column=29, value="Yes" if r.get("wants_membership") else "No")
        ws.cell(row=row, column=30, value=r.get("membership_type", ""))
        ws.cell(row=row, column=31, value=r.get("membership_fee", 0))
        ws.cell(row=row, column=32, value=r.get("registration_fee", 0))
        ws.cell(row=row, column=33, value=r.get("accommodation_fee", 0))
        ws.cell(row=row, column=34, value=r.get("total_amount", 0))
        ws.cell(row=row, column=35, value=r.get("payment_status", ""))
        ws.cell(row=row, column=36, value=r.get("payment_mode", ""))
        ws.cell(row=row, column=37, value=r.get("assigned_room_no", ""))
        ws.cell(row=row, column=38, value=r.get("assigned_location", ""))
        ws.cell(row=row, column=39, value="Yes" if r.get("accommodation_notified") else "No")
        ws.cell(row=row, column=40, value=r.get("created_at", ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    title_clean = event.get("title", "Event").replace(" ", "_")[:30]
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": f"attachment; filename=IDSEA_{title_clean}_Registrations.xlsx"})


@api_router.put("/admin/event-registrations/{reg_id}/payment")
async def admin_update_registration_payment(reg_id: str, data: dict, admin=Depends(get_current_admin)):
    status = data.get("payment_status", "paid")
    await db.event_registrations.update_one({"id": reg_id}, {"$set": {"payment_status": status}})
    return {"message": "Payment status updated"}


@api_router.put("/admin/event-registrations/{reg_id}/accommodation")
async def admin_update_registration_accommodation(reg_id: str, data: dict, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    update = {
        "assigned_room_no": data.get("assigned_room_no", ""),
        "assigned_location": data.get("assigned_location", ""),
        "assigned_location_type": data.get("assigned_location_type", ""),
        "assigned_map_link": data.get("assigned_map_link", ""),
    }
    await db.event_registrations.update_one({"id": reg_id}, {"$set": update})

    # Send WhatsApp notification for room allotment
    reg = await db.event_registrations.find_one({"id": reg_id}, {"_id": 0})
    if reg and reg.get("phone") and update.get("assigned_room_no"):
        event = await db.events.find_one({"id": reg.get("event_id", "")}, {"_id": 0})
        map_link = f"Map: {update['assigned_map_link']}" if update.get("assigned_map_link") else ""
        background_tasks.add_task(
            send_whatsapp_notification, reg["phone"], "room_allotment",
            {"name": reg.get("name", ""), "event_title": event.get("title", "") if event else "",
             "room_no": update["assigned_room_no"], "location": update.get("assigned_location", ""), "map_link": map_link}
        )

    return {"message": "Accommodation details updated"}


@api_router.put("/admin/event-registrations/{reg_id}")
async def admin_update_registration(reg_id: str, data: dict, admin=Depends(get_current_admin)):
    reg = await db.event_registrations.find_one({"id": reg_id}, {"_id": 0})
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    allowed = ["name", "email", "phone", "qualification", "organization", "state",
               "is_member", "member_id", "member_category", "registration_category",
               "address_line1", "address_line2", "district", "address_state", "country", "pincode",
               "college", "university",
               "accommodation_choice", "hotel_name", "hotel_room_type",
               "wants_membership", "membership_type", "selected_addons",
               "registration_fee", "accommodation_fee", "addon_fee",
               "additional_persons_fee", "membership_fee", "total_amount",
               "payment_status", "payment_mode"]
    float_fields = {"registration_fee", "accommodation_fee", "addon_fee",
                    "additional_persons_fee", "membership_fee", "total_amount"}
    bool_fields = {"is_member", "wants_membership"}
    update = {}
    for f in allowed:
        if f in data:
            val = data[f]
            if f in float_fields:
                val = float(val)
            elif f in bool_fields:
                val = bool(val)
            update[f] = val
    if update:
        await db.event_registrations.update_one({"id": reg_id}, {"$set": update})
    return {"message": "Registration updated"}


@api_router.delete("/admin/event-registrations/{reg_id}")
async def admin_delete_registration(reg_id: str, admin=Depends(get_current_admin)):
    reg = await db.event_registrations.find_one({"id": reg_id}, {"_id": 0})
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    await db.event_registrations.delete_one({"id": reg_id})
    return {"message": "Registration deleted"}


@api_router.post("/admin/events/{event_id}/register-manual")
async def admin_manual_registration(event_id: str, data: dict, admin=Depends(get_current_admin)):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    reg = EventRegistration(
        event_id=event_id,
        is_member=data.get("is_member", False),
        member_id=data.get("member_id", ""),
        member_category=data.get("member_category", ""),
        registration_category=data.get("registration_category", "non_member"),
        name=data.get("name", ""),
        email=data.get("email", ""),
        phone=data.get("phone", ""),
        qualification=data.get("qualification", ""),
        organization=data.get("organization", ""),
        state=data.get("state", data.get("address_state", "")),
        address_line1=data.get("address_line1", ""),
        address_line2=data.get("address_line2", ""),
        district=data.get("district", ""),
        address_state=data.get("address_state", ""),
        country=data.get("country", "India"),
        pincode=data.get("pincode", ""),
        college=data.get("college", ""),
        university=data.get("university", ""),
        accommodation_choice=data.get("accommodation_choice", "none"),
        hotel_name=data.get("hotel_name", ""),
        hotel_room_type=data.get("hotel_room_type", ""),
        additional_persons=data.get("additional_persons", []),
        additional_persons_fee=float(data.get("additional_persons_fee", 0)),
        selected_addons=data.get("selected_addons", []),
        addon_fee=float(data.get("addon_fee", 0)),
        wants_membership=data.get("wants_membership", False),
        membership_type=data.get("membership_type", ""),
        registration_fee=float(data.get("registration_fee", 0)),
        accommodation_fee=float(data.get("accommodation_fee", 0)),
        membership_fee=float(data.get("membership_fee", 0)),
        total_amount=float(data.get("total_amount", 0)),
        payment_status=data.get("payment_status", "pending"),
        payment_mode=data.get("payment_mode", "manual"),
    )
    await db.event_registrations.insert_one(reg.model_dump())
    result = reg.model_dump()
    result.pop("_id", None)
    return {"message": "Registration created", "registration": result}


@api_router.get("/admin/events/{event_id}/registrations/export/pdf")
async def admin_export_registrations_pdf(event_id: str, admin=Depends(get_current_admin)):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    regs = await db.event_registrations.find({"event_id": event_id}, {"_id": 0}).sort("created_at", -1).to_list(5000)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#0c3c60"), spaceAfter=6)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.grey, alignment=TA_CENTER, spaceAfter=12)
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7, leading=9)

    elements.append(Paragraph(f"IDSEA - {event.get('title', 'Event')}", title_style))
    elements.append(Paragraph(f"Date: {event.get('date', '')} | Venue: {event.get('venue', '')} | Total: {len(regs)} registrations", sub_style))

    pdf_headers = ["#", "Name", "Email", "Phone", "Organization", "Reg Category", "State", "Accom", "Hotel", "Room", "Total", "Payment"]
    data = [pdf_headers]
    cat_labels = {"member": "Member", "non_member": "Non-Member", "student": "Student", "international": "Intl"}
    for idx, r in enumerate(regs, 1):
        accom = r.get("accommodation_choice", "-")
        if accom in ("premium_hotel", "hotel"):
            accom = r.get("hotel_name", "Hotel")[:16]
        data.append([
            str(idx),
            Paragraph(r.get("name", "")[:26], cell_style),
            Paragraph(r.get("email", "")[:24], cell_style),
            r.get("phone", "")[:14],
            Paragraph(r.get("organization", "")[:22], cell_style),
            cat_labels.get(r.get("registration_category", ""), r.get("registration_category", "-"))[:12],
            r.get("address_state", r.get("state", ""))[:12],
            accom[:14],
            r.get("hotel_name", "")[:14] if r.get("accommodation_choice") in ("premium_hotel", "hotel") else "-",
            r.get("assigned_room_no", "") or "-",
            f"{r.get('total_amount', 0)}",
            r.get("payment_status", "")[:10],
        ])

    col_widths = [22, 78, 78, 58, 72, 48, 48, 52, 52, 35, 40, 42]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0c3c60")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (-2, 1), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    title_clean = event.get("title", "Event").replace(" ", "_")[:30]
    return StreamingResponse(buf, media_type="application/pdf",
                           headers={"Content-Disposition": f"attachment; filename=IDSEA_{title_clean}_Registrations.pdf"})


@api_router.get("/admin/events/{event_id}/registrations/accommodation-report")
async def admin_export_accommodation_report(event_id: str, admin=Depends(get_current_admin)):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    regs = await db.event_registrations.find(
        {"event_id": event_id, "accommodation_choice": {"$ne": "self"}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accommodation"
    headers = ["#", "Name", "Phone", "Email", "Category", "Accom Choice", "Hotel/Location", "Room No", "Assigned Location", "Location Type", "Map Link", "Accom Fee", "Notified"]
    header_fill = openpyxl.styles.PatternFill(start_color="0C3C60", end_color="0C3C60", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, r in enumerate(regs, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=r.get("name", ""))
        ws.cell(row=i + 1, column=3, value=r.get("phone", ""))
        ws.cell(row=i + 1, column=4, value=r.get("email", ""))
        ws.cell(row=i + 1, column=5, value=r.get("member_category", ""))
        ws.cell(row=i + 1, column=6, value=r.get("accommodation_choice", ""))
        ws.cell(row=i + 1, column=7, value=r.get("hotel_name", ""))
        ws.cell(row=i + 1, column=8, value=r.get("assigned_room_no", ""))
        ws.cell(row=i + 1, column=9, value=r.get("assigned_location", ""))
        ws.cell(row=i + 1, column=10, value=r.get("assigned_location_type", ""))
        ws.cell(row=i + 1, column=11, value=r.get("assigned_map_link", ""))
        ws.cell(row=i + 1, column=12, value=r.get("accommodation_fee", 0))
        ws.cell(row=i + 1, column=13, value="Yes" if r.get("accommodation_notified") else "No")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    title_clean = event.get("title", "Event").replace(" ", "_")[:30]
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": f"attachment; filename=IDSEA_{title_clean}_Accommodation.xlsx"})


@api_router.post("/admin/event-registrations/{reg_id}/send-accommodation")
async def admin_send_accommodation_details(reg_id: str, data: dict, admin=Depends(get_current_admin)):
    reg = await db.event_registrations.find_one({"id": reg_id}, {"_id": 0})
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    event = await db.events.find_one({"id": reg["event_id"]}, {"_id": 0})
    channel = data.get("channel", "email")

    if channel == "email":
        smtp = await db.smtp_settings.find_one({}, {"_id": 0})
        if smtp and smtp.get("smtp_host") and smtp.get("smtp_user"):
            try:
                import smtplib
                from email.mime.multipart import MIMEMultipart
                from email.mime.text import MIMEText

                room = reg.get("assigned_room_no", "TBD")
                location = reg.get("assigned_location", reg.get("hotel_name", "TBD"))
                loc_type = reg.get("assigned_location_type", "")
                map_link = reg.get("assigned_map_link", "")

                body = f"""<html><body style="font-family: Arial, sans-serif;">
                <h2 style="color:#0c3c60;">Accommodation Details - {event.get('title','Event')}</h2>
                <p>Dear {reg.get('name','')},</p>
                <p>Your accommodation details for <strong>{event.get('title','')}</strong> are confirmed:</p>
                <table style="border-collapse:collapse;margin:16px 0;">
                <tr><td style="padding:8px;border:1px solid #ddd;font-weight:bold;">Location Type</td><td style="padding:8px;border:1px solid #ddd;">{loc_type}</td></tr>
                <tr><td style="padding:8px;border:1px solid #ddd;font-weight:bold;">Location</td><td style="padding:8px;border:1px solid #ddd;">{location}</td></tr>
                <tr><td style="padding:8px;border:1px solid #ddd;font-weight:bold;">Room No</td><td style="padding:8px;border:1px solid #ddd;">{room}</td></tr>
                <tr><td style="padding:8px;border:1px solid #ddd;font-weight:bold;">Event Dates</td><td style="padding:8px;border:1px solid #ddd;">{event.get('date','')} - {event.get('end_date','')}</td></tr>
                <tr><td style="padding:8px;border:1px solid #ddd;font-weight:bold;">Venue</td><td style="padding:8px;border:1px solid #ddd;">{event.get('venue','')}</td></tr>
                </table>
                {"<p><strong>Map Link:</strong> <a href='" + map_link + "'>" + map_link + "</a></p>" if map_link else ""}
                <p>For any queries, please contact IDSEA.</p>
                <p>Regards,<br/>IDSEA Team</p></body></html>"""

                msg = MIMEMultipart("alternative")
                msg["Subject"] = f"Accommodation Details - {event.get('title','')}"
                msg["From"] = smtp.get("from_email", smtp["smtp_user"])
                msg["To"] = reg.get("email", "")
                msg.attach(MIMEText(body, "html"))

                server = smtplib.SMTP(smtp["smtp_host"], int(smtp.get("smtp_port", 587)))
                server.starttls()
                server.login(smtp["smtp_user"], smtp["smtp_pass"])
                server.send_message(msg)
                server.quit()
                await db.event_registrations.update_one({"id": reg_id}, {"$set": {"accommodation_notified": True}})
                return {"message": f"Email sent to {reg.get('email', '')}"}
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Email failed: {str(e)}")
        else:
            return {"message": "SMTP not configured. Please configure SMTP settings first.", "warning": True}

    elif channel == "whatsapp":
        await db.event_registrations.update_one({"id": reg_id}, {"$set": {"accommodation_notified": True}})
        room = reg.get("assigned_room_no", "TBD")
        location = reg.get("assigned_location", reg.get("hotel_name", "TBD"))
        loc_type = reg.get("assigned_location_type", "")
        map_link = reg.get("assigned_map_link", "")
        text = f"*IDSEA - Accommodation Details*\n\n*Event:* {event.get('title','')}\n*Location:* {loc_type} - {location}\n*Room No:* {room}\n*Dates:* {event.get('date','')} - {event.get('end_date','')}\n*Venue:* {event.get('venue','')}"
        if map_link:
            text += f"\n*Map:* {map_link}"
        phone = reg.get("phone", "").replace(" ", "").replace("+", "")
        if not phone.startswith("91"):
            phone = "91" + phone
        wa_url = f"https://wa.me/{phone}?text={text}"
        return {"message": "WhatsApp link generated", "whatsapp_url": wa_url}

    return {"message": "Invalid channel"}


@api_router.post("/admin/events/{event_id}/registrations/send-all-accommodation")
async def admin_send_all_accommodation(event_id: str, data: dict, admin=Depends(get_current_admin)):
    channel = data.get("channel", "email")
    regs = await db.event_registrations.find(
        {"event_id": event_id, "accommodation_choice": {"$ne": "self"},
         "assigned_room_no": {"$ne": ""}, "assigned_location": {"$ne": ""}},
        {"_id": 0}
    ).to_list(5000)

    if not regs:
        return {"message": "No registrations with assigned accommodation found", "sent": 0}

    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    sent = 0
    errors = 0

    if channel == "email":
        smtp = await db.smtp_settings.find_one({}, {"_id": 0})
        if not smtp or not smtp.get("smtp_host"):
            return {"message": "SMTP not configured", "sent": 0, "warning": True}
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            server = smtplib.SMTP(smtp["smtp_host"], int(smtp.get("smtp_port", 587)))
            server.starttls()
            server.login(smtp["smtp_user"], smtp["smtp_pass"])

            for reg in regs:
                try:
                    room = reg.get("assigned_room_no", "TBD")
                    location = reg.get("assigned_location", "TBD")
                    loc_type = reg.get("assigned_location_type", "")
                    map_link = reg.get("assigned_map_link", "")
                    body = f"""<html><body style="font-family:Arial,sans-serif;">
                    <h2 style="color:#0c3c60;">Accommodation Details - {event.get('title','')}</h2>
                    <p>Dear {reg.get('name','')},</p>
                    <p>Your accommodation: <strong>{loc_type} - {location}, Room {room}</strong></p>
                    {"<p>Map: <a href='" + map_link + "'>" + map_link + "</a></p>" if map_link else ""}
                    <p>Event: {event.get('date','')} - {event.get('end_date','')} at {event.get('venue','')}</p>
                    <p>Regards, IDSEA Team</p></body></html>"""
                    msg = MIMEMultipart("alternative")
                    msg["Subject"] = f"Accommodation Details - {event.get('title','')}"
                    msg["From"] = smtp.get("from_email", smtp["smtp_user"])
                    msg["To"] = reg.get("email", "")
                    msg.attach(MIMEText(body, "html"))
                    server.send_message(msg)
                    await db.event_registrations.update_one({"id": reg["id"]}, {"$set": {"accommodation_notified": True}})
                    sent += 1
                except:
                    errors += 1
            server.quit()
        except Exception as e:
            return {"message": f"SMTP error: {str(e)}", "sent": sent}

    return {"message": f"Sent to {sent} registrants ({errors} errors)", "sent": sent, "errors": errors}


# =================== ADMIN NEWS ROUTES ===================

@api_router.get("/admin/news")
async def admin_get_news(admin=Depends(get_current_admin)):
    return await db.news.find({}, {"_id": 0}).sort("published_date", -1).to_list(200)


@api_router.post("/admin/news")
async def admin_create_news(data: NewsCreate, admin=Depends(get_current_admin)):
    news_item = News(**data.model_dump())
    await db.news.insert_one(news_item.model_dump())
    return news_item.model_dump()


@api_router.put("/admin/news/{news_id}")
async def admin_update_news(news_id: str, data: NewsCreate, admin=Depends(get_current_admin)):
    await db.news.update_one({"id": news_id}, {"$set": data.model_dump()})
    return {"message": "News updated"}


@api_router.delete("/admin/news/{news_id}")
async def admin_delete_news(news_id: str, admin=Depends(get_current_admin)):
    await db.news.delete_one({"id": news_id})
    return {"message": "News deleted"}


# =================== ADMIN GALLERY ROUTES ===================

@api_router.get("/admin/gallery/albums")
async def admin_get_albums(admin=Depends(get_current_admin)):
    albums = await db.gallery_albums.find({}, {"_id": 0}).to_list(100)
    for album in albums:
        photos = await db.gallery_photos.find({"album_id": album["id"]}, {"_id": 0}).to_list(100)
        album["photos"] = photos
        album["photo_count"] = len(photos)
    return albums


@api_router.post("/admin/gallery/albums")
async def admin_create_album(data: GalleryAlbumCreate, admin=Depends(get_current_admin)):
    album = GalleryAlbum(**data.model_dump())
    await db.gallery_albums.insert_one(album.model_dump())
    return album.model_dump()


@api_router.put("/admin/gallery/albums/{album_id}")
async def admin_update_album(album_id: str, data: GalleryAlbumCreate, admin=Depends(get_current_admin)):
    await db.gallery_albums.update_one({"id": album_id}, {"$set": data.model_dump()})
    return {"message": "Album updated"}


@api_router.delete("/admin/gallery/albums/{album_id}")
async def admin_delete_album(album_id: str, admin=Depends(get_current_admin)):
    await db.gallery_albums.delete_one({"id": album_id})
    await db.gallery_photos.delete_many({"album_id": album_id})
    return {"message": "Album deleted"}


@api_router.get("/admin/gallery/photos/{album_id}")
async def admin_get_photos(album_id: str, admin=Depends(get_current_admin)):
    return await db.gallery_photos.find({"album_id": album_id}, {"_id": 0}).to_list(200)


@api_router.post("/admin/gallery/photos")
async def admin_add_photo(data: GalleryPhotoCreate, admin=Depends(get_current_admin)):
    photo = GalleryPhoto(**data.model_dump())
    await db.gallery_photos.insert_one(photo.model_dump())
    album = await db.gallery_albums.find_one({"id": data.album_id})
    if album and not album.get("cover_image"):
        await db.gallery_albums.update_one({"id": data.album_id}, {"$set": {"cover_image": data.image_url}})
    return photo.model_dump()


@api_router.delete("/admin/gallery/photos/{photo_id}")
async def admin_delete_photo(photo_id: str, admin=Depends(get_current_admin)):
    await db.gallery_photos.delete_one({"id": photo_id})
    return {"message": "Photo deleted"}


# =================== ADMIN PUBLICATIONS ===================

@api_router.get("/admin/publications")
async def admin_get_publications(admin=Depends(get_current_admin)):
    return await db.publications.find({}, {"_id": 0}).to_list(200)


@api_router.post("/admin/publications")
async def admin_create_publication(data: PublicationCreate, admin=Depends(get_current_admin)):
    pub = Publication(**data.model_dump())
    await db.publications.insert_one(pub.model_dump())
    return pub.model_dump()


@api_router.put("/admin/publications/{pub_id}")
async def admin_update_publication(pub_id: str, data: PublicationCreate, admin=Depends(get_current_admin)):
    await db.publications.update_one({"id": pub_id}, {"$set": data.model_dump()})
    return {"message": "Publication updated"}


@api_router.delete("/admin/publications/{pub_id}")
async def admin_delete_publication(pub_id: str, admin=Depends(get_current_admin)):
    await db.publications.delete_one({"id": pub_id})
    return {"message": "Publication deleted"}


# =================== ADMIN EXECUTIVE ===================

@api_router.get("/admin/executive")
async def admin_get_executive(category: Optional[str] = None, admin=Depends(get_current_admin)):
    query = {}
    if category:
        query["category"] = category
    return await db.executive_committee.find(query, {"_id": 0}).sort("order", 1).to_list(100)


@api_router.post("/admin/executive")
async def admin_create_executive(data: ExecutiveCreate, admin=Depends(get_current_admin)):
    exec_member = ExecutiveCommittee(**data.model_dump())
    await db.executive_committee.insert_one(exec_member.model_dump())
    return exec_member.model_dump()


@api_router.put("/admin/executive/{exec_id}")
async def admin_update_executive(exec_id: str, data: ExecutiveCreate, admin=Depends(get_current_admin)):
    await db.executive_committee.update_one({"id": exec_id}, {"$set": data.model_dump()})
    return {"message": "Updated"}


@api_router.delete("/admin/executive/{exec_id}")
async def admin_delete_executive(exec_id: str, admin=Depends(get_current_admin)):
    await db.executive_committee.delete_one({"id": exec_id})
    return {"message": "Deleted"}


# =================== ADMIN EMAIL ===================

@api_router.post("/admin/email/send")
async def admin_send_email(data: EmailRequest, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    recipients = list(data.recipients)
    if data.recipient_group != "custom":
        query: dict = {"status": "approved"}
        if data.recipient_group in ["academic", "entrepreneur", "corporate"]:
            query["membership_type"] = data.recipient_group
        members = await db.members.find(query, {"email": 1, "_id": 0}).to_list(1000)
        recipients = [m["email"] for m in members if m.get("email")]
    log_id = str(uuid.uuid4())
    log = {
        "id": log_id,
        "subject": data.subject,
        "body": data.body[:300],
        "recipients_count": len(recipients),
        "recipient_group": data.recipient_group,
        "sent_by": admin["email"],
        "created_at": now_iso(),
        "sent_at": "",
        "status": "sending"
    }
    await db.email_logs.insert_one(log)
    smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})

    async def _send_and_update():
        try:
            success = send_email_smtp(recipients, data.subject, data.body, smtp_settings)
            if success:
                await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "sent", "sent_at": now_iso()}})
            else:
                await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "failed", "error": "SMTP send returned False"}})
        except Exception as e:
            await db.email_logs.update_one({"id": log_id}, {"$set": {"status": "failed", "error": str(e)[:200]}})

    background_tasks.add_task(_send_and_update)
    return {"message": f"Email sending to {len(recipients)} recipients", "log_id": log_id}


@api_router.get("/admin/email/logs")
async def admin_get_email_logs(admin=Depends(get_current_admin)):
    logs = await db.email_logs.find({}, {"_id": 0}).sort("sent_at", -1).to_list(100)
    return logs


# =================== PAYMENTS ===================

@api_router.post("/payments/create-order")
async def create_payment_order(data: PaymentOrder):
    payment = Payment(
        member_id=data.member_id,
        member_name=data.member_name,
        member_email=data.member_email,
        membership_type=data.membership_type,
        event_registration_id=data.event_registration_id,
        purpose=data.purpose,
        amount=data.amount,
        currency=data.currency
    )
    razorpay_order_id = ""
    rzp_client, key_id = await get_razorpay_client_async()
    if rzp_client:
        try:
            order = rzp_client.order.create({
                "amount": int(data.amount * 100),
                "currency": data.currency,
                "payment_capture": 1,
                "receipt": f"rcpt_{payment.id[:20]}"
            })
            razorpay_order_id = order["id"]
            payment.razorpay_order_id = razorpay_order_id
        except Exception as e:
            logging.error(f"Razorpay error: {e}")
    await db.payments.insert_one(payment.model_dump())
    result = payment.model_dump()
    result.pop("_id", None)
    return {
        "payment_id": payment.id,
        "razorpay_order_id": razorpay_order_id,
        "amount": data.amount,
        "currency": data.currency,
        "key_id": key_id
    }


@api_router.post("/payments/verify")
async def verify_payment(data: PaymentVerify):
    _, key_secret = await get_effective_razorpay_keys()
    if key_secret:
        msg = f"{data.razorpay_order_id}|{data.razorpay_payment_id}"
        expected = hmac.new(
            key_secret.encode('utf-8'),
            msg.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        if not hmac.compare_digest(expected, data.razorpay_signature):
            raise HTTPException(status_code=400, detail="Payment verification failed")
    await db.payments.update_one(
        {"id": data.payment_id},
        {"$set": {"razorpay_payment_id": data.razorpay_payment_id, "status": "paid"}}
    )
    payment = await db.payments.find_one({"id": data.payment_id}, {"_id": 0})
    if payment:
        if payment.get("member_id"):
            await db.members.update_one(
                {"id": payment["member_id"]},
                {"$set": {"payment_status": "paid", "amount_paid": payment["amount"]}}
            )
            # WhatsApp: Payment received for membership
            member = await db.members.find_one({"id": payment["member_id"]}, {"_id": 0})
            if member and member.get("phone"):
                await send_whatsapp_notification(member["phone"], "payment_received", {
                    "name": member.get("name", ""), "amount": str(payment.get("amount", 0)),
                    "reference": data.razorpay_payment_id,
                })
        if payment.get("event_registration_id"):
            await db.event_registrations.update_one(
                {"id": payment["event_registration_id"]},
                {"$set": {"payment_status": "paid", "payment_mode": "razorpay"}}
            )
            # WhatsApp: Event payment confirmed
            reg = await db.event_registrations.find_one({"id": payment["event_registration_id"]}, {"_id": 0})
            if reg and reg.get("phone"):
                event = await db.events.find_one({"id": reg.get("event_id", "")}, {"_id": 0})
                await send_whatsapp_notification(reg["phone"], "event_payment_confirmed", {
                    "name": reg.get("name", ""), "event_title": event.get("title", "") if event else "",
                    "amount": str(payment.get("amount", 0)), "transaction_id": data.razorpay_payment_id,
                    "registration_id": reg.get("id", ""),
                })
    return {"message": "Payment verified", "status": "paid"}


@api_router.post("/payments/submit-utr")
async def submit_utr_payment(data: dict):
    utr = data.get("utr_number", "").strip()
    payment_method = data.get("payment_method", "upi")
    amount = float(data.get("amount", 0))
    purpose = data.get("purpose", "membership")
    if not utr:
        raise HTTPException(status_code=400, detail="UTR number is required")

    payment = Payment(
        member_id=data.get("member_id", ""),
        member_name=data.get("name", ""),
        member_email=data.get("email", ""),
        membership_type=data.get("membership_type", ""),
        event_registration_id=data.get("event_registration_id", ""),
        purpose=purpose,
        amount=amount,
        payment_method=payment_method,
        utr_number=utr,
        status="verification_pending"
    )
    await db.payments.insert_one(payment.model_dump())

    if data.get("event_registration_id"):
        await db.event_registrations.update_one(
            {"id": data["event_registration_id"]},
            {"$set": {"payment_status": "verification_pending", "payment_mode": payment_method}}
        )
    if data.get("member_id"):
        await db.members.update_one(
            {"id": data["member_id"]},
            {"$set": {"payment_status": "verification_pending"}}
        )

    result = payment.model_dump()
    result.pop("_id", None)
    return {"message": "Payment submitted for verification", "payment": result}


@api_router.get("/payments/upi-qr")
async def generate_upi_qr(amount: float, name: str = "IDSEA"):
    settings = await db.payment_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("upi_ids"):
        raise HTTPException(status_code=400, detail="UPI not configured")
    upi_id = settings["upi_ids"][0].get("upi_id", "")
    if not upi_id:
        raise HTTPException(status_code=400, detail="UPI ID not set")

    upi_url = f"upi://pay?pa={upi_id}&pn={name}&am={amount}&cu=INR"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=8, border=2)
    qr.add_data(upi_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#0c3c60", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    img_b64 = base64.b64encode(buf.getvalue()).decode()
    return {"qr_image": f"data:image/png;base64,{img_b64}", "upi_url": upi_url, "upi_id": upi_id, "amount": amount}


@api_router.get("/public/payment-settings")
async def get_public_payment_settings():
    settings = await db.payment_settings.find_one({}, {"_id": 0})
    key_id, key_secret = await get_effective_razorpay_keys()
    if not settings:
        return {"bank_accounts": [], "upi_ids": [], "razorpay_enabled": bool(key_id and key_secret),
                "methods_enabled": {"razorpay": bool(key_id), "upi": True, "bank_transfer": True}}
    methods = settings.get("methods_enabled", {"razorpay": True, "upi": True, "bank_transfer": True})
    rzp_configured = bool(key_id and key_secret)
    return {
        "bank_accounts": settings.get("bank_accounts", []) if methods.get("bank_transfer", True) else [],
        "upi_ids": [{"upi_id": u.get("upi_id", ""), "name": u.get("name", "")} for u in settings.get("upi_ids", [])] if methods.get("upi", True) else [],
        "razorpay_enabled": rzp_configured and methods.get("razorpay", True),
        "methods_enabled": methods
    }


# =================== ADMIN PAYMENT SETTINGS ===================

@api_router.get("/admin/payment-settings")
async def admin_get_payment_settings(admin=Depends(get_current_admin)):
    settings = await db.payment_settings.find_one({}, {"_id": 0})
    key_id, key_secret = await get_effective_razorpay_keys()
    if not settings:
        settings = {"bank_accounts": [], "upi_ids": []}
    settings["razorpay_key_id"] = settings.get("razorpay_key_id", "") or RAZORPAY_KEY_ID
    settings["razorpay_configured"] = bool(key_id and key_secret)
    if "methods_enabled" not in settings:
        settings["methods_enabled"] = {"razorpay": True, "upi": True, "bank_transfer": True}
    # Mask secret for display
    settings["razorpay_key_secret_masked"] = ("*" * 20 + key_secret[-4:]) if key_secret else ""
    return settings


@api_router.put("/admin/payment-settings")
async def admin_update_payment_settings(data: dict, admin=Depends(get_current_admin)):
    update = {
        "bank_accounts": data.get("bank_accounts", []),
        "upi_ids": data.get("upi_ids", []),
        "methods_enabled": data.get("methods_enabled", {"razorpay": True, "upi": True, "bank_transfer": True}),
    }
    # Store Razorpay keys in DB if provided
    if data.get("razorpay_key_id"):
        update["razorpay_key_id"] = data["razorpay_key_id"]
    if data.get("razorpay_key_secret") and not data["razorpay_key_secret"].startswith("*"):
        update["razorpay_key_secret"] = data["razorpay_key_secret"]
    await db.payment_settings.update_one({}, {"$set": update}, upsert=True)
    return {"message": "Payment settings updated"}


@api_router.put("/admin/payments/{payment_id}/verify-utr")
async def admin_verify_utr(payment_id: str, data: dict, admin=Depends(get_current_admin)):
    action = data.get("action", "approve")
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    if action == "approve":
        await db.payments.update_one({"id": payment_id}, {"$set": {"status": "paid"}})
        if payment.get("member_id"):
            await db.members.update_one({"id": payment["member_id"]}, {"$set": {"payment_status": "paid", "amount_paid": payment["amount"]}})
            # WhatsApp: Payment received
            member = await db.members.find_one({"id": payment["member_id"]}, {"_id": 0})
            if member and member.get("phone"):
                await send_whatsapp_notification(member["phone"], "payment_received", {
                    "name": member.get("name", ""), "amount": str(payment.get("amount", 0)),
                    "reference": payment.get("utr_number", payment_id[:8]),
                })
        if payment.get("event_registration_id"):
            await db.event_registrations.update_one({"id": payment["event_registration_id"]}, {"$set": {"payment_status": "paid"}})
            # WhatsApp: Event payment confirmed
            reg = await db.event_registrations.find_one({"id": payment["event_registration_id"]}, {"_id": 0})
            if reg and reg.get("phone"):
                event = await db.events.find_one({"id": reg.get("event_id", "")}, {"_id": 0})
                await send_whatsapp_notification(reg["phone"], "event_payment_confirmed", {
                    "name": reg.get("name", ""), "event_title": event.get("title", "") if event else "",
                    "amount": str(payment.get("amount", 0)),
                    "transaction_id": payment.get("utr_number", payment_id[:8]),
                    "registration_id": reg.get("id", ""),
                })
        return {"message": "Payment approved"}
    else:
        await db.payments.update_one({"id": payment_id}, {"$set": {"status": "rejected"}})
        if payment.get("event_registration_id"):
            await db.event_registrations.update_one({"id": payment["event_registration_id"]}, {"$set": {"payment_status": "rejected"}})
        if payment.get("member_id"):
            await db.members.update_one({"id": payment["member_id"]}, {"$set": {"payment_status": "rejected"}})
        return {"message": "Payment rejected"}


@api_router.put("/admin/payments/{payment_id}")
async def admin_update_payment(payment_id: str, data: dict, admin=Depends(get_current_admin)):
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    update = {}
    for field in ["utr_number", "notes", "member_name", "member_email", "amount", "payment_method"]:
        if field in data:
            update[field] = data[field]
    if "amount" in update:
        update["amount"] = float(update["amount"])
    if update:
        await db.payments.update_one({"id": payment_id}, {"$set": update})
    return {"message": "Payment updated"}


@api_router.delete("/admin/payments/{payment_id}")
async def admin_delete_payment(payment_id: str, admin=Depends(get_current_admin)):
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    await db.payments.delete_one({"id": payment_id})
    # Reset linked registration/member payment status
    if payment.get("event_registration_id"):
        await db.event_registrations.update_one({"id": payment["event_registration_id"]}, {"$set": {"payment_status": "pending"}})
    if payment.get("member_id"):
        await db.members.update_one({"id": payment["member_id"]}, {"$set": {"payment_status": "pending"}})
    return {"message": "Payment deleted"}


@api_router.post("/admin/payments/{payment_id}/refund")
async def admin_refund_payment(payment_id: str, data: dict, admin=Depends(get_current_admin)):
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    refund_notes = data.get("notes", "")
    await db.payments.update_one({"id": payment_id}, {"$set": {
        "status": "refunded", "refund_status": "completed",
        "refund_date": now_iso(), "notes": refund_notes
    }})
    if payment.get("event_registration_id"):
        await db.event_registrations.update_one({"id": payment["event_registration_id"]}, {"$set": {"payment_status": "refunded"}})
    if payment.get("member_id"):
        await db.members.update_one({"id": payment["member_id"]}, {"$set": {"payment_status": "refunded"}})
    return {"message": "Payment refunded"}


@api_router.get("/admin/payments")
async def admin_get_payments(admin=Depends(get_current_admin)):
    return await db.payments.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.post("/admin/payments/manual")
async def admin_add_manual_payment(data: dict, admin=Depends(get_current_admin)):
    method = data.get("payment_method", "manual")
    status = "paid" if method in ["manual", "razorpay"] else "verification_pending"
    payment = Payment(
        member_id=data.get("member_id", ""),
        member_name=data.get("member_name", ""),
        member_email=data.get("member_email", ""),
        membership_type=data.get("membership_type", ""),
        event_registration_id=data.get("event_registration_id", ""),
        purpose=data.get("purpose", "membership"),
        amount=float(data.get("amount", 0)),
        payment_method=method,
        utr_number=data.get("utr_number", ""),
        razorpay_payment_id=data.get("utr_number", "") if method == "razorpay" else "",
        status=status,
        notes=data.get("notes", "")
    )
    await db.payments.insert_one(payment.model_dump())
    result = payment.model_dump()
    result.pop("_id", None)
    return result


# =================== REPORTS & DASHBOARD ===================

@api_router.get("/admin/reports/stats")
async def admin_reports_stats(admin=Depends(get_current_admin)):
    total_members = await db.members.count_documents({})
    approved = await db.members.count_documents({"status": "approved"})
    pending = await db.members.count_documents({"status": "pending"})
    start_of_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    new_month = await db.members.count_documents({"created_at": {"$gte": start_of_month}})
    paid_payments = await db.payments.find({"status": "paid"}, {"amount": 1, "_id": 0}).to_list(1000)
    total_revenue = sum(p.get("amount", 0) for p in paid_payments)
    academic = await db.members.count_documents({"membership_type": "academic", "status": "approved"})
    entrepreneur = await db.members.count_documents({"membership_type": "entrepreneur", "status": "approved"})
    corporate = await db.members.count_documents({"membership_type": "corporate", "status": "approved"})
    pipeline = [
        {"$match": {"status": "approved", "state": {"$nin": ["", None]}}},
        {"$group": {"_id": "$state", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    state_stats = await db.members.aggregate(pipeline).to_list(10)
    state_stats = [{"state": s["_id"], "count": s["count"]} for s in state_stats if s["_id"]]
    return {
        "total_members": total_members,
        "approved_members": approved,
        "pending_members": pending,
        "new_this_month": new_month,
        "total_revenue": total_revenue,
        "membership_types": {"academic": academic, "entrepreneur": entrepreneur, "corporate": corporate},
        "events": {
            "total": await db.events.count_documents({}),
            "upcoming": await db.events.count_documents({"status": "upcoming"})
        },
        "state_stats": state_stats
    }


@api_router.get("/admin/dashboard")
async def admin_dashboard(admin=Depends(get_current_admin)):
    total_members = await db.members.count_documents({})
    pending_approvals = await db.members.count_documents({"status": "pending"})
    start_of_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    new_this_month = await db.members.count_documents({"created_at": {"$gte": start_of_month}})
    upcoming_events = await db.events.count_documents({"status": "upcoming"})
    recent_members = await db.members.find({}, {"_id": 0, "name": 1, "membership_type": 1, "status": 1, "created_at": 1, "email": 1}).sort("created_at", -1).to_list(5)
    recent_payments = await db.payments.find({}, {"_id": 0, "member_name": 1, "amount": 1, "status": 1, "created_at": 1, "membership_type": 1}).sort("created_at", -1).to_list(5)
    return {
        "total_members": total_members,
        "pending_approvals": pending_approvals,
        "new_this_month": new_this_month,
        "upcoming_events": upcoming_events,
        "recent_members": recent_members,
        "recent_payments": recent_payments
    }


# =================== CMS ===================

@api_router.get("/admin/cms")
async def admin_get_cms(admin=Depends(get_current_admin)):
    settings = await db.cms_settings.find_one({}, {"_id": 0})
    return settings or {}


@api_router.put("/admin/cms")
async def admin_update_cms(data: CMSSettings, admin=Depends(get_current_admin)):
    settings_dict = data.model_dump()
    settings_dict["updated_at"] = now_iso()
    await db.cms_settings.update_one({}, {"$set": settings_dict}, upsert=True)
    return {"message": "CMS updated"}


@api_router.get("/admin/page-content/{page}")
async def admin_get_page_content(page: str, admin=Depends(get_current_admin)):
    doc = await db.page_contents.find_one({"page": page}, {"_id": 0})
    return doc.get("content", {}) if doc else {}


@api_router.put("/admin/page-content/{page}")
async def admin_update_page_content(page: str, data: dict, admin=Depends(get_current_admin)):
    await db.page_contents.update_one(
        {"page": page},
        {"$set": {"page": page, "content": data, "updated_at": now_iso()}},
        upsert=True
    )
    return {"message": f"Page content for '{page}' updated"}


# =================== ADMIN USER MANAGEMENT ===================

@api_router.get("/admin/users")
async def admin_get_users(admin=Depends(get_current_admin)):
    return await db.admins.find({}, {"_id": 0, "password_hash": 0}).to_list(100)


@api_router.post("/admin/users")
async def admin_create_user(data: dict, admin=Depends(get_current_admin)):
    if admin.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Only super admin can create users")
    new_admin = AdminUser(
        username=data["username"],
        email=data["email"],
        password_hash=hash_password(data["password"]),
        role=data.get("role", "admin")
    )
    await db.admins.insert_one(new_admin.model_dump())
    return {"id": new_admin.id, "email": new_admin.email, "role": new_admin.role, "username": new_admin.username}


@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, admin=Depends(get_current_admin)):
    if admin.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Only super admin can delete users")
    await db.admins.delete_one({"id": user_id})
    return {"message": "User deleted"}


# =================== CERTIFICATES ===================

@api_router.post("/admin/certificates/generate")
async def generate_certificate(data: CertificateRequest, admin=Depends(get_current_admin)):
    member = await db.members.find_one({"id": data.member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    cert_id = str(uuid.uuid4())[:8].upper()
    issue_date = data.issue_date or datetime.now().strftime("%d %B %Y")
    certificate = {
        "id": cert_id,
        "member_id": data.member_id,
        "member_name": member["name"],
        "membership_id": member.get("membership_id", ""),
        "certificate_type": data.certificate_type,
        "event_name": data.event_name,
        "issue_date": issue_date,
        "issued_by": admin["username"],
        "created_at": now_iso()
    }
    await db.certificates.insert_one(certificate)
    # Remove MongoDB _id before returning
    certificate.pop("_id", None)
    return certificate


@api_router.get("/admin/certificates")
async def admin_get_certificates(admin=Depends(get_current_admin)):
    return await db.certificates.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)


# =================== CERTIFICATE PDF ===================

@api_router.get("/admin/certificates/{cert_id}/pdf")
async def download_certificate_pdf(cert_id: str, admin=Depends(get_current_admin)):
    cert = await db.certificates.find_one({"id": cert_id}, {"_id": 0})
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    w, h = landscape(A4)

    # Border
    c.setStrokeColor(colors.HexColor('#0c3c60'))
    c.setLineWidth(4)
    c.rect(30, 30, w - 60, h - 60)
    c.setStrokeColor(colors.HexColor('#1e7a4d'))
    c.setLineWidth(2)
    c.rect(40, 40, w - 80, h - 80)

    # Header
    c.setFillColor(colors.HexColor('#0c3c60'))
    c.setFont("Helvetica-Bold", 28)
    c.drawCentredString(w / 2, h - 100, "Indian Dairy Scientists and")
    c.drawCentredString(w / 2, h - 135, "Entrepreneurs Association")
    c.setFont("Helvetica", 12)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(w / 2, h - 158, "IDSEA")

    # Certificate type
    cert_type_label = {
        "membership": "Certificate of Membership",
        "event": "Certificate of Participation",
        "appreciation": "Certificate of Appreciation"
    }.get(cert.get("certificate_type", ""), "Certificate")

    c.setFillColor(colors.HexColor('#1e7a4d'))
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(w / 2, h - 210, cert_type_label)

    # "This is to certify..."
    c.setFillColor(colors.HexColor('#374151'))
    c.setFont("Helvetica", 14)
    c.drawCentredString(w / 2, h - 260, "This is to certify that")

    # Name
    c.setFillColor(colors.HexColor('#0c3c60'))
    c.setFont("Helvetica-Bold", 26)
    c.drawCentredString(w / 2, h - 300, cert.get("member_name", ""))

    # Membership ID
    c.setFillColor(colors.HexColor('#6b7280'))
    c.setFont("Helvetica", 12)
    mid = cert.get("membership_id", "")
    if mid:
        c.drawCentredString(w / 2, h - 325, f"Member ID: {mid}")

    # Body text
    c.setFillColor(colors.HexColor('#374151'))
    c.setFont("Helvetica", 13)
    if cert.get("certificate_type") == "event" and cert.get("event_name"):
        c.drawCentredString(w / 2, h - 358, f"has successfully participated in")
        c.setFont("Helvetica-Bold", 15)
        c.setFillColor(colors.HexColor('#1e7a4d'))
        c.drawCentredString(w / 2, h - 380, cert["event_name"])
    elif cert.get("certificate_type") == "membership":
        c.drawCentredString(w / 2, h - 358, "is a registered member of IDSEA")
    else:
        c.drawCentredString(w / 2, h - 358, "is recognized by IDSEA for their contributions")

    # Date and signature
    c.setFillColor(colors.HexColor('#6b7280'))
    c.setFont("Helvetica", 11)
    c.drawString(80, 90, f"Date: {cert.get('issue_date', '')}")
    c.drawString(80, 75, f"Certificate ID: {cert_id}")
    c.line(w - 280, 100, w - 80, 100)
    c.setFont("Helvetica", 11)
    c.drawCentredString(w - 180, 82, "Authorized Signatory")
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(w - 180, 110, cert.get("issued_by", "IDSEA"))

    c.save()
    buf.seek(0)
    filename = f"IDSEA_Certificate_{cert_id}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ============ CERTIFICATE TEMPLATE DESIGNER ============

def _map_font(family, weight="normal", style="normal"):
    fm = {
        ("Helvetica","normal","normal"): "Helvetica", ("Helvetica","bold","normal"): "Helvetica-Bold",
        ("Helvetica","normal","italic"): "Helvetica-Oblique", ("Helvetica","bold","italic"): "Helvetica-BoldOblique",
        ("Times-Roman","normal","normal"): "Times-Roman", ("Times-Roman","bold","normal"): "Times-Bold",
        ("Times-Roman","normal","italic"): "Times-Italic", ("Times-Roman","bold","italic"): "Times-BoldItalic",
        ("Courier","normal","normal"): "Courier", ("Courier","bold","normal"): "Courier-Bold",
        ("Courier","normal","italic"): "Courier-Oblique", ("Courier","bold","italic"): "Courier-BoldOblique",
    }
    return fm.get((family, weight, style), "Helvetica")


def _resolve_img_path(url_or_path: str) -> str:
    if not url_or_path:
        return ""
    p = url_or_path
    # Handle full HTTP URLs - download to temp file
    if p.startswith("http://") or p.startswith("https://"):
        try:
            import urllib.request, tempfile
            suffix = os.path.splitext(p.split("?")[0])[-1] or ".png"
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            urllib.request.urlretrieve(p, tmp.name)
            return tmp.name
        except Exception:
            return ""
    # Handle /api/uploads/... paths (strip /api prefix)
    if p.startswith("/api/"):
        p = p[5:]  # Remove "/api/" -> "uploads/..."
    elif p.startswith("/"):
        p = p.lstrip("/")
    # Try relative path first (CWD is /app/backend)
    if os.path.exists(p):
        return p
    # Try absolute path
    abs_path = os.path.join("/app/backend", p)
    if os.path.exists(abs_path):
        return abs_path
    return ""


def generate_template_pdf(template: dict, data: dict, cert_id: str = "") -> bytes:
    from reportlab.platypus import Paragraph as RLParagraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    import qrcode
    import tempfile

    pw = template.get("page_width", 1000)
    ph = template.get("page_height", 707)
    orient = template.get("orientation", "landscape")
    if orient == "landscape":
        pdf_w, pdf_h = landscape(A4)
    else:
        pdf_w, pdf_h = A4
    sx, sy = pdf_w / pw, pdf_h / ph

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(pdf_w, pdf_h))

    bg_color = template.get("background_color", "#ffffff")
    if bg_color and bg_color.lower() != "#ffffff":
        c.setFillColor(colors.HexColor(bg_color))
        c.rect(0, 0, pdf_w, pdf_h, fill=1, stroke=0)

    bg_img = _resolve_img_path(template.get("background_image_url", ""))
    if bg_img:
        try:
            c.drawImage(bg_img, 0, 0, width=pdf_w, height=pdf_h, preserveAspectRatio=False)
        except Exception:
            pass

    for el in template.get("elements", []):
        etype = el.get("type", "text")
        ex = el.get("x", 0) * sx
        ew = el.get("width", 200) * sx
        eh = el.get("height", 30) * sy
        ey = pdf_h - (el.get("y", 0) * sy) - eh
        clr = el.get("color", "#000000")
        opacity = el.get("opacity", 100) / 100.0
        fs = max(6, el.get("font_size", 16) * min(sx, sy))

        if etype in ("text", "placeholder"):
            content = el.get("content", "")
            if etype == "placeholder":
                key = el.get("placeholder_key", "")
                content = str(data.get(key, f"{{{{{key}}}}}"))
            if not content:
                continue
            ff = el.get("font_family", "Helvetica")
            fw_val = el.get("font_weight", "normal")
            fi = el.get("font_style", "normal")
            ta = el.get("text_align", "left")
            td = el.get("text_decoration", "none")
            fn = _map_font(ff, fw_val, fi)

            # Use Paragraph for multi-line text wrapping
            align_map = {"left": TA_LEFT, "center": TA_CENTER, "right": TA_RIGHT}
            leading = fs * 1.3
            style = ParagraphStyle(
                name='el', fontName=fn, fontSize=fs,
                leading=leading, textColor=colors.HexColor(clr),
                alignment=align_map.get(ta, TA_LEFT),
                spaceBefore=0, spaceAfter=0,
            )
            # Handle multi-line content (newlines from textarea)
            content_clean = content.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            content_clean = content_clean.replace("\n", "<br/>")
            if td == "underline":
                content_clean = f"<u>{content_clean}</u>"
            para = RLParagraph(content_clean, style)
            para_w, para_h = para.wrap(ew, max(eh, fs * 10))
            # Top-align text within the element box (matches CSS flex align-items: flex-start)
            y_offset = ey + eh - min(para_h, eh)
            para.drawOn(c, ex, max(ey, y_offset))

        elif etype == "image":
            ip = _resolve_img_path(el.get("image_url", ""))
            if ip:
                try:
                    c.drawImage(ip, ex, ey, width=ew, height=eh, preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass

        elif etype == "signature_block":
            sn = el.get("signer_name", "")
            st_val = el.get("signer_title", "")
            si = _resolve_img_path(el.get("signature_image_url", ""))
            c.setFillColor(colors.HexColor(clr))
            if si:
                try:
                    sig_h = eh * 0.45
                    c.drawImage(si, ex + ew * 0.1, ey + eh * 0.5, width=ew * 0.8, height=sig_h, preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass
            c.setStrokeColor(colors.HexColor("#999999"))
            c.setLineWidth(0.6)
            c.line(ex + 8, ey + eh * 0.42, ex + ew - 8, ey + eh * 0.42)
            c.setFont("Helvetica-Bold", max(6, fs * 0.85))
            c.drawCentredString(ex + ew / 2, ey + eh * 0.22, sn)
            c.setFont("Helvetica", max(6, fs * 0.7))
            c.drawCentredString(ex + ew / 2, ey + eh * 0.06, st_val)

        elif etype == "line":
            lc = el.get("line_color", "#000000")
            lw_val = max(0.5, el.get("line_width", 2) * sx)
            c.setStrokeColor(colors.HexColor(lc))
            c.setLineWidth(lw_val)
            c.line(ex, ey + eh / 2, ex + ew, ey + eh / 2)

        elif etype == "qrcode":
            qr_data = cert_id or data.get("certificate_id", "") or data.get("membership_id", "SAMPLE")
            verify_base = el.get("verify_url", "")
            # Auto-construct verify URL if not set: use site_url from data or fallback
            if not verify_base:
                site_base = data.get("_site_url", "")
                if site_base:
                    verify_base = f"{site_base.rstrip('/')}/verify"
            if verify_base:
                qr_url = f"{verify_base}?id={qr_data}"
            else:
                qr_url = qr_data
            try:
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=1)
                qr.add_data(qr_url)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white")
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                qr_img.save(tmp.name)
                qr_size = min(ew, eh)
                qr_x = ex + (ew - qr_size) / 2
                qr_y = ey + (eh - qr_size) / 2
                c.drawImage(tmp.name, qr_x, qr_y, width=qr_size, height=qr_size, preserveAspectRatio=True)
                # Draw cert ID text below QR
                c.setFillColor(colors.HexColor(clr))
                c.setFont("Helvetica", max(5, fs * 0.55))
                c.drawCentredString(ex + ew / 2, qr_y - max(5, fs * 0.6), qr_data)
                os.unlink(tmp.name)
            except Exception:
                pass

    c.showPage()
    c.save()
    buf.seek(0)
    return buf.getvalue()


@api_router.get("/admin/certificate-templates")
async def list_cert_templates(admin=Depends(get_current_admin)):
    return await db.certificate_templates.find({}, {"_id": 0}).sort("updated_at", -1).to_list(100)


@api_router.post("/admin/certificate-templates")
async def create_cert_template(request: Request, admin=Depends(get_current_admin)):
    data = await request.json()
    orient = data.get("orientation", "landscape")
    tpl = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", "Untitled Template"),
        "type": data.get("type", "custom"),
        "orientation": orient,
        "page_width": 1000 if orient == "landscape" else 707,
        "page_height": 707 if orient == "landscape" else 1000,
        "background_color": data.get("background_color", "#ffffff"),
        "background_image_url": data.get("background_image_url", ""),
        "linked_membership_type": data.get("linked_membership_type", ""),
        "linked_membership_types": data.get("linked_membership_types", []),
        "elements": data.get("elements", []),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.certificate_templates.insert_one(tpl)
    tpl.pop("_id", None)
    return tpl


@api_router.get("/admin/certificate-templates/{tpl_id}")
async def get_cert_template(tpl_id: str, admin=Depends(get_current_admin)):
    t = await db.certificate_templates.find_one({"id": tpl_id}, {"_id": 0})
    if not t:
        raise HTTPException(404, "Template not found")
    return t


@api_router.put("/admin/certificate-templates/{tpl_id}")
async def update_cert_template(tpl_id: str, request: Request, admin=Depends(get_current_admin)):
    data = await request.json()
    data["updated_at"] = now_iso()
    data.pop("id", None)
    data.pop("_id", None)
    data.pop("created_at", None)
    r = await db.certificate_templates.update_one({"id": tpl_id}, {"$set": data})
    if r.matched_count == 0:
        raise HTTPException(404, "Template not found")
    return await db.certificate_templates.find_one({"id": tpl_id}, {"_id": 0})


@api_router.delete("/admin/certificate-templates/{tpl_id}")
async def delete_cert_template(tpl_id: str, admin=Depends(get_current_admin)):
    r = await db.certificate_templates.delete_one({"id": tpl_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Template not found")
    return {"message": "Template deleted"}


@api_router.post("/admin/certificate-templates/{tpl_id}/clone")
async def clone_cert_template(tpl_id: str, admin=Depends(get_current_admin)):
    t = await db.certificate_templates.find_one({"id": tpl_id}, {"_id": 0})
    if not t:
        raise HTTPException(404, "Template not found")
    t["id"] = str(uuid.uuid4())
    t["name"] = t.get("name", "") + " (Copy)"
    t["created_at"] = now_iso()
    t["updated_at"] = now_iso()
    await db.certificate_templates.insert_one(t)
    t.pop("_id", None)
    return t


@api_router.put("/admin/certificate-templates/{tpl_id}/link-plan")
async def link_template_to_plan(tpl_id: str, request: Request, admin=Depends(get_current_admin)):
    data = await request.json()
    mtype = data.get("membership_type", "")
    action = data.get("action", "toggle")  # "toggle", "add", "remove", "set"

    tpl = await db.certificate_templates.find_one({"id": tpl_id}, {"_id": 0})
    if not tpl:
        raise HTTPException(404, "Template not found")

    current_types = tpl.get("linked_membership_types", [])
    # Migrate from old single field if needed
    if not current_types and tpl.get("linked_membership_type"):
        current_types = [tpl["linked_membership_type"]]

    if not mtype:
        # Clear all links
        new_types = []
    elif action == "remove":
        new_types = [t for t in current_types if t != mtype]
    elif action == "add":
        new_types = list(set(current_types + [mtype]))
    else:
        # toggle
        if mtype in current_types:
            new_types = [t for t in current_types if t != mtype]
        else:
            new_types = list(set(current_types + [mtype]))

    await db.certificate_templates.update_one(
        {"id": tpl_id},
        {"$set": {
            "linked_membership_types": new_types,
            "linked_membership_type": new_types[0] if new_types else "",
            "updated_at": now_iso()
        }}
    )
    return {"message": f"Template linked to {', '.join(new_types)}" if new_types else "Template unlinked", "linked_membership_types": new_types}


@api_router.get("/admin/certificate-templates/by-plan/{membership_type}")
async def get_template_by_plan(membership_type: str, admin=Depends(get_current_admin)):
    # Search in both old and new fields for backward compatibility
    t = await db.certificate_templates.find_one(
        {"$or": [{"linked_membership_types": membership_type}, {"linked_membership_type": membership_type}]},
        {"_id": 0}
    )
    if not t:
        raise HTTPException(404, f"No template linked to {membership_type}")
    return t




@api_router.post("/admin/certificate-templates/{tpl_id}/preview")
async def preview_cert_template(tpl_id: str, request: Request, admin=Depends(get_current_admin)):
    t = await db.certificate_templates.find_one({"id": tpl_id}, {"_id": 0})
    if not t:
        raise HTTPException(404, "Template not found")
    sample = {
        "name": "Dr. Sample Person", "membership_id": "ACD/IDSEA/2026/0001",
        "date": datetime.now().strftime("%d.%m.%Y"), "year": str(datetime.now().year),
        "email": "sample@example.com", "phone": "+91 98765 43210",
        "qualification": "Ph.D. in Dairy Science", "organization": "IDSEA Research Institute",
        "membership_type": "Academic", "state": "Tamil Nadu",
        "event_title": "IDSEA Annual Conference 2026", "event_date": "20-22 March 2026",
        "event_venue": "VCRI, Namakkal", "registration_id": "REG-2026-0001",
        "paper_title": "Advances in Dairy Processing", "specialization": "Dairy Technology",
        "certificate_id": "IDSEA-SAMPLE-0000",
    }
    body = await request.json() if request.headers.get("content-length") and int(request.headers.get("content-length", 0)) > 0 else {}
    if body.get("data"):
        sample.update(body["data"])
    site_url = await _get_site_url()
    sample["_site_url"] = site_url
    pdf = generate_template_pdf(t, sample, cert_id="IDSEA-SAMPLE-0000")
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="preview.pdf"'})


def _gen_cert_id(prefix="CERT"):
    import random, string
    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    return f"IDSEA-{prefix}-{code}"


_MTYPE_LABELS = {"academic": "Academic", "entrepreneur": "Entrepreneur", "corporate": "Corporate", "international": "International", "life": "Life", "student": "Student"}

def _membership_label(raw_type: str) -> str:
    """Convert membership type key to clean label (e.g., 'academic' -> 'Academic', 'Academic Member' -> 'Academic')"""
    if not raw_type:
        return ""
    # Remove trailing ' Member' if present
    label = raw_type.replace(" Member", "").replace(" member", "").strip()
    # Look up from known labels
    return _MTYPE_LABELS.get(label.lower(), label.capitalize())


async def _get_site_url() -> str:
    """Get site URL from CMS settings for QR code generation"""
    cms = await db.cms_settings.find_one({}, {"_id": 0, "site_url": 1})
    return (cms or {}).get("site_url", "") or ""


@api_router.post("/admin/certificate-templates/{tpl_id}/generate-member/{member_id}")
async def gen_cert_member(tpl_id: str, member_id: str, admin=Depends(get_current_admin)):
    t = await db.certificate_templates.find_one({"id": tpl_id}, {"_id": 0})
    m = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not t or not m:
        raise HTTPException(404, "Not found")
    cert_id = _gen_cert_id("MEM")
    site_url = await _get_site_url()
    data = {
        "name": f"{m.get('prefix','')} {m.get('name','')}".strip(),
        "membership_id": m.get("membership_id", ""),
        "date": datetime.now().strftime("%d.%m.%Y"), "year": str(datetime.now().year),
        "email": m.get("email", ""), "phone": m.get("phone", ""),
        "qualification": m.get("qualification", ""), "specialization": m.get("specialization", ""),
        "organization": m.get("organization", ""), "membership_type": _membership_label(m.get("membership_type", "")),
        "state": m.get("state", ""), "country": m.get("country", "India"),
        "certificate_id": cert_id, "_site_url": site_url,
    }
    # Store certificate record (metadata only, no PDF)
    await db.certificate_records.insert_one({
        "cert_id": cert_id, "type": "membership", "template_id": tpl_id,
        "member_id": member_id, "recipient_name": data["name"],
        "membership_id": data.get("membership_id", ""),
        "membership_type": data.get("membership_type", ""),
        "issued_date": now_iso(), "data": data,
    })
    pdf = generate_template_pdf(t, data, cert_id=cert_id)
    fn = f"certificate_{m.get('name','member').replace(' ','_')}.pdf"
    # WhatsApp: Certificate issued notification with PDF
    if m.get("phone"):
        wa_msg = await _render_wa_template("certificate_issued", {
            "name": data["name"], "certificate_type": "Membership",
            "certificate_id": cert_id,
        })
        if wa_msg:
            await send_whatsapp_document_bytes(m["phone"], pdf, fn, wa_msg)
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@api_router.post("/admin/certificate-templates/{tpl_id}/generate-event/{event_id}")
async def gen_cert_event_bulk(tpl_id: str, event_id: str, admin=Depends(get_current_admin)):
    import zipfile as zf
    t = await db.certificate_templates.find_one({"id": tpl_id}, {"_id": 0})
    ev = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not t or not ev:
        raise HTTPException(404, "Not found")
    regs = await db.event_registrations.find({"event_id": event_id, "payment_status": "paid"}, {"_id": 0}).to_list(5000)
    site_url = await _get_site_url()
    zbuf = io.BytesIO()
    with zf.ZipFile(zbuf, 'w', zf.ZIP_DEFLATED) as z:
        for r in regs:
            cert_id = _gen_cert_id("EVT")
            data = {
                "name": r.get("name", ""), "email": r.get("email", ""),
                "phone": r.get("phone", ""), "qualification": r.get("qualification", ""),
                "organization": r.get("organization", ""), "state": r.get("state", ""),
                "event_title": ev.get("title", ""), "event_venue": ev.get("venue", ""),
                "event_date": f"{ev.get('date','')} to {ev.get('end_date','')}",
                "registration_id": r.get("id", ""),
                "date": datetime.now().strftime("%d.%m.%Y"), "year": str(datetime.now().year),
                "certificate_id": cert_id, "_site_url": site_url,
            }
            # Store record
            await db.certificate_records.insert_one({
                "cert_id": cert_id, "type": "event", "template_id": tpl_id,
                "event_id": event_id, "registration_id": r.get("id", ""),
                "recipient_name": data["name"], "event_title": ev.get("title", ""),
                "issued_date": now_iso(), "data": data,
            })
            pdf = generate_template_pdf(t, data, cert_id=cert_id)
            sn = r.get("name", "p").replace(" ", "_")[:40]
            z.writestr(f"certificate_{sn}.pdf", pdf)
    zbuf.seek(0)
    fn = f"certificates_{ev.get('title','event').replace(' ','_')[:30]}.zip"
    return StreamingResponse(zbuf, media_type="application/zip",
                             headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# =================== PUBLIC CERTIFICATE VERIFICATION ===================

@api_router.get("/public/certificates/verify/{cert_id}")
async def verify_certificate(cert_id: str):
    rec = await db.certificate_records.find_one({"cert_id": cert_id}, {"_id": 0})
    if not rec:
        return {"verified": False, "message": "Certificate not found"}
    return {
        "verified": True,
        "cert_id": rec["cert_id"],
        "type": rec.get("type", ""),
        "recipient_name": rec.get("recipient_name", ""),
        "membership_id": rec.get("membership_id", ""),
        "membership_type": rec.get("membership_type", ""),
        "event_title": rec.get("event_title", ""),
        "issued_date": rec.get("issued_date", ""),
    }


@api_router.get("/public/certificates/download/{cert_id}")
async def download_certificate(cert_id: str):
    rec = await db.certificate_records.find_one({"cert_id": cert_id}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "Certificate not found")
    tpl = await db.certificate_templates.find_one({"id": rec.get("template_id")}, {"_id": 0})
    if not tpl:
        raise HTTPException(404, "Certificate template not found")
    data = rec.get("data", {})
    data["certificate_id"] = cert_id
    site_url = await _get_site_url()
    data["_site_url"] = site_url
    pdf = generate_template_pdf(tpl, data, cert_id=cert_id)
    fn = f"certificate_{cert_id}.pdf"
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fn}"'})



# =================== MEMBER EXPORT ===================

@api_router.get("/admin/members/export/excel")
async def export_members_excel(status: Optional[str] = None, membership_type: Optional[str] = None, admin=Depends(get_current_admin)):
    query = {}
    if status:
        query["status"] = status
    if membership_type:
        query["membership_type"] = membership_type
    members = await db.members.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = ["S.No", "Member ID", "Name", "Email", "Phone", "Qualification", "Organization", "State", "Membership Type", "Status", "Payment Status", "Join Date"]
    header_fill = openpyxl.styles.PatternFill(start_color="0C3C60", end_color="0C3C60", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, m in enumerate(members, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=m.get("membership_id", ""))
        ws.cell(row=i + 1, column=3, value=m.get("name", ""))
        ws.cell(row=i + 1, column=4, value=m.get("email", ""))
        ws.cell(row=i + 1, column=5, value=m.get("phone", ""))
        ws.cell(row=i + 1, column=6, value=m.get("qualification", ""))
        ws.cell(row=i + 1, column=7, value=m.get("organization", ""))
        ws.cell(row=i + 1, column=8, value=m.get("state", ""))
        ws.cell(row=i + 1, column=9, value=m.get("membership_type", ""))
        ws.cell(row=i + 1, column=10, value=m.get("status", ""))
        ws.cell(row=i + 1, column=11, value=m.get("payment_status", ""))
        ws.cell(row=i + 1, column=12, value=m.get("join_date", ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": "attachment; filename=IDSEA_Members.xlsx"})


@api_router.get("/admin/members/export/pdf")
async def export_members_pdf(status: Optional[str] = None, membership_type: Optional[str] = None, admin=Depends(get_current_admin)):
    query = {}
    if status:
        query["status"] = status
    if membership_type:
        query["membership_type"] = membership_type
    members = await db.members.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    elements = []

    title_style = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=16, alignment=TA_CENTER, spaceAfter=6, textColor=colors.HexColor('#0c3c60'))
    subtitle_style = ParagraphStyle('subtitle', fontName='Helvetica', fontSize=10, alignment=TA_CENTER, spaceAfter=16, textColor=colors.HexColor('#6b7280'))
    elements.append(Paragraph("IDSEA - Member Directory", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%d %B %Y')} | Total: {len(members)} members", subtitle_style))

    data = [["#", "Member ID", "Name", "Email", "Organization", "State", "Type", "Status"]]
    for i, m in enumerate(members, 1):
        data.append([
            str(i), m.get("membership_id", "")[:12], m.get("name", "")[:25],
            m.get("email", "")[:28], m.get("organization", "")[:30],
            m.get("state", "")[:15], m.get("membership_type", "")[:12],
            m.get("status", "")
        ])

    col_widths = [30, 70, 120, 140, 150, 80, 70, 60]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c3c60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=IDSEA_Members.pdf"})


# =================== FILE UPLOAD ===================

ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), admin=Depends(get_current_admin)):
    try:
        ext = Path(file.filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=400, detail=f"File type {ext} not allowed")
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="File too large (max 10MB)")
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        unique_name = f"{uuid.uuid4().hex[:12]}{ext}"
        file_path = UPLOAD_DIR / unique_name
        with open(file_path, 'wb') as f:
            f.write(content)
        if not file_path.exists():
            raise HTTPException(status_code=500, detail="File verification failed")
        file_record = {
            "id": str(uuid.uuid4()),
            "original_name": file.filename,
            "stored_name": unique_name,
            "file_url": f"/api/uploads/{unique_name}",
            "file_type": ext,
            "file_size": len(content),
            "uploaded_by": admin["email"],
            "uploaded_at": now_iso()
        }
        await db.uploads.insert_one(file_record)
        logging.info(f"Upload OK: {file_path} size={len(content)} exists={file_path.exists()}")
        return {"file_url": file_record["file_url"], "original_name": file.filename, "id": file_record["id"]}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Admin upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@api_router.get("/uploads/{filename}")
async def serve_upload(filename: str):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        logging.error(f"Upload not found: {file_path} (UPLOAD_DIR={UPLOAD_DIR}, exists={UPLOAD_DIR.exists()}, files={len(list(UPLOAD_DIR.iterdir())) if UPLOAD_DIR.exists() else 0})")
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")
    ext = file_path.suffix.lower()
    content_types = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
        '.gif': 'image/gif', '.webp': 'image/webp', '.pdf': 'application/pdf',
        '.doc': 'application/msword', '.docx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.document',
        '.xls': 'application/vnd.ms-excel', '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    }
    ct = content_types.get(ext, 'application/octet-stream')
    headers = {"Cache-Control": "public, max-age=2592000", "Access-Control-Allow-Origin": "*"}
    return StreamingResponse(open(file_path, 'rb'), media_type=ct, headers=headers)


@api_router.get("/admin/upload-debug")
async def upload_debug(admin=Depends(get_current_admin)):
    """Debug endpoint to check upload directory on VPS"""
    files = list(UPLOAD_DIR.iterdir()) if UPLOAD_DIR.exists() else []
    return {
        "upload_dir": str(UPLOAD_DIR),
        "exists": UPLOAD_DIR.exists(),
        "writable": os.access(str(UPLOAD_DIR), os.W_OK) if UPLOAD_DIR.exists() else False,
        "file_count": len(files),
        "recent_files": [f.name for f in sorted(files, key=lambda x: x.stat().st_mtime, reverse=True)[:10]] if files else [],
        "dir_owner": str(UPLOAD_DIR.stat().st_uid) if UPLOAD_DIR.exists() else "N/A",
    }


# =================== SMTP SETTINGS ===================

@api_router.get("/admin/smtp-settings")
async def admin_get_smtp_settings(admin=Depends(get_current_admin)):
    settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if settings:
        settings["smtp_pass"] = "••••••••" if settings.get("smtp_pass") else ""
    return settings or {"smtp_host": SMTP_HOST, "smtp_port": SMTP_PORT_VAL, "smtp_user": SMTP_USER, "smtp_pass": "••••••••" if SMTP_PASS else "", "from_email": FROM_EMAIL, "is_configured": bool(SMTP_HOST and SMTP_USER)}


@api_router.put("/admin/smtp-settings")
async def admin_update_smtp_settings(data: dict, admin=Depends(get_current_admin)):
    if admin.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Only super admin can update SMTP settings")
    settings = {"smtp_host": data.get("smtp_host", ""), "smtp_port": int(data.get("smtp_port", 587)), "smtp_user": data.get("smtp_user", ""), "from_email": data.get("from_email", "noreply@idsea.org"), "updated_at": now_iso()}
    if data.get("smtp_pass") and data["smtp_pass"] != "••••••••":
        settings["smtp_pass"] = data["smtp_pass"]
    else:
        existing = await db.smtp_settings.find_one({}, {"_id": 0})
        if existing:
            settings["smtp_pass"] = existing.get("smtp_pass", "")
    settings["is_configured"] = bool(settings["smtp_host"] and settings["smtp_user"])
    await db.smtp_settings.update_one({}, {"$set": settings}, upsert=True)
    return {"message": "SMTP settings updated", "is_configured": settings["is_configured"]}


@api_router.post("/admin/smtp-settings/test")
async def test_smtp_settings(admin=Depends(get_current_admin)):
    settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("smtp_host"):
        return {"success": False, "message": "SMTP not configured. Please save settings first."}
    success = send_email_smtp([admin["email"]], "IDSEA - SMTP Test Email", "<h2>SMTP Configuration Test</h2><p>This is a test email from IDSEA admin panel. If you received this, your SMTP settings are working correctly!</p>", settings)
    return {"success": success, "message": "Test email sent to " + admin["email"] if success else "Failed to send. Check SMTP credentials."}


# =================== EMAIL TEMPLATES ===================

@api_router.get("/admin/email-templates")
async def admin_get_email_templates(admin=Depends(get_current_admin)):
    templates = await db.email_templates.find({}, {"_id": 0}).to_list(50)
    existing_keys = {t["key"] for t in templates}
    # Merge with defaults for any missing templates
    for key, default in DEFAULT_EMAIL_TEMPLATES.items():
        if key not in existing_keys:
            templates.append({"key": key, **default, "is_default": True, "is_custom": False})
        else:
            for t in templates:
                if t["key"] == key:
                    t["is_default"] = t.get("updated_at") is None
                    t["is_custom"] = False
    # Mark custom templates
    for t in templates:
        if t["key"] not in DEFAULT_EMAIL_TEMPLATES:
            t["is_custom"] = True
            t["is_default"] = False
    default_keys = list(DEFAULT_EMAIL_TEMPLATES.keys())
    return sorted(templates, key=lambda t: default_keys.index(t["key"]) if t["key"] in DEFAULT_EMAIL_TEMPLATES else 99)


@api_router.get("/admin/email-templates/{template_key}")
async def admin_get_email_template(template_key: str, admin=Depends(get_current_admin)):
    template = await db.email_templates.find_one({"key": template_key}, {"_id": 0})
    if not template:
        default = DEFAULT_EMAIL_TEMPLATES.get(template_key)
        if default:
            return {"key": template_key, **default, "is_default": True}
        raise HTTPException(status_code=404, detail="Template not found")
    template["is_default"] = False
    return template


@api_router.put("/admin/email-templates/{template_key}")
async def admin_update_email_template(template_key: str, data: dict, admin=Depends(get_current_admin)):
    is_default = template_key in DEFAULT_EMAIL_TEMPLATES
    if is_default:
        update = {
            "key": template_key,
            "name": data.get("name", DEFAULT_EMAIL_TEMPLATES[template_key]["name"]),
            "subject": data.get("subject", ""),
            "body": data.get("body", ""),
            "description": data.get("description", ""),
            "variables": data.get("variables", DEFAULT_EMAIL_TEMPLATES[template_key]["variables"]),
            "updated_at": now_iso(),
            "updated_by": admin["email"]
        }
    else:
        existing = await db.email_templates.find_one({"key": template_key})
        if not existing:
            raise HTTPException(status_code=404, detail="Template not found")
        update = {
            "name": data.get("name", existing.get("name", "")),
            "subject": data.get("subject", ""),
            "body": data.get("body", ""),
            "description": data.get("description", ""),
            "variables": data.get("variables", existing.get("variables", [])),
            "updated_at": now_iso(),
            "updated_by": admin["email"]
        }
    await db.email_templates.update_one({"key": template_key}, {"$set": update}, upsert=True)
    return {"message": f"Template '{template_key}' updated"}


@api_router.post("/admin/email-templates/{template_key}/reset")
async def admin_reset_email_template(template_key: str, admin=Depends(get_current_admin)):
    if template_key not in DEFAULT_EMAIL_TEMPLATES:
        raise HTTPException(status_code=400, detail="Unknown template key")
    await db.email_templates.delete_one({"key": template_key})
    return {"message": f"Template '{template_key}' reset to default"}


@api_router.post("/admin/email-templates/{template_key}/preview")
async def admin_preview_email_template(template_key: str, admin=Depends(get_current_admin)):
    template = await db.email_templates.find_one({"key": template_key}, {"_id": 0})
    if not template:
        template = DEFAULT_EMAIL_TEMPLATES.get(template_key)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    sample_vars = {
        "member_name": "Dr. John Doe",
        "email": "john.doe@example.com",
        "phone": "+91 98765 43210",
        "qualification": "Ph.D. (Dairy Science)",
        "organization": "National Dairy Research Institute",
        "membership_type": "Academic",
        "membership_id": "IDSEA20260001",
        "state": "Tamil Nadu",
        "payment_status": "Pending",
        "application_date": datetime.now().strftime("%d %B %Y"),
        "join_date": datetime.now().strftime("%d %B %Y"),
        "event_title": "International Dairy Conference 2026",
        "event_date": "15 March 2026",
        "event_end_date": "17 March 2026",
        "event_venue": "VCRI Namakkal, Tamil Nadu",
        "event_description": "A premier conference bringing together dairy scientists and entrepreneurs.",
        "registration_fee": "2,500",
        "speaker_details": "Dr. A. Elango, Dr. G. Kumaresan",
        "venue_map_link": "",
        "rejection_reason": "Incomplete documentation submitted",
        "participant_name": "Dr. John Doe",
        "participant_email": "john.doe@example.com",
        "registration_id": "REG-20260001",
        "registration_type": "Academic",
        "amount_paid": "Rs. 2,500",
        "certificate_type": "Participation",
        "certificate_id": "CERT-20260001",
        "verify_url": "https://idsea.org/verify/CERT-20260001",
    }
    rendered_subject = render_template(template.get("subject", ""), sample_vars)
    rendered_body = render_template(template.get("body", ""), sample_vars)
    return {"subject": rendered_subject, "body": rendered_body, "variables": template.get("variables", [])}


# =================== CUSTOM TEMPLATE CREATION ===================

@api_router.post("/admin/email-templates")
async def admin_create_custom_template(data: dict, admin=Depends(get_current_admin)):
    key = data.get("key", "").strip().lower().replace(" ", "_")
    if not key or not data.get("name") or not data.get("subject"):
        raise HTTPException(status_code=400, detail="Key, name, and subject required")
    if key in DEFAULT_EMAIL_TEMPLATES:
        raise HTTPException(status_code=400, detail="Cannot create template with reserved key")
    existing = await db.email_templates.find_one({"key": key})
    if existing:
        raise HTTPException(status_code=400, detail="Template key already exists")
    template = {
        "key": key,
        "name": data.get("name", ""),
        "subject": data.get("subject", ""),
        "body": data.get("body", DEFAULT_EMAIL_TEMPLATES["registration_submitted"]["body"]),
        "description": data.get("description", ""),
        "variables": data.get("variables", []),
        "is_custom": True,
        "created_at": now_iso(),
        "created_by": admin["email"]
    }
    await db.email_templates.insert_one(template)
    return {"message": f"Template '{key}' created", "key": key}


@api_router.delete("/admin/email-templates/{template_key}")
async def admin_delete_custom_template(template_key: str, admin=Depends(get_current_admin)):
    if template_key in DEFAULT_EMAIL_TEMPLATES:
        raise HTTPException(status_code=400, detail="Cannot delete default template. Use reset instead.")
    result = await db.email_templates.delete_one({"key": template_key})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": f"Template '{template_key}' deleted"}


# =================== EMAIL QUEUE & CAMPAIGNS ===================

EMAIL_BATCH_SIZE = 50
EMAIL_BATCH_INTERVAL = 300  # 5 minutes in seconds
_email_scheduler_running = False


async def process_email_queue():
    """Background task: process pending emails in batches of 50 every 5 minutes"""
    global _email_scheduler_running
    _email_scheduler_running = True
    while True:
        try:
            pending = await db.email_queue.find(
                {"status": "pending"}
            ).sort("created_at", 1).limit(EMAIL_BATCH_SIZE).to_list(EMAIL_BATCH_SIZE)

            if pending:
                smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})
                sent_count = 0
                failed_count = 0
                for email_item in pending:
                    try:
                        await db.email_queue.update_one(
                            {"id": email_item["id"]},
                            {"$set": {"status": "processing"}}
                        )
                        success = send_email_smtp(
                            [email_item["recipient_email"]],
                            email_item["subject"],
                            email_item["body"],
                            smtp_settings
                        )
                        if success:
                            await db.email_queue.update_one(
                                {"id": email_item["id"]},
                                {"$set": {"status": "sent", "sent_at": now_iso()}}
                            )
                            sent_count += 1
                        else:
                            await db.email_queue.update_one(
                                {"id": email_item["id"]},
                                {"$set": {"status": "failed", "error": "SMTP send returned False", "retry_count": email_item.get("retry_count", 0) + 1}}
                            )
                            failed_count += 1
                    except Exception as e:
                        await db.email_queue.update_one(
                            {"id": email_item["id"]},
                            {"$set": {"status": "failed", "error": str(e)[:200], "retry_count": email_item.get("retry_count", 0) + 1}}
                        )
                        failed_count += 1

                logging.info(f"Email queue batch: sent={sent_count}, failed={failed_count}, remaining pending")

        except Exception as e:
            logging.error(f"Email queue processor error: {e}")

        await asyncio.sleep(EMAIL_BATCH_INTERVAL)


async def queue_email(recipient_email: str, recipient_name: str, subject: str, body: str, campaign_id: str = "", campaign_name: str = "", template_key: str = ""):
    """Add an email to the queue for batch processing"""
    await db.email_queue.insert_one({
        "id": str(uuid.uuid4()),
        "recipient_email": recipient_email,
        "recipient_name": recipient_name,
        "subject": subject,
        "body": body,
        "template_key": template_key,
        "campaign_id": campaign_id,
        "campaign_name": campaign_name,
        "status": "pending",
        "error": "",
        "retry_count": 0,
        "created_at": now_iso(),
        "scheduled_at": now_iso(),
        "sent_at": ""
    })


async def queue_bulk_emails(template_key: str, recipients: list, variables_list: list, campaign_name: str = ""):
    """Queue bulk emails using a template for batch sending"""
    template = await db.email_templates.find_one({"key": template_key}, {"_id": 0})
    if not template:
        template = DEFAULT_EMAIL_TEMPLATES.get(template_key)
    if not template:
        return 0

    campaign_id = str(uuid.uuid4())
    count = 0
    for i, recipient in enumerate(recipients):
        variables = variables_list[i] if i < len(variables_list) else {}
        subject = render_template(template.get("subject", ""), variables)
        body = render_template(template.get("body", ""), variables)
        await queue_email(
            recipient_email=recipient.get("email", ""),
            recipient_name=recipient.get("name", ""),
            subject=subject,
            body=body,
            campaign_id=campaign_id,
            campaign_name=campaign_name or template.get("name", ""),
            template_key=template_key
        )
        count += 1

    await db.email_logs.insert_one({
        "id": str(uuid.uuid4()),
        "subject": campaign_name or template.get("name", ""),
        "body": f"Queued {count} emails for batch sending",
        "recipients_count": count,
        "recipient_group": campaign_name,
        "campaign_id": campaign_id,
        "sent_by": "system",
        "sent_at": now_iso(),
        "status": "queued"
    })
    return count


@api_router.get("/admin/email-queue")
async def admin_get_email_queue(status: Optional[str] = None, limit: int = 100, admin=Depends(get_current_admin)):
    query = {}
    if status:
        query["status"] = status
    items = await db.email_queue.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return items


@api_router.get("/admin/email-queue/stats")
async def admin_email_queue_stats(admin=Depends(get_current_admin)):
    pipeline = [{"$group": {"_id": "$status", "count": {"$sum": 1}}}]
    queue_stats = await db.email_queue.aggregate(pipeline).to_list(10)
    log_stats = await db.email_logs.aggregate(pipeline).to_list(10)
    queue_counts = {s["_id"]: s["count"] for s in queue_stats if s["_id"]}
    log_counts = {s["_id"]: s["count"] for s in log_stats if s["_id"]}
    return {
        "queue": queue_counts,
        "queue_total": sum(queue_counts.values()),
        "logs": log_counts,
        "logs_total": sum(log_counts.values()),
        "batch_size": EMAIL_BATCH_SIZE,
        "batch_interval_mins": EMAIL_BATCH_INTERVAL // 60,
        "scheduler_running": _email_scheduler_running,
    }


@api_router.post("/admin/email-queue/{email_id}/retry")
async def admin_retry_queued_email(email_id: str, admin=Depends(get_current_admin)):
    result = await db.email_queue.update_one(
        {"id": email_id, "status": "failed"},
        {"$set": {"status": "pending", "error": ""}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Email not found or not in failed state")
    return {"message": "Email re-queued for sending"}


@api_router.post("/admin/email-queue/retry-all-failed")
async def admin_retry_all_failed(admin=Depends(get_current_admin)):
    result = await db.email_queue.update_many(
        {"status": "failed"},
        {"$set": {"status": "pending", "error": ""}}
    )
    return {"message": f"{result.modified_count} failed emails re-queued"}


@api_router.delete("/admin/email-queue/clear-sent")
async def admin_clear_sent_queue(admin=Depends(get_current_admin)):
    result = await db.email_queue.delete_many({"status": "sent"})
    return {"message": f"{result.deleted_count} sent emails cleared"}


@api_router.post("/admin/email-campaign/send")
async def admin_send_campaign(data: dict, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    """Send email campaign to selected recipients using a template"""
    template_key = data.get("template_key")
    recipient_group = data.get("recipient_group", "all_members")
    campaign_name = data.get("campaign_name", "Manual Campaign")
    custom_variables = data.get("custom_variables", {})

    if not template_key:
        raise HTTPException(status_code=400, detail="template_key required")

    # Validate template exists
    tmpl = await db.email_templates.find_one({"key": template_key}, {"_id": 0})
    if not tmpl and template_key not in DEFAULT_EMAIL_TEMPLATES:
        raise HTTPException(status_code=400, detail=f"Template '{template_key}' not found")

    query = {"status": "approved"}
    if recipient_group == "academic":
        query["membership_type"] = "academic"
    elif recipient_group == "entrepreneur":
        query["membership_type"] = "entrepreneur"
    elif recipient_group == "corporate":
        query["membership_type"] = "corporate"

    members = await db.members.find(query, {"_id": 0}).to_list(10000)
    members_with_email = [m for m in members if m.get("email")]

    if not members_with_email:
        raise HTTPException(status_code=400, detail="No recipients found")

    recipients = [{"email": m["email"], "name": m.get("name", "")} for m in members_with_email]
    variables_list = []
    for m in members_with_email:
        v = {
            "member_name": m.get("name", "Member"),
            "email": m.get("email", ""),
            "membership_id": m.get("membership_id", ""),
            "membership_type": m.get("membership_type", ""),
            "phone": m.get("phone", ""),
            **custom_variables,
        }
        variables_list.append(v)

    count = await queue_bulk_emails(template_key, recipients, variables_list, campaign_name)
    return {"message": f"Campaign '{campaign_name}' queued: {count} emails scheduled", "count": count}

@api_router.get("/admin/reports/monthly-growth")
async def admin_monthly_growth(admin=Depends(get_current_admin)):
    pipeline = [
        {"$addFields": {"month": {"$substr": ["$created_at", 0, 7]}}},
        {"$group": {"_id": "$month", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
        {"$limit": 12}
    ]
    member_growth = await db.members.aggregate(pipeline).to_list(12)

    payment_pipeline = [
        {"$match": {"status": "paid"}},
        {"$addFields": {"month": {"$substr": ["$created_at", 0, 7]}}},
        {"$group": {"_id": "$month", "revenue": {"$sum": "$amount"}, "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
        {"$limit": 12}
    ]
    revenue_growth = await db.payments.aggregate(payment_pipeline).to_list(12)

    type_pipeline = [
        {"$match": {"status": "approved"}},
        {"$group": {"_id": "$membership_type", "count": {"$sum": 1}}}
    ]
    type_dist = await db.members.aggregate(type_pipeline).to_list(10)

    status_pipeline = [
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]
    status_dist = await db.members.aggregate(status_pipeline).to_list(10)

    state_pipeline = [
        {"$match": {"status": "approved", "state": {"$nin": ["", None]}}},
        {"$group": {"_id": "$state", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 15}
    ]
    state_dist = await db.members.aggregate(state_pipeline).to_list(15)

    return {
        "member_growth": [{"month": m["_id"], "members": m["count"]} for m in member_growth],
        "revenue_growth": [{"month": r["_id"], "revenue": r["revenue"], "transactions": r["count"]} for r in revenue_growth],
        "type_distribution": [{"type": t["_id"], "count": t["count"]} for t in type_dist if t["_id"]],
        "status_distribution": [{"status": s["_id"], "count": s["count"]} for s in status_dist if s["_id"]],
        "state_distribution": [{"state": s["_id"], "count": s["count"]} for s in state_dist if s["_id"]],
    }


# =================== SLIDER MANAGEMENT ===================

@api_router.get("/public/sliders")
async def get_public_sliders():
    sliders = await db.sliders.find({"is_active": True}, {"_id": 0}).sort("order", 1).to_list(20)
    return sliders


@api_router.get("/admin/sliders")
async def admin_get_sliders(admin=Depends(get_current_admin)):
    return await db.sliders.find({}, {"_id": 0}).sort("order", 1).to_list(50)


@api_router.post("/admin/sliders")
async def admin_create_slider(data: SliderCreate, admin=Depends(get_current_admin)):
    slider = Slider(**data.model_dump())
    await db.sliders.insert_one(slider.model_dump())
    result = slider.model_dump()
    result.pop("_id", None)
    return result


@api_router.put("/admin/sliders/reorder")
async def admin_reorder_sliders(data: dict, admin=Depends(get_current_admin)):
    """Reorder sliders - must be before {slider_id} route"""
    items = data.get("items", [])
    for item in items:
        await db.sliders.update_one({"id": item["id"]}, {"$set": {"order": item["order"]}})
    return {"message": "Sliders reordered"}


@api_router.put("/admin/sliders/{slider_id}")
async def admin_update_slider(slider_id: str, data: SliderCreate, admin=Depends(get_current_admin)):
    update_data = data.model_dump()
    await db.sliders.update_one({"id": slider_id}, {"$set": update_data})
    return {"message": "Slider updated"}


@api_router.delete("/admin/sliders/{slider_id}")
async def admin_delete_slider(slider_id: str, admin=Depends(get_current_admin)):
    await db.sliders.delete_one({"id": slider_id})
    return {"message": "Slider deleted"}




# =================== MEMBERSHIP PLANS ===================

@api_router.get("/admin/membership-plans")
async def admin_get_membership_plans(admin=Depends(get_current_admin)):
    plans = await db.membership_plans.find({}, {"_id": 0}).to_list(100)
    if not plans:
        defaults = [
            {"id": str(uuid.uuid4()), "key": "academic", "label": "Academic", "fee_inr": 3100, "fee_usd": 50, "enabled": True, "description": "For academic professionals"},
            {"id": str(uuid.uuid4()), "key": "entrepreneur", "label": "Entrepreneur", "fee_inr": 5100, "fee_usd": 75, "enabled": True, "description": "For entrepreneurs in dairy industry"},
            {"id": str(uuid.uuid4()), "key": "corporate", "label": "Corporate", "fee_inr": 25100, "fee_usd": 300, "enabled": True, "description": "For corporate organizations"},
            {"id": str(uuid.uuid4()), "key": "international", "label": "International Delegates", "fee_inr": 0, "fee_usd": 100, "enabled": True, "description": "For international delegates"},
        ]
        for d in defaults:
            await db.membership_plans.insert_one(d)
        plans = defaults
    return plans


@api_router.post("/admin/membership-plans")
async def admin_create_membership_plan(data: dict, admin=Depends(get_current_admin)):
    plan = {
        "id": str(uuid.uuid4()),
        "key": data.get("key", "").lower().replace(" ", "_"),
        "label": data.get("label", ""),
        "fee_inr": float(data.get("fee_inr", 0)),
        "fee_usd": float(data.get("fee_usd", 0)),
        "enabled": data.get("enabled", True),
        "description": data.get("description", ""),
        "created_at": now_iso(),
    }
    await db.membership_plans.insert_one(plan)
    plan.pop("_id", None)
    return plan


@api_router.put("/admin/membership-plans/{plan_id}")
async def admin_update_membership_plan(plan_id: str, data: dict, admin=Depends(get_current_admin)):
    update = {k: v for k, v in data.items() if k in ("key", "label", "fee_inr", "fee_usd", "enabled", "description")}
    if "fee_inr" in update:
        update["fee_inr"] = float(update["fee_inr"])
    if "fee_usd" in update:
        update["fee_usd"] = float(update["fee_usd"])
    update["updated_at"] = now_iso()
    await db.membership_plans.update_one({"id": plan_id}, {"$set": update})
    return {"message": "Plan updated"}


@api_router.delete("/admin/membership-plans/{plan_id}")
async def admin_delete_membership_plan(plan_id: str, admin=Depends(get_current_admin)):
    await db.membership_plans.delete_one({"id": plan_id})
    return {"message": "Plan deleted"}


@api_router.get("/public/membership-plans")
async def public_get_membership_plans():
    plans = await db.membership_plans.find({"enabled": True}, {"_id": 0}).to_list(100)
    if not plans:
        plans = [
            {"key": "academic", "label": "Academic", "fee_inr": 3100, "fee_usd": 50, "enabled": True},
            {"key": "entrepreneur", "label": "Entrepreneur", "fee_inr": 5100, "fee_usd": 75, "enabled": True},
            {"key": "corporate", "label": "Corporate", "fee_inr": 25100, "fee_usd": 300, "enabled": True},
            {"key": "international", "label": "International Delegates", "fee_inr": 0, "fee_usd": 100, "enabled": True},
        ]
    return plans


# =================== MEMBERSHIP ID CONFIG ===================

DEFAULT_ID_PREFIXES = {
    "academic": "ACD",
    "entrepreneur": "ENT",
    "corporate": "COP",
    "international": "INT",
}


async def get_membership_prefix(membership_type: str) -> str:
    """Get the configured prefix for a membership type from DB, fallback to defaults"""
    config = await db.membership_id_config.find_one({"type": membership_type}, {"_id": 0})
    if config and config.get("prefix"):
        return config["prefix"]
    return DEFAULT_ID_PREFIXES.get(membership_type, "MEM")


async def generate_membership_id(membership_type: str) -> str:
    """Generate next membership ID using configured prefix — continuous serial"""
    prefix = await get_membership_prefix(membership_type)
    all_members = await db.members.find(
        {"status": "approved", "membership_id": {"$regex": f"^{prefix}/"}},
        {"_id": 0, "membership_id": 1}
    ).to_list(None)
    max_serial = 0
    for m in all_members:
        try:
            s = int(m["membership_id"].split("/")[-1])
            if s > max_serial:
                max_serial = s
        except (ValueError, IndexError):
            pass
    serial = str(max_serial + 1).zfill(4)
    return f"{prefix}/IDSEA/{serial}"


# =================== JOURNAL MANAGEMENT ===================

@api_router.get("/admin/journal/settings")
async def admin_get_journal_settings(admin=Depends(get_current_admin)):
    settings = await db.journal_settings.find_one({}, {"_id": 0})
    if not settings:
        settings = {"coming_soon": True, "journal_name": "Journal of Dairy Science and Enterprise", "abbreviation": "JDSE", "description": "", "cover_image": "", "meta_title": "", "meta_description": "", "meta_keywords": ""}
    return settings


@api_router.put("/admin/journal/settings")
async def admin_update_journal_settings(data: dict, admin=Depends(get_current_admin)):
    data["updated_at"] = now_iso()
    await db.journal_settings.update_one({}, {"$set": data}, upsert=True)
    return {"message": "Journal settings updated"}


@api_router.get("/admin/journal/board")
async def admin_get_journal_board(admin=Depends(get_current_admin)):
    sections = await db.journal_board.find({}, {"_id": 0}).sort("order", 1).to_list(50)
    return sections


@api_router.post("/admin/journal/board")
async def admin_add_journal_section(data: dict, admin=Depends(get_current_admin)):
    section = {
        "id": str(uuid.uuid4()),
        "title": data.get("title", "New Section"),
        "order": data.get("order", 99),
        "members": data.get("members", []),
        "created_at": now_iso(),
    }
    await db.journal_board.insert_one(section)
    return {"message": "Section added", "id": section["id"]}


@api_router.put("/admin/journal/board/{section_id}")
async def admin_update_journal_section(section_id: str, data: dict, admin=Depends(get_current_admin)):
    data.pop("_id", None)
    data["updated_at"] = now_iso()
    await db.journal_board.update_one({"id": section_id}, {"$set": data})
    return {"message": "Section updated"}


@api_router.delete("/admin/journal/board/{section_id}")
async def admin_delete_journal_section(section_id: str, admin=Depends(get_current_admin)):
    await db.journal_board.delete_one({"id": section_id})
    return {"message": "Section deleted"}


@api_router.get("/public/journal")
async def get_public_journal():
    settings = await db.journal_settings.find_one({}, {"_id": 0}) or {"coming_soon": True, "journal_name": "Journal of Dairy Science and Enterprise", "abbreviation": "JDSE"}
    board = await db.journal_board.find({}, {"_id": 0}).sort("order", 1).to_list(50) if not settings.get("coming_soon") else []
    return {"settings": settings, "board": board}


@api_router.get("/admin/members/search")
async def admin_search_members(q: str = "", admin=Depends(get_current_admin)):
    if len(q) < 2:
        return []
    query = {"name": {"$regex": q, "$options": "i"}, "status": "approved"}
    members = await db.members.find(query, {"_id": 0, "id": 1, "name": 1, "qualification": 1, "organization": 1, "email": 1, "phone": 1, "photo_url": 1}).limit(15).to_list(15)
    return members


@api_router.get("/admin/membership-id-config")
async def admin_get_membership_id_config(admin=Depends(get_current_admin)):
    configs = await db.membership_id_config.find({}, {"_id": 0}).to_list(20)
    existing_types = {c["type"] for c in configs}
    for mtype, prefix in DEFAULT_ID_PREFIXES.items():
        if mtype not in existing_types:
            configs.append({"type": mtype, "prefix": prefix, "is_default": True})
        else:
            for c in configs:
                if c["type"] == mtype:
                    c["is_default"] = False
    # Add sample next ID
    for c in configs:
        c["sample_id"] = f"{c['prefix']}/IDSEA/0001"
    return sorted(configs, key=lambda c: list(DEFAULT_ID_PREFIXES.keys()).index(c["type"]) if c["type"] in DEFAULT_ID_PREFIXES else 99)


@api_router.put("/admin/membership-id-config/{membership_type}")
async def admin_update_membership_id_prefix(membership_type: str, data: dict, admin=Depends(get_current_admin)):
    prefix = data.get("prefix", "").strip().upper()
    if not prefix:
        raise HTTPException(status_code=400, detail="Prefix cannot be empty")
    if "/" in prefix or " " in prefix:
        raise HTTPException(status_code=400, detail="Prefix cannot contain / or spaces")
    await db.membership_id_config.update_one(
        {"type": membership_type},
        {"$set": {"type": membership_type, "prefix": prefix, "updated_at": now_iso(), "updated_by": admin["email"]}},
        upsert=True
    )
    return {"message": f"Prefix for '{membership_type}' updated to '{prefix}'", "sample_id": f"{prefix}/IDSEA/0001"}


# Public PDF upload for registration documents
@api_router.post("/public/upload-pdf")
async def public_upload_pdf(file: UploadFile = File(...)):
    try:
        if not file.filename.lower().endswith('.pdf'):
            raise HTTPException(status_code=400, detail="Only PDF files are allowed")
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large. Max 10MB.")
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        ext = os.path.splitext(file.filename)[1]
        fname = f"{uuid.uuid4()}{ext}"
        fpath = UPLOAD_DIR / fname
        with open(fpath, 'wb') as f:
            f.write(content)
        if not fpath.exists():
            raise HTTPException(status_code=500, detail="File verification failed")
        logging.info(f"PDF upload OK: {fpath} size={len(content)}")
        return {"url": f"/api/uploads/{fname}", "file_url": f"/api/uploads/{fname}", "filename": fname}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"PDF upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


# =================== WHATSAPP (AK NEXUS) SERVICE ===================


# WhatsApp Server Integration (whatsapp-server-4)


WHATSAPP_SERVER_URL = os.environ.get("WHATSAPP_SERVER_URL", "")


async def get_whatsapp_settings():
    settings = await db.whatsapp_settings.find_one({}, {"_id": 0})
    return settings or {}


async def wa_server_post(endpoint: str, body: dict, is_json: bool = True):
    """Make POST request to WhatsApp Server API with Bearer auth."""
    settings = await get_whatsapp_settings()
    token = settings.get("api_key", "")
    server_url = settings.get("server_url", "") or WHATSAPP_SERVER_URL
    if not token or not server_url:
        return {"ok": False, "error": "WhatsApp server URL or API key not configured"}
    url = f"{server_url.rstrip('/')}/{endpoint.lstrip('/')}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    async with httpx.AsyncClient(timeout=30) as client:
        try:
            resp = await client.post(url, json=body, headers=headers)
            logging.info(f"WhatsApp Server {endpoint} -> HTTP {resp.status_code}")
            if resp.status_code == 401:
                return {"ok": False, "error": "Authentication failed. Check API key."}
            if resp.status_code == 429:
                return {"ok": False, "error": "Rate limited. Try again later."}
            try:
                data = resp.json()
            except Exception:
                return {"ok": False, "error": f"Invalid response: {resp.text[:200]}"}
            # Normalize: if response has 'detail' (error), convert to standard format
            if resp.status_code >= 400 or "detail" in data:
                return {"ok": False, "error": data.get("detail", data.get("error", f"HTTP {resp.status_code}"))}
            # Ensure ok field exists
            if "ok" not in data:
                data["ok"] = resp.status_code < 300
            return data
        except httpx.TimeoutException:
            return {"ok": False, "error": "Request timed out"}
        except Exception as e:
            logging.error(f"WhatsApp Server API error: {e}")
            return {"ok": False, "error": str(e)}


async def wa_server_get(endpoint: str):
    """Make GET request to WhatsApp Server API."""
    settings = await get_whatsapp_settings()
    token = settings.get("api_key", "")
    server_url = settings.get("server_url", "") or WHATSAPP_SERVER_URL
    if not token or not server_url:
        return {"ok": False, "error": "WhatsApp server URL or API key not configured"}
    url = f"{server_url.rstrip('/')}/{endpoint.lstrip('/')}"
    headers = {"Authorization": f"Bearer {token}"}
    async with httpx.AsyncClient(timeout=30) as client:
        try:
            resp = await client.get(url, headers=headers)
            if resp.status_code == 401:
                return {"ok": False, "error": "Auth failed. Check API key."}
            try:
                data = resp.json()
            except Exception:
                return {"ok": False, "error": resp.text[:200]}
            if resp.status_code >= 400 or "detail" in data:
                return {"ok": False, "error": data.get("detail", data.get("error", f"HTTP {resp.status_code}"))}
            if "ok" not in data:
                data["ok"] = resp.status_code < 300
            return data
        except Exception as e:
            return {"ok": False, "error": str(e)}


async def send_whatsapp_message(phone: str, message: str, session_id: str = None):
    """Send a WhatsApp text message via WhatsApp Server /api/v1/send/text."""
    settings = await get_whatsapp_settings()
    if not settings.get("enabled", False):
        return False
    sid = session_id or settings.get("session_id", "primary")
    phone = phone.strip().replace("+", "").replace(" ", "").replace("-", "")
    if len(phone) == 10:
        phone = "91" + phone
    body = {"session_id": sid, "to": phone, "text": message}
    result = await wa_server_post("api/v1/send/text", body)
    is_success = result.get("ok", False) is True
    await db.whatsapp_logs.insert_one({
        "id": str(uuid.uuid4()), "phone": phone, "message": message[:200],
        "session_id": sid, "msg_type": "text",
        "status": "sent" if is_success else "failed",
        "error": "" if is_success else result.get("error", ""),
        "response": str(result)[:500], "sent_at": now_iso()
    })
    if not is_success:
        logging.warning(f"WhatsApp text to {phone}: {result.get('error', 'unknown error')}")
    return is_success


async def send_whatsapp_image(phone: str, media_url: str, caption: str = "", session_id: str = None):
    """Send image via WhatsApp Server /api/v1/send/media."""
    settings = await get_whatsapp_settings()
    if not settings.get("enabled", False):
        return False
    sid = session_id or settings.get("session_id", "primary")
    phone = phone.strip().replace("+", "").replace(" ", "").replace("-", "")
    if len(phone) == 10:
        phone = "91" + phone
    # Download file and send via multipart
    server_url = settings.get("server_url", "") or WHATSAPP_SERVER_URL
    token = settings.get("api_key", "")
    if not server_url or not token:
        return False
    url = f"{server_url.rstrip('/')}/api/v1/send/media"
    async with httpx.AsyncClient(timeout=60) as client:
        try:
            # Download the image first
            img_resp = await client.get(media_url)
            if img_resp.status_code != 200:
                logging.warning(f"Failed to download image from {media_url}")
                return False
            files = {"file": ("image.jpg", img_resp.content, "image/jpeg")}
            data = {"session_id": sid, "to": phone, "media_type": "image"}
            if caption:
                data["caption"] = caption
            resp = await client.post(url, headers={"Authorization": f"Bearer {token}"}, files=files, data=data)
            result = resp.json() if resp.status_code == 200 else {"ok": False}
        except Exception as e:
            logging.error(f"WhatsApp image send error: {e}")
            result = {"ok": False, "error": str(e)}
    is_success = result.get("ok", False)
    await db.whatsapp_logs.insert_one({
        "id": str(uuid.uuid4()), "phone": phone, "message": caption[:200],
        "session_id": sid, "msg_type": "image", "media_url": media_url,
        "status": "sent" if is_success else "failed",
        "response": str(result)[:500], "sent_at": now_iso()
    })
    return is_success


async def send_whatsapp_document(phone: str, media_url: str, caption: str = "", session_id: str = None):
    """Send document via WhatsApp Server /api/v1/send/media."""
    settings = await get_whatsapp_settings()
    if not settings.get("enabled", False):
        return False
    sid = session_id or settings.get("session_id", "primary")
    phone = phone.strip().replace("+", "").replace(" ", "").replace("-", "")
    if len(phone) == 10:
        phone = "91" + phone
    server_url = settings.get("server_url", "") or WHATSAPP_SERVER_URL
    token = settings.get("api_key", "")
    if not server_url or not token:
        return False
    url = f"{server_url.rstrip('/')}/api/v1/send/media"
    async with httpx.AsyncClient(timeout=60) as client:
        try:
            doc_resp = await client.get(media_url)
            if doc_resp.status_code != 200:
                return False
            files = {"file": ("document.pdf", doc_resp.content, "application/pdf")}
            data = {"session_id": sid, "to": phone, "media_type": "document"}
            if caption:
                data["caption"] = caption
            resp = await client.post(url, headers={"Authorization": f"Bearer {token}"}, files=files, data=data)
            result = resp.json() if resp.status_code == 200 else {"ok": False}
        except Exception as e:
            logging.error(f"WhatsApp document send error: {e}")
            result = {"ok": False, "error": str(e)}
    is_success = result.get("ok", False) is True
    await db.whatsapp_logs.insert_one({
        "id": str(uuid.uuid4()), "phone": phone, "message": caption[:200],
        "session_id": sid, "msg_type": "document", "media_url": media_url,
        "status": "sent" if is_success else "failed",
        "response": str(result)[:500], "sent_at": now_iso()
    })
    return is_success


async def send_whatsapp_document_bytes(phone: str, file_bytes: bytes, filename: str, caption: str = "", session_id: str = None):
    """Send document from raw bytes via WhatsApp Server /api/v1/send/media."""
    settings = await get_whatsapp_settings()
    if not settings.get("enabled", False):
        return False
    sid = session_id or settings.get("session_id", "primary")
    phone = phone.strip().replace("+", "").replace(" ", "").replace("-", "")
    if len(phone) == 10:
        phone = "91" + phone
    server_url = settings.get("server_url", "") or WHATSAPP_SERVER_URL
    token = settings.get("api_key", "")
    if not server_url or not token:
        return False
    url = f"{server_url.rstrip('/')}/api/v1/send/media"
    async with httpx.AsyncClient(timeout=60) as client:
        try:
            files = {"file": (filename, file_bytes, "application/pdf")}
            data = {"session_id": sid, "to": phone, "media_type": "document"}
            if caption:
                data["caption"] = caption
            resp = await client.post(url, headers={"Authorization": f"Bearer {token}"}, files=files, data=data)
            result = resp.json() if resp.status_code < 300 else {"ok": False, "error": f"HTTP {resp.status_code}"}
        except Exception as e:
            logging.error(f"WhatsApp document bytes send error: {e}")
            result = {"ok": False, "error": str(e)}
    is_success = result.get("ok", False) is True
    await db.whatsapp_logs.insert_one({
        "id": str(uuid.uuid4()), "phone": phone, "message": caption[:200],
        "session_id": sid, "msg_type": "document",
        "status": "sent" if is_success else "failed",
        "response": str(result)[:500], "sent_at": now_iso()
    })
    return is_success


async def _render_wa_template(template_key: str, variables: dict) -> str:
    """Render a WhatsApp template message with variables. Returns empty string if template disabled."""
    db_tpl = await db.whatsapp_templates.find_one({"key": template_key}, {"_id": 0})
    if db_tpl and db_tpl.get("enabled", True):
        text = db_tpl.get("message", "")
    else:
        defaults = {d["key"]: d["message"] for d in WA_DEFAULT_TEMPLATES}
        text = defaults.get(template_key, "")
    if not text:
        return ""
    try:
        return text.format(**variables)
    except KeyError:
        for k, v in variables.items():
            text = text.replace("{" + k + "}", str(v))
        return text


async def send_whatsapp_notification(phone: str, notification_type: str, variables: dict):
    """Send WhatsApp notification using DB-stored templates with optional attachment."""
    # Try DB template first
    db_template = await db.whatsapp_templates.find_one({"key": notification_type}, {"_id": 0})
    if db_template and db_template.get("enabled", True):
        template_text = db_template.get("message", "")
        attachment_url = db_template.get("attachment_url", "")
        attachment_type = db_template.get("attachment_type", "")
    else:
        # Fallback to hardcoded defaults
        defaults = {d["key"]: d["message"] for d in WA_DEFAULT_TEMPLATES}
        template_text = defaults.get(notification_type, "")
        attachment_url = ""
        attachment_type = ""
    if not template_text:
        return False
    try:
        message = template_text.format(**variables)
    except KeyError:
        message = template_text
        for k, v in variables.items():
            message = message.replace("{" + k + "}", str(v))
    # Send with attachment if configured
    if attachment_url and attachment_type == "image":
        return await send_whatsapp_image(phone, attachment_url, message)
    elif attachment_url and attachment_type == "document":
        return await send_whatsapp_document(phone, attachment_url, message)
    return await send_whatsapp_message(phone, message)


# =================== WHATSAPP ADMIN ENDPOINTS ===================

@api_router.get("/admin/whatsapp-settings")
async def admin_get_whatsapp_settings(admin=Depends(get_current_admin)):
    settings = await get_whatsapp_settings()
    if not settings:
        settings = {"server_url": "", "api_key": "", "session_id": "primary", "enabled": False, "auto_notifications": {}}
    token = settings.get("api_key", "")
    settings["api_key_masked"] = ("*" * 8 + token[-5:]) if len(token) > 5 else token
    settings["api_key"] = token
    return settings


@api_router.put("/admin/whatsapp-settings")
async def admin_update_whatsapp_settings(data: dict, admin=Depends(get_current_admin)):
    update = {
        "enabled": data.get("enabled", False),
        "session_id": data.get("session_id", "primary"),
        "auto_notifications": data.get("auto_notifications", {
            "membership_submitted": True, "membership_approved": True,
            "membership_denied": True, "event_registered": True,
            "room_allotment": True, "payment_received": True,
            "event_invite": True, "event_certificate": True,
            "event_payment_confirmed": True, "certificate_issued": True,
        }),
        "updated_at": now_iso()
    }
    if data.get("api_key") and not data["api_key"].startswith("*"):
        update["api_key"] = data["api_key"]
    if data.get("server_url"):
        update["server_url"] = data["server_url"].rstrip("/")
    await db.whatsapp_settings.update_one({}, {"$set": update}, upsert=True)
    return {"message": "WhatsApp settings updated"}


@api_router.get("/admin/whatsapp/status")
async def admin_whatsapp_status(admin=Depends(get_current_admin)):
    settings = await get_whatsapp_settings()
    server_url = settings.get("server_url", "") or WHATSAPP_SERVER_URL
    sid = settings.get("session_id", "primary")
    token = settings.get("api_key", "")
    if not server_url or not token:
        return {"status": "not_configured", "session_id": sid}
    # WhatsApp server may not have a status endpoint — verify by checking server reachability
    url = f"{server_url.rstrip('/')}/api/v1/send/text"
    async with httpx.AsyncClient(timeout=10) as client:
        try:
            # OPTIONS/HEAD to check server is alive without sending a message
            resp = await client.options(url, headers={"Authorization": f"Bearer {token}"})
            if resp.status_code in (200, 204, 405):
                return {"status": "success", "session_id": sid, "message": f"Server reachable, session: {sid}"}
            if resp.status_code == 401:
                return {"status": "error", "session_id": sid, "message": "API key invalid"}
            return {"status": "error", "session_id": sid, "message": f"Server returned HTTP {resp.status_code}"}
        except Exception as e:
            return {"status": "error", "session_id": sid, "message": f"Cannot reach server: {str(e)[:100]}"}


@api_router.post("/admin/whatsapp/send-test")
async def admin_whatsapp_send_test(data: dict, admin=Depends(get_current_admin)):
    phone = data.get("phone", "")
    message = data.get("message", "This is a test message from IDSEA Admin Panel.")
    media_url = data.get("media_url", "")
    media_type = data.get("media_type", "")
    if not phone:
        raise HTTPException(status_code=400, detail="Phone number required")
    settings = await get_whatsapp_settings()
    sid = settings.get("session_id", "primary")

    if media_type == "image" and media_url:
        success = await send_whatsapp_image(phone, media_url, message, sid)
    elif media_type == "document" and media_url:
        success = await send_whatsapp_document(phone, media_url, message, sid)
    else:
        success = await send_whatsapp_message(phone, message, sid)

    if not success:
        return {"status": "error", "message": "Send failed"}
    return {"status": "success", "message": "Test message sent"}


@api_router.post("/admin/whatsapp/send-bulk")
async def admin_whatsapp_send_bulk(data: dict, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    message = data.get("message", "")
    target = data.get("target", "all_members")
    event_id = data.get("event_id", "")
    membership_filter = data.get("membership_type", "")
    if not message:
        raise HTTPException(status_code=400, detail="Message is required")

    async def _send_bulk():
        recipients = []
        if target == "event_registered" and event_id:
            regs = await db.event_registrations.find({"event_id": event_id}, {"_id": 0}).to_list(5000)
            recipients = [{"name": r.get("name", ""), "phone": r.get("phone", "")} for r in regs if r.get("phone")]
        else:
            query = {"status": "approved"}
            if membership_filter and membership_filter != "all":
                query["membership_type"] = membership_filter
            members = await db.members.find(query, {"_id": 0}).to_list(10000)
            recipients = [{"name": m.get("name", ""), "phone": m.get("phone", "")} for m in members if m.get("phone")]
        sent, failed = 0, 0
        for r in recipients:
            personalized = message.replace("{name}", r["name"])
            success = await send_whatsapp_message(r["phone"], personalized)
            if success:
                sent += 1
            else:
                failed += 1
            await asyncio.sleep(1)
        await db.whatsapp_logs.insert_one({
            "id": str(uuid.uuid4()), "type": "bulk", "target": target,
            "event_id": event_id, "message": message[:200],
            "total_recipients": len(recipients), "sent": sent, "failed": failed, "sent_at": now_iso()
        })

    background_tasks.add_task(_send_bulk)
    return {"message": "Bulk message sending started in background"}


# ============ WHATSAPP MARKETING CAMPAIGN ============

def _gen_ref_code():
    import string, random
    chars = string.ascii_uppercase + string.digits
    return "IDSEA-" + "".join(random.choices(chars, k=6))


@api_router.post("/admin/whatsapp/campaign")
async def create_whatsapp_campaign(data: dict, background_tasks: BackgroundTasks, admin=Depends(get_current_admin)):
    message = data.get("message", "")
    if not message:
        raise HTTPException(400, "Message is required")

    campaign_id = str(uuid.uuid4())
    campaign = {
        "id": campaign_id,
        "name": data.get("name", "Untitled Campaign"),
        "message": message,
        "batch_size": max(1, min(50, int(data.get("batch_size", 10)))),
        "interval_seconds": max(5, int(data.get("interval_seconds", 30))),
        "target": data.get("target", "all_members"),
        "membership_type": data.get("membership_type", "all"),
        "event_id": data.get("event_id", ""),
        "custom_phones": data.get("custom_phones", []),
        "media_url": data.get("media_url", ""),
        "media_type": data.get("media_type", ""),
        "add_reference_code": data.get("add_reference_code", True),
        "status": "running",
        "total": 0, "sent": 0, "failed": 0,
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.whatsapp_campaigns.insert_one(campaign)
    campaign.pop("_id", None)

    async def _run_campaign():
        try:
            recipients = []
            target = campaign["target"]
            if target == "event_registered" and campaign["event_id"]:
                regs = await db.event_registrations.find({"event_id": campaign["event_id"]}, {"_id": 0}).to_list(5000)
                recipients = [{"name": r.get("name", ""), "phone": r.get("phone", "")} for r in regs if r.get("phone")]
            elif target == "custom":
                recipients = [p for p in campaign["custom_phones"] if p.get("phone")]
            else:
                query = {"status": "approved"}
                mf = campaign["membership_type"]
                if mf and mf != "all":
                    query["membership_type"] = mf
                members = await db.members.find(query, {"_id": 0}).to_list(10000)
                recipients = [{"name": m.get("name", ""), "phone": m.get("phone", "")} for m in members if m.get("phone")]

            total = len(recipients)
            await db.whatsapp_campaigns.update_one({"id": campaign_id}, {"$set": {"total": total, "updated_at": now_iso()}})

            batch_size = campaign["batch_size"]
            interval = campaign["interval_seconds"]
            media_url = campaign["media_url"]
            media_type = campaign["media_type"]
            add_ref = campaign["add_reference_code"]
            sent, failed = 0, 0

            for batch_start in range(0, total, batch_size):
                batch = recipients[batch_start:batch_start + batch_size]
                for r in batch:
                    ref = _gen_ref_code() if add_ref else ""
                    name = r.get("name", "Member")
                    personalized = message.replace("{name}", name)
                    if ref:
                        personalized += f"\n\nRef: {ref}"

                    success = False
                    if media_type == "image" and media_url:
                        success = await send_whatsapp_image(r["phone"], media_url, personalized)
                    elif media_type == "document" and media_url:
                        success = await send_whatsapp_document(r["phone"], media_url, personalized)
                    else:
                        success = await send_whatsapp_message(r["phone"], personalized)

                    if success:
                        sent += 1
                    else:
                        failed += 1

                    await asyncio.sleep(0.5)

                await db.whatsapp_campaigns.update_one(
                    {"id": campaign_id},
                    {"$set": {"sent": sent, "failed": failed, "updated_at": now_iso()}}
                )

                if batch_start + batch_size < total:
                    await asyncio.sleep(interval)

            await db.whatsapp_campaigns.update_one(
                {"id": campaign_id},
                {"$set": {"status": "completed", "sent": sent, "failed": failed, "updated_at": now_iso()}}
            )
        except Exception as e:
            logging.error(f"Campaign {campaign_id} error: {e}")
            await db.whatsapp_campaigns.update_one(
                {"id": campaign_id},
                {"$set": {"status": "failed", "updated_at": now_iso()}}
            )

    background_tasks.add_task(_run_campaign)
    return {"message": "Campaign started", "campaign_id": campaign_id, "campaign": campaign}


@api_router.get("/admin/whatsapp/campaigns")
async def list_whatsapp_campaigns(admin=Depends(get_current_admin)):
    return await db.whatsapp_campaigns.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)


@api_router.get("/admin/whatsapp/campaigns/{campaign_id}")
async def get_whatsapp_campaign(campaign_id: str, admin=Depends(get_current_admin)):
    c = await db.whatsapp_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not c:
        raise HTTPException(404, "Campaign not found")
    return c


@api_router.get("/admin/whatsapp/logs")
async def admin_whatsapp_logs(admin=Depends(get_current_admin), limit: int = 100):
    logs = await db.whatsapp_logs.find({}, {"_id": 0}).sort("sent_at", -1).to_list(limit)
    return logs


@api_router.post("/whatsapp/webhook")
async def whatsapp_webhook(data: dict):
    logging.info(f"WhatsApp webhook received: {str(data)[:500]}")
    await db.whatsapp_webhooks.insert_one({
        "id": str(uuid.uuid4()), "data": data, "received_at": now_iso()
    })
    return {"status": "ok"}



# =================== WHATSAPP TEMPLATES CRUD ===================

WA_DEFAULT_TEMPLATES = [
    {
        "key": "membership_submitted", "name": "Membership Submitted",
        "message": "Hello {name},\n\nThank you for applying for IDSEA {membership_type} membership.\n\nYour application has been received and is under review. We will notify you once it is processed.\n\nRegards,\nIDSEA Team",
        "variables": ["name", "membership_type", "email", "phone"],
        "description": "Sent when a new membership application is received",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "membership_approved", "name": "Membership Approved",
        "message": "Congratulations {name}!\n\nYour IDSEA membership has been approved.\n\nMembership ID: {membership_id}\nType: {membership_type}\n\nWelcome to the IDSEA family!\n\nRegards,\nIDSEA Team",
        "variables": ["name", "membership_id", "membership_type"],
        "description": "Sent when membership is approved (certificate auto-attached)",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "membership_denied", "name": "Membership Denied",
        "message": "Dear {name},\n\nWe regret to inform you that your IDSEA membership application has been denied.\n\nPlease contact us at info@idsea.org for more details.\n\nRegards,\nIDSEA Team",
        "variables": ["name", "membership_type"],
        "description": "Sent when membership is denied by admin",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "event_registered", "name": "Event Registration Confirmed",
        "message": "Hello {name},\n\nYou have been successfully registered for:\n\n*{event_title}*\nDate: {event_date}\nVenue: {event_venue}\n{venue_map_link}\nTotal Amount: Rs. {total_amount}\n\nPayment Status: {payment_status}\n\nRegards,\nIDSEA Team",
        "variables": ["name", "event_title", "event_date", "event_venue", "venue_map_link", "total_amount", "payment_status"],
        "description": "Sent after successful event registration",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "room_allotment", "name": "Room Allotment",
        "message": "Hello {name},\n\nYour accommodation details for *{event_title}*:\n\nRoom No: {room_no}\nLocation: {location}\n{map_link}\n\nRegards,\nIDSEA Team",
        "variables": ["name", "event_title", "room_no", "location", "map_link"],
        "description": "Sent when room is allotted for an event",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "payment_received", "name": "Payment Received",
        "message": "Hello {name},\n\nWe have received your payment of Rs. {amount}.\n\nTransaction Reference: {reference}\n\nThank you!\nIDSEA Team",
        "variables": ["name", "amount", "reference"],
        "description": "Sent when membership payment is confirmed",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "payment_reminder", "name": "Payment Reminder",
        "message": "Dear {name},\n\nThis is a friendly reminder that your IDSEA {membership_type} membership payment of Rs. {amount} is pending.\n\nPlease complete your payment at your earliest convenience.\n\nRegards,\nIDSEA Team",
        "variables": ["name", "membership_type", "amount", "membership_id"],
        "description": "Sent as a reminder for pending membership payments",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "event_invite", "name": "Event Invitation",
        "message": "Dear {name},\n\nWe are pleased to invite you to:\n\n*{event_title}*\nDate: {event_date}\nVenue: {event_venue}\n\n{event_description}\n\nRegister now to secure your spot!\n\nRegards,\nIDSEA Team",
        "variables": ["name", "event_title", "event_date", "event_venue", "event_description"],
        "description": "Sent when admin broadcasts a new event to members",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "event_certificate", "name": "Event Participation Certificate",
        "message": "Dear {name},\n\nThank you for participating in *{event_title}* held on {event_date} at {event_venue}.\n\nPlease find your participation certificate attached.\n\nWe look forward to seeing you at future IDSEA events!\n\nRegards,\nIDSEA Team",
        "variables": ["name", "event_title", "event_date", "event_venue"],
        "description": "Sent with participation certificate when event is closed",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "membership_renewal", "name": "Membership Renewal Reminder",
        "message": "Dear {name},\n\nYour IDSEA {membership_type} membership (ID: {membership_id}) is due for renewal.\n\nPlease renew your membership to continue enjoying all IDSEA benefits.\n\nRegards,\nIDSEA Team",
        "variables": ["name", "membership_type", "membership_id"],
        "description": "Sent as a reminder before membership expiry",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "event_payment_confirmed", "name": "Event Payment Confirmed",
        "message": "Hello {name},\n\nYour payment of Rs. {amount} for *{event_title}* has been confirmed.\n\nTransaction ID: {transaction_id}\nRegistration ID: {registration_id}\n\nSee you at the event!\n\nRegards,\nIDSEA Team",
        "variables": ["name", "event_title", "amount", "transaction_id", "registration_id"],
        "description": "Sent when event registration payment is verified",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "certificate_issued", "name": "Certificate Issued",
        "message": "Dear {name},\n\nYour {certificate_type} certificate has been issued by IDSEA.\n\nCertificate ID: {certificate_id}\n\nPlease find the certificate attached.\n\nRegards,\nIDSEA Team",
        "variables": ["name", "certificate_type", "certificate_id"],
        "description": "Sent when any certificate is manually generated for a member",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
    {
        "key": "custom_message", "name": "Custom Admin Message",
        "message": "Dear {name},\n\n{message}\n\nRegards,\nIDSEA Team",
        "variables": ["name", "message"],
        "description": "Sent when admin sends a direct message to a member",
        "attachment_url": "", "attachment_type": "", "enabled": True,
    },
]


@api_router.get("/admin/whatsapp-templates")
async def admin_get_whatsapp_templates(admin=Depends(get_current_admin)):
    templates = await db.whatsapp_templates.find({}, {"_id": 0}).to_list(50)
    # Merge defaults with DB entries
    db_keys = {t["key"] for t in templates}
    result = list(templates)
    for d in WA_DEFAULT_TEMPLATES:
        if d["key"] not in db_keys:
            result.append(d)
    result.sort(key=lambda t: next((i for i, d in enumerate(WA_DEFAULT_TEMPLATES) if d["key"] == t["key"]), 99))
    return result


@api_router.get("/admin/whatsapp-templates/{template_key}")
async def admin_get_whatsapp_template(template_key: str, admin=Depends(get_current_admin)):
    template = await db.whatsapp_templates.find_one({"key": template_key}, {"_id": 0})
    if not template:
        default = next((d for d in WA_DEFAULT_TEMPLATES if d["key"] == template_key), None)
        if default:
            return default
        raise HTTPException(404, "Template not found")
    return template


@api_router.put("/admin/whatsapp-templates/{template_key}")
async def admin_update_whatsapp_template(template_key: str, data: dict, admin=Depends(get_current_admin)):
    update = {
        "key": template_key,
        "name": data.get("name", template_key),
        "message": data.get("message", ""),
        "description": data.get("description", ""),
        "attachment_url": data.get("attachment_url", ""),
        "attachment_type": data.get("attachment_type", ""),
        "enabled": data.get("enabled", True),
        "updated_at": now_iso(),
    }
    # Preserve variables from default if not provided
    default = next((d for d in WA_DEFAULT_TEMPLATES if d["key"] == template_key), None)
    update["variables"] = data.get("variables", default["variables"] if default else [])
    await db.whatsapp_templates.update_one({"key": template_key}, {"$set": update}, upsert=True)
    return {"message": f"WhatsApp template '{template_key}' updated"}


@api_router.post("/admin/whatsapp-templates/{template_key}/reset")
async def admin_reset_whatsapp_template(template_key: str, admin=Depends(get_current_admin)):
    await db.whatsapp_templates.delete_one({"key": template_key})
    return {"message": f"Template '{template_key}' reset to default"}


@api_router.post("/admin/whatsapp-templates/{template_key}/preview")
async def admin_preview_whatsapp_template(template_key: str, admin=Depends(get_current_admin)):
    template = await db.whatsapp_templates.find_one({"key": template_key}, {"_id": 0})
    if not template:
        default = next((d for d in WA_DEFAULT_TEMPLATES if d["key"] == template_key), None)
        if not default:
            raise HTTPException(404, "Template not found")
        template = default
    sample_vars = {
        "name": "Dr. Sample Person", "membership_type": "Academic",
        "membership_id": "ACD/IDSEA/0001", "email": "sample@idsea.org",
        "phone": "919876543210", "event_title": "IDSEA Conference 2026",
        "event_date": "15 March 2026", "event_venue": "VCRI Namakkal",
        "venue_map_link": "https://maps.google.com", "total_amount": "3,100",
        "payment_status": "Paid", "room_no": "A-101", "location": "Guest House Block A",
        "map_link": "https://maps.google.com", "amount": "3,100", "reference": "TXN-123456",
        "event_description": "Join us for the annual IDSEA conference featuring keynote speakers and workshops.",
        "transaction_id": "PAY-RZP-ABC123", "registration_id": "REG-2026-0001",
        "certificate_type": "Membership", "certificate_id": "IDSEA-MEM-A3K7X9",
        "message": "This is a custom message from the IDSEA admin team.",
    }
    message = template.get("message", "")
    try:
        preview = message.format(**sample_vars)
    except KeyError:
        preview = message
        for k, v in sample_vars.items():
            preview = preview.replace("{" + k + "}", v)
    return {"preview": preview, "attachment_url": template.get("attachment_url", ""), "attachment_type": template.get("attachment_type", "")}




# =================== BACKUP, RESTORE & FACTORY RESET ===================

import zipfile
import shutil
import tempfile
import json as json_lib
from bson import json_util


@api_router.get("/admin/backup/database")
async def admin_backup_database(admin=Depends(get_current_admin)):
    """Download full MongoDB database as JSON ZIP"""
    collections = ["admins", "members", "events", "event_details", "event_registrations",
                    "executive_committee", "cms_settings", "page_contents", "news", "gallery_albums",
                    "publications", "payments", "payment_settings", "email_templates", "email_logs",
                    "email_queue", "smtp_settings", "certificate_records", "certificate_templates",
                    "whatsapp_settings", "whatsapp_campaigns", "slider_settings", "membership_plans"]

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    with zipfile.ZipFile(tmp.name, 'w', zipfile.ZIP_DEFLATED) as zf:
        for coll_name in collections:
            try:
                docs = await db[coll_name].find({}).to_list(100000)
                json_str = json_util.dumps(docs, indent=2)
                zf.writestr(f"{coll_name}.json", json_str)
            except Exception:
                pass
        zf.writestr("_backup_info.json", json_lib.dumps({
            "backup_date": now_iso(),
            "backup_by": admin["email"],
            "db_name": db.name,
            "collections": collections
        }, indent=2))

    return FileResponse(tmp.name, media_type="application/zip",
                        filename=f"idsea_db_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                        background=BackgroundTask(lambda: os.remove(tmp.name)))


@api_router.get("/admin/backup/uploads")
async def admin_backup_uploads(admin=Depends(get_current_admin)):
    """Download all uploaded files as ZIP"""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    with zipfile.ZipFile(tmp.name, 'w', zipfile.ZIP_DEFLATED) as zf:
        upload_path = UPLOAD_DIR
        if upload_path.exists():
            for file_path in upload_path.rglob("*"):
                if file_path.is_file():
                    arcname = str(file_path.relative_to(upload_path))
                    zf.write(str(file_path), arcname)
        zf.writestr("_backup_info.json", json_lib.dumps({
            "backup_date": now_iso(),
            "backup_by": admin["email"],
            "uploads_path": str(upload_path),
            "file_count": sum(1 for f in upload_path.rglob("*") if f.is_file()) if upload_path.exists() else 0
        }, indent=2))

    return FileResponse(tmp.name, media_type="application/zip",
                        filename=f"idsea_uploads_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                        background=BackgroundTask(lambda: os.remove(tmp.name)))


@api_router.get("/admin/backup/info")
async def admin_backup_info(admin=Depends(get_current_admin)):
    """Get backup info — file counts, DB stats, restore paths"""
    collections_info = {}
    for coll_name in ["admins", "members", "events", "event_details", "event_registrations",
                       "executive_committee", "cms_settings", "page_contents", "news", "gallery_albums",
                       "publications", "payments", "email_templates", "email_logs", "certificate_records"]:
        try:
            count = await db[coll_name].count_documents({})
            collections_info[coll_name] = count
        except Exception:
            collections_info[coll_name] = 0

    upload_count = sum(1 for f in UPLOAD_DIR.rglob("*") if f.is_file()) if UPLOAD_DIR.exists() else 0
    upload_size_mb = round(sum(f.stat().st_size for f in UPLOAD_DIR.rglob("*") if f.is_file()) / (1024*1024), 2) if UPLOAD_DIR.exists() else 0

    return {
        "database": {
            "name": db.name,
            "collections": collections_info,
            "total_documents": sum(collections_info.values()),
        },
        "uploads": {
            "path": str(UPLOAD_DIR),
            "restore_path": "/var/www/idsea/backend/uploads/",
            "file_count": upload_count,
            "size_mb": upload_size_mb,
        },
        "restore_instructions": {
            "database": "Upload the database backup ZIP. Each JSON file inside will replace the corresponding MongoDB collection.",
            "uploads": "Upload the uploads backup ZIP. Files will be extracted to the uploads directory, restoring all images and documents.",
        }
    }


@api_router.post("/admin/restore/database")
async def admin_restore_database(file: UploadFile = File(...), admin=Depends(get_current_admin)):
    """Restore MongoDB database from backup ZIP"""
    if not file.filename.endswith('.zip'):
        raise HTTPException(status_code=400, detail="Only ZIP files accepted")

    tmp_path = f"/tmp/idsea_restore_{uuid.uuid4().hex}.zip"
    with open(tmp_path, 'wb') as f:
        content = await file.read()
        f.write(content)

    restored = []
    try:
        with zipfile.ZipFile(tmp_path, 'r') as zf:
            for name in zf.namelist():
                if name.startswith('_') or not name.endswith('.json'):
                    continue
                coll_name = name.replace('.json', '')
                json_str = zf.read(name).decode('utf-8')
                docs = json_util.loads(json_str)
                if docs:
                    await db[coll_name].delete_many({})
                    await db[coll_name].insert_many(docs)
                    restored.append(f"{coll_name}: {len(docs)} documents")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")
    finally:
        os.remove(tmp_path)

    return {"message": "Database restored successfully", "restored": restored}


@api_router.post("/admin/restore/uploads")
async def admin_restore_uploads(file: UploadFile = File(...), admin=Depends(get_current_admin)):
    """Restore uploads folder from backup ZIP"""
    if not file.filename.endswith('.zip'):
        raise HTTPException(status_code=400, detail="Only ZIP files accepted")

    tmp_path = f"/tmp/idsea_uploads_restore_{uuid.uuid4().hex}.zip"
    with open(tmp_path, 'wb') as f:
        content = await file.read()
        f.write(content)

    file_count = 0
    try:
        with zipfile.ZipFile(tmp_path, 'r') as zf:
            for name in zf.namelist():
                if name == '_backup_info.json':
                    continue
                target = UPLOAD_DIR / name
                # Prevent zip-slip attack
                if not str(target.resolve()).startswith(str(UPLOAD_DIR.resolve())):
                    continue
                target.parent.mkdir(parents=True, exist_ok=True)
                with zf.open(name) as src, open(target, 'wb') as dst:
                    dst.write(src.read())
                file_count += 1
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")
    finally:
        os.remove(tmp_path)

    return {"message": f"Uploads restored: {file_count} files", "count": file_count, "path": str(UPLOAD_DIR)}


@api_router.post("/admin/factory-reset")
async def admin_factory_reset(data: dict, admin=Depends(get_current_admin)):
    """Factory reset — delete ALL data. Requires confirmation code."""
    if data.get("confirm") != "DELETE_ALL_DATA":
        raise HTTPException(status_code=400, detail="Send {\"confirm\": \"DELETE_ALL_DATA\"} to confirm")

    keep_admin = data.get("keep_admin", True)
    collections_cleared = []

    all_collections = ["members", "events", "event_details", "event_registrations",
                       "executive_committee", "cms_settings", "page_contents", "news", "gallery_albums",
                       "publications", "payments", "payment_settings", "email_templates", "email_logs",
                       "email_queue", "smtp_settings", "certificate_records", "certificate_templates",
                       "whatsapp_settings", "whatsapp_campaigns", "slider_settings", "membership_plans"]

    for coll_name in all_collections:
        result = await db[coll_name].delete_many({})
        collections_cleared.append(f"{coll_name}: {result.deleted_count}")

    if not keep_admin:
        await db.admins.delete_many({})
        collections_cleared.append("admins: all deleted")

    # Clear uploads folder
    if data.get("clear_uploads", False):
        if UPLOAD_DIR.exists():
            for f in UPLOAD_DIR.iterdir():
                if f.is_file() and f.name != 'idsea_logo.png':
                    f.unlink()
            collections_cleared.append("uploads: files cleared (kept logo)")

    return {"message": "Factory reset complete. Restart backend to re-seed defaults.", "cleared": collections_cleared}


# =================== APP SETUP ===================

app.include_router(api_router)

# Serve uploaded files (images, PDFs, etc.)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def startup_event():
    # Start email queue processor
    asyncio.create_task(process_email_queue())
    logging.info("Email queue scheduler started (batch=50, interval=5min)")

    if not await db.admins.find_one({"email": "admin@idsea.org"}):
        await db.admins.insert_one(AdminUser(
            username="Super Admin",
            email="admin@idsea.org",
            password_hash=hash_password("Admin@123"),
            role="super_admin"
        ).model_dump())
        logger.info("Default admin created: admin@idsea.org / Admin@123")

    # Seed default CMS page contents (needed for site to render)
    default_pages = {
        "home": {
            "about_title": "About IDSEA",
            "about_subtitle": "A Platform for Dairy Science & Entrepreneurship",
            "about_description": "IDSEA is a national professional and scientific body that bridges the gap between dairy scientists and industry professionals. We bring together academicians, technologists, entrepreneurs, and students under one umbrella.",
            "about_description2": "Headquartered at VCRI, Namakkal, Tamil Nadu, we operate with an all-India mandate to foster innovation, research, and sustainable growth.",
            "about_image": "https://images.unsplash.com/photo-1532094349884-543559059a6d?w=600&q=80",
            "membership_title": "Membership Types",
            "membership_subtitle": "Join IDSEA and be part of India's premier dairy science community",
            "events_title": "Upcoming Events",
            "events_subtitle": "Conferences, workshops, and seminars",
            "news_title": "Latest News",
            "news_subtitle": "Scientific updates and announcements",
            "cta_title": "Join the IDSEA Community Today",
            "cta_description": "Become part of India's premier dairy science and entrepreneurship network. Connect, collaborate, and grow.",
            "cta_button_text": "Apply for Membership",
            "cta_button_link": "/apply",
            "seo_title": "IDSEA - Indian Dairy Scientists and Entrepreneurs Association",
            "seo_description": "IDSEA is India's premier platform for dairy scientists, academicians, technologists and entrepreneurs. Join us for conferences, research and networking.",
            "seo_keywords": "IDSEA, dairy science, dairy entrepreneurs, India dairy, veterinary science, dairy technology, dairy research, membership",
        },
        "about": {
            "hero_title": "About IDSEA",
            "hero_subtitle": "Indian Dairy Scientists and Entrepreneurs Association - bridging dairy science with industry since our founding.",
            "objectives": "Promote advancement and dissemination of knowledge in dairy science and entrepreneurship\nProvide a common national platform for dairy scientists, academicians, and entrepreneurs\nSupport dairy startups, MSMEs, cooperatives, FPOs through knowledge sharing\nOrganize conferences, seminars, workshops, training programs, and exhibitions\nFacilitate academia-industry-startup collaboration for innovation and technology transfer\nPublish journals, proceedings, newsletters, reports, and technical bulletins\nCollaborate with universities, research institutes, and international organizations\nRecognize excellence through awards, fellowships, and professional recognitions",
            "hq_title": "Headquarters",
            "council_title": "IDSEA Executive Council",
            "council_subtitle": "Term: 3 years",
            "founders_title": "Patron / Founders",
            "founders_subtitle": "The visionaries who established IDSEA",
            "cert_title": "Registration Certificate",
            "cert_subtitle": "Officially registered under the Tamil Nadu Societies Registration Act",
            "cert_org_name": "Indian Dairy Scientists and Entrepreneurs Association (IDSEA)",
            "cert_reg_number": "",
            "cert_act": "Tamil Nadu Societies Registration Act, 1975 (Tamil Nadu Act 27 of 1975)",
            "cert_image_url": "",
        },
        "events": {
            "hero_title": "Events & Conferences",
            "hero_subtitle": "Dairy science conferences, workshops, and seminars",
        },
        "gallery": {
            "hero_title": "Photo Gallery",
            "hero_subtitle": "Conferences, field visits, workshops, and research events",
        },
        "publications": {
            "hero_title": "Publications & Research",
            "hero_subtitle": "Research papers, journals, and technical articles by our members",
        },
        "members": {
            "hero_title": "Member Directory",
            "hero_subtitle": "Search and discover IDSEA members across India",
        },
        "contact": {
            "hero_title": "Contact Us",
            "hero_subtitle": "Get in touch with IDSEA",
            "form_title": "Send Message",
            "form_success_title": "Message Sent!",
            "form_success_message": "Thank you for contacting IDSEA. We'll get back to you soon.",
            "membership_cta_title": "Become a Member",
            "membership_cta_description": "Join IDSEA to connect with dairy scientists and entrepreneurs across India.",
        },
        "navbar": {
            "org_name": "Indian Dairy Scientists and Entrepreneurs Association",
            "org_short": "(IDSEA)",
            "registration_number": "",
        },
        "footer": {
            "description": "Bridging dairy science, innovation & entrepreneurship for the sustainable growth of the Indian dairy sector.",
            "copyright_text": "Indian Dairy Scientists and Entrepreneurs Association (IDSEA). All rights reserved.",
        },
    }
    for page_key, content in default_pages.items():
        if await db.page_contents.count_documents({"page": page_key}) == 0:
            await db.page_contents.insert_one({"page": page_key, "content": content, "updated_at": now_iso()})
    logger.info("Page contents seeded")


@app.on_event("shutdown")
async def shutdown_db_client():
    db_client.close()
